from flask import Blueprint, render_template, request, redirect, url_for, session, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from .database import get_db
from functools import wraps
from io import BytesIO
import csv
import json
import urllib.request
import urllib.error
from psycopg2.extras import execute_values
import re
import openpyxl
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import os
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, datetime, timedelta

auth = Blueprint('auth', __name__)


# ================================================================
# TIMEOUT DE INACTIVIDAD — renueva la sesión en cada request
# ================================================================

@auth.before_request
def renovar_sesion():
    session.modified = True
    session.permanent = True
    # Resolución automática de promociones provisorias vencidas.
    # Se dispara con la actividad de cualquier usuario logueado.
    _chequear_promociones_vencidas()


# ================================================================
# HELPERS
# ================================================================

def login_requerido(roles_permitidos):
    def decorador(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if 'rol' not in session:
                return redirect(url_for('auth.login'))
            if session['rol'] not in roles_permitidos:
                return redirect(url_for('auth.dashboard'))
            return f(*args, **kwargs)
        return wrapper
    return decorador


# ================================================================
# PROMOCIÓN PROVISORIA POR CORRELATIVA ADEUDADA
#
# Regla institucional: el alumno que promociona una materia y adeuda
# el final de su correlativa del año anterior conserva la promoción
# sólo hasta la fecha límite del ciclo lectivo (por defecto el 31/12).
# Tiene las mesas de ese mismo año para aprobar el final. Si llega la
# fecha límite sin aprobarlo, la promoción se cae y la materia queda
# como 'regular': pasa a deber final, porque los profesores no guardan
# la promoción de un año para el otro.
#
# Mientras tanto la cursada queda marcada con promocion_provisoria.
# ================================================================

# Marcas que el sistema escribe solo en observaciones. Se limpian en cada
# guardado antes de volver a evaluar la regla, para que no se acumulen ni
# queden colgadas cuando la condicion cambia.
_MARCAS_AUTOMATICAS = ('PROMOCIÓN PROVISORIA', 'PROMOCIÓN CAÍDA', 'PROMOCIÓN CONDICIONADA')


def _limpiar_marcas_automaticas(obs):
    """
    Devuelve las observaciones sin las marcas que escribe el sistema,
    conservando intacto lo que haya escrito la preceptora.
    """
    if not obs:
        return ''
    partes = [p.strip() for p in str(obs).split('|')]
    quedan = [p for p in partes
              if p and not any(p.startswith(m) for m in _MARCAS_AUTOMATICAS)]
    return ' | '.join(quedan)


def _fecha_limite_promocion(cur, anio):
    """
    Devuelve la fecha límite del ciclo lectivo indicado, leída del
    parámetro 'fecha_limite_promocion' (formato DD-MM).
    Si el parámetro falta o está mal cargado, cae al 31 de diciembre.
    """
    try:
        cur.execute("SELECT valor FROM configuracion WHERE clave = 'fecha_limite_promocion'")
        row = cur.fetchone()
        dia, mes = [int(x) for x in (row[0] if row and row[0] else '31-12').strip().split('-')]
        return date(anio, mes, dia)
    except Exception:
        return date(anio, 12, 31)


def resolver_promociones_provisorias(conn):
    """
    Baja a 'regular' las promociones provisorias cuyo plazo ya venció.

    Cada cursada se evalúa contra la fecha límite de SU PROPIO ciclo
    lectivo (inscripciones.anio_lectivo), no contra el ciclo configurado
    como vigente. Esto la hace inmune a dos situaciones:

      • Servidor apagado durante meses: al prender, resuelve todo lo
        que quedó pendiente de ciclos anteriores.
      • Cambio de anio_lectivo_actual antes de la resolución: si el
        admin pasa el sistema a 2027 en febrero, las promociones
        provisorias de 2026 se resuelven igual contra el 31/12/2026.

    Es idempotente: una vez resuelta, la cursada deja de estar marcada,
    así que correrla de más no hace nada.

    NO toca cursadas cerradas: por decisión institucional una cursada
    cerrada es inmodificable. Por eso el cierre de notas está bloqueado
    hasta pasada la fecha límite cuando hay promociones provisorias.

    Cada baja queda registrada en cursadas_auditoria con
    modificado_por_id NULL, que identifica al sistema como autor.

    Devuelve la cantidad de promociones dadas de baja.
    """
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT cu.id, cu.condicion, cu.observaciones, i.anio_lectivo
            FROM cursadas cu
            JOIN inscripciones i ON i.id = cu.inscripcion_id
            WHERE cu.promocion_provisoria
              AND NOT cu.cerrada
        """)
        pendientes = cur.fetchall()
        if not pendientes:
            return 0

        hoy = date.today()
        motivo = ('Baja automática del sistema: venció el plazo para aprobar '
                  'el final de la correlativa y la promoción no se confirmó.')
        bajadas = 0

        for cursada_id, condicion_ant, obs_ant, anio_lectivo in pendientes:
            limite = _fecha_limite_promocion(cur, anio_lectivo)
            if hoy <= limite:
                continue  # todavía está en plazo

            nota_obs = (f'PROMOCIÓN CAÍDA — venció el plazo del '
                        f'{limite.strftime("%d/%m/%Y")} sin aprobar el final '
                        f'de la correlativa. Se asienta como regular.')
            obs_nueva = (obs_ant + ' | ' + nota_obs).strip(' |') if obs_ant else nota_obs

            cur.execute("""
                UPDATE cursadas
                SET condicion = 'regular',
                    promocion_provisoria = FALSE,
                    observaciones = %s
                WHERE id = %s
            """, (obs_nueva, cursada_id))

            cur.execute("""
                INSERT INTO cursadas_auditoria
                    (cursada_id, campo, valor_anterior, valor_nuevo,
                     modificado_por_id, motivo)
                VALUES (%s, 'condicion', %s, 'regular', NULL, %s)
            """, (cursada_id, condicion_ant, motivo))
            bajadas += 1

        conn.commit()
        return bajadas
    except Exception:
        conn.rollback()
        return 0
    finally:
        cur.close()


# Evita golpear la base en cada request: se chequea una vez por día
# por worker. El parámetro en la base evita que dos workers procesen
# el mismo ciclo dos veces.
_ULTIMO_CHEQUEO_PROMOCIONES = None


def _chequear_promociones_vencidas():
    global _ULTIMO_CHEQUEO_PROMOCIONES
    if 'rol' not in session:
        return
    hoy = date.today()
    if _ULTIMO_CHEQUEO_PROMOCIONES == hoy:
        return
    _ULTIMO_CHEQUEO_PROMOCIONES = hoy
    conn = None
    try:
        conn = get_db()
        resolver_promociones_provisorias(conn)
    except Exception:
        # Nunca romper la navegación por esto.
        pass
    finally:
        if conn:
            conn.close()


def limpiar_dni(dni):
    """Elimina puntos, espacios y guiones del DNI antes de guardar o buscar en la DB."""
    return str(dni).replace('.', '').replace(' ', '').replace('-', '').strip()


def formatear_dni(dni):
    """Agrega puntos al DNI para mostrar en pantalla. Ej: 12345678 → 12.345.678"""
    try:
        return f"{int(str(dni).replace('.', '').replace(' ', '')):,}".replace(',', '.')
    except:
        return str(dni)


def formatear_documento(tipo, numero):
    """
    Devuelve el documento formateado para mostrar según su tipo:
      - DNI / DNI_EXT → con puntos (Ej: 12.345.678)
      - PAS / CI      → tal cual, en mayúsculas
    """
    if not numero:
        return ''
    t = (tipo or 'DNI').upper()
    if t in ('DNI', 'DNI_EXT'):
        return formatear_dni(numero)
    return str(numero).upper()


def etiqueta_documento(tipo):
    """Etiqueta corta para mostrar el tipo de documento (para PDFs, tablas, etc.)"""
    return {
        'DNI':     'DNI',
        'DNI_EXT': 'DNI',     # se muestra igual que DNI argentino (es DNI argentino emitido a extranjero)
        'PAS':     'Pasaporte',
        'CI':      'CI',
    }.get((tipo or 'DNI').upper(), 'DNI')


def validar_documento(tipo, numero):
    """
    Valida el número de documento según su tipo.
    Devuelve None si es válido, o un mensaje de error.

    Reglas:
      - DNI / DNI_EXT: solo dígitos, 7 u 8 caracteres
      - PAS / CI:      alfanumérico (letras, dígitos, guiones), 5 a 20 caracteres
    """
    if not numero:
        return 'El número de documento es obligatorio'

    n = str(numero).strip().upper().replace('.', '').replace(' ', '')
    t = (tipo or 'DNI').upper()

    if t not in ('DNI', 'DNI_EXT', 'PAS', 'CI'):
        return 'Tipo de documento inválido (debe ser DNI, DNI_EXT, PAS o CI)'

    if t in ('DNI', 'DNI_EXT'):
        if not n.isdigit():
            return 'El DNI solo debe contener números'
        if not (7 <= len(n) <= 8):
            return 'El DNI debe tener 7 u 8 dígitos'
    else:  # PAS / CI
        # Alfanumérico (letras, números, guiones)
        if not re.match(r'^[A-Z0-9\-]{5,20}$', n):
            return 'El pasaporte/CI debe tener entre 5 y 20 caracteres (letras, números o guiones)'

    return None  # válido


def limpiar_documento(numero):
    """Normaliza un número de documento para guardar en DB (sin puntos ni espacios, en mayúsculas)."""
    if not numero:
        return ''
    return str(numero).strip().upper().replace('.', '').replace(' ', '')


def validar_cuil(cuil):
    """
    Valida el CUIL/CUIT con el algoritmo oficial de AFIP (dígito verificador módulo 11).
    Es opcional: si está vacío, devuelve None (válido).

    Prefijos válidos para personas humanas (Res. AFIP/ANSES 5007/2021):
      20, 23, 24, 27
    """
    if not cuil:
        return None  # opcional → vacío es válido

    # Quitar guiones, puntos y espacios
    c = str(cuil).replace('-', '').replace('.', '').replace(' ', '').strip()

    if not c.isdigit():
        return 'El CUIL solo debe contener números'

    if len(c) != 11:
        return 'El CUIL debe tener 11 dígitos'

    if c[:2] not in ('20', '23', '24', '27'):
        return 'Prefijo de CUIL inválido (debe comenzar con 20, 23, 24 o 27)'

    # Algoritmo de dígito verificador (módulo 11)
    multiplicadores = [5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
    suma = sum(int(c[i]) * multiplicadores[i] for i in range(10))
    resto = suma % 11
    if resto == 0:
        dv = 0
    elif resto == 1:
        dv = 9  # caso especial (raro)
    else:
        dv = 11 - resto

    if int(c[10]) != dv:
        return 'El CUIL no es válido (dígito verificador incorrecto)'

    return None  # válido ✅


def limpiar_cuil(cuil):
    """Normaliza un CUIL para guardar en DB (solo dígitos, sin guiones)."""
    if not cuil:
        return None
    c = str(cuil).replace('-', '').replace('.', '').replace(' ', '').strip()
    return c if c else None


# Lista oficial de las 24 jurisdicciones argentinas (23 provincias + CABA)
PROVINCIAS_ARG = {
    'Buenos Aires', 'Ciudad Autónoma de Buenos Aires', 'Catamarca', 'Chaco',
    'Chubut', 'Córdoba', 'Corrientes', 'Entre Ríos', 'Formosa', 'Jujuy',
    'La Pampa', 'La Rioja', 'Mendoza', 'Misiones', 'Neuquén', 'Río Negro',
    'Salta', 'San Juan', 'San Luis', 'Santa Cruz', 'Santa Fe',
    'Santiago del Estero', 'Tierra del Fuego', 'Tucumán'
}


def validar_provincia(provincia):
    """Valida que la provincia sea una de las 24 jurisdicciones argentinas.
    Si está vacía, devuelve None (es opcional)."""
    if not provincia:
        return None
    if provincia.strip() not in PROVINCIAS_ARG:
        return f'Provincia inválida. Debe ser una de las 24 jurisdicciones argentinas.'
    return None


def validar_fecha_nacimiento(fecha_iso):
    """
    Valida la fecha de nacimiento del alumno:
      - Si está vacía → válido (es opcional)
      - Debe poder parsearse como AAAA-MM-DD
      - Edad mínima: 17 años cumplidos
      - Edad máxima: 100 años (defensivo contra typos)
      - No puede ser fecha futura

    Devuelve None si es válida, o mensaje de error si no.
    """
    if not fecha_iso:
        return None  # opcional

    try:
        f = date.fromisoformat(fecha_iso)
    except (ValueError, TypeError):
        return 'Fecha de nacimiento inválida (formato debe ser AAAA-MM-DD)'

    hoy = date.today()
    if f > hoy:
        return 'La fecha de nacimiento no puede ser futura'

    # Calcular edad
    edad = hoy.year - f.year
    if (hoy.month, hoy.day) < (f.month, f.day):
        edad -= 1

    # Helper local para singular/plural
    def _anios(n):
        return f'{n} año' if n == 1 else f'{n} años'

    if edad < 17:
        return f'El alumno tendría solo {_anios(edad)}. La edad mínima es 17 años.'

    if edad > 100:
        return f'El alumno tendría {_anios(edad)}. La edad máxima permitida es 100 años (verificá el año de nacimiento).'

    return None


def validar_password_fuerte(password):
    """
    Valida que la contraseña cumpla los requisitos mostrados en vivo
    en la pantalla de cambio de contraseña: mínimo 8 caracteres,
    al menos una mayúscula, una minúscula, un número y un carácter especial.
    Devuelve None si es válida, o un mensaje de error si no lo es.
    """
    if len(password) < 8:
        return 'La contraseña debe tener al menos 8 caracteres'
    if not re.search(r'[A-Z]', password):
        return 'La contraseña debe tener al menos una mayúscula'
    if not re.search(r'[a-z]', password):
        return 'La contraseña debe tener al menos una minúscula'
    if not re.search(r'[0-9]', password):
        return 'La contraseña debe tener al menos un número'
    if not re.search(r'[!@#$%^&*(),.?":{}|<>_\-]', password):
        return 'La contraseña debe tener al menos un carácter especial (!@#$%...)'
    return None




# ================================================================
# CICLO LECTIVO — calculado automáticamente por fecha del servidor
# ================================================================

def get_ciclo_lectivo():
    """
    Calcula el ciclo lectivo vigente según la fecha actual del servidor.
    El ciclo va del 01/04 de un año al 31/03 del siguiente.
    Ejemplo: hoy = 18/06/2026 → ciclo 2026-2027
             hoy = 15/02/2027 → ciclo 2026-2027 (aún no terminó)
    """
    hoy = date.today()
    if hoy.month >= 4:
        anio_inicio = hoy.year
    else:
        anio_inicio = hoy.year - 1
    return {
        'anio_inicio': anio_inicio,
        'anio_fin':    anio_inicio + 1,
        'inicio':      date(anio_inicio, 4, 1),
        'fin':         date(anio_inicio + 1, 3, 31),
        'label':       f"{anio_inicio} — {anio_inicio + 1}",
        'vencido':     hoy > date(anio_inicio + 1, 3, 31),
    }


def get_pendientes_libro_folio(carrera_id):
    """
    Retorna lista de alumnos promocionados sin libro/folio en la carrera dada.
    """
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("""
        SELECT a.apellido, a.nombre, a.dni, m.nombre AS materia,
               m.anio AS anio_materia, cu.id AS cursada_id,
               i.anio_lectivo
        FROM cursadas cu
        JOIN inscripciones i  ON i.id  = cu.inscripcion_id
        JOIN alumnos a        ON a.id  = i.alumno_id
        JOIN materias m       ON m.id  = i.materia_id
        WHERE cu.condicion = 'promocionado'
          AND NOT cu.promocion_provisoria
          AND (cu.libro IS NULL OR cu.folio IS NULL)
          AND m.carrera_id = %s
        ORDER BY m.anio, a.apellido, a.nombre
    """, (carrera_id,))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [{
        'apellido':    r[0], 'nombre':       r[1],
        'dni':         formatear_dni(r[2]),
        'materia':     r[3], 'anio_materia': r[4],
        'cursada_id':  r[5], 'anio_lectivo': r[6],
    } for r in rows]


def get_estado_inscripciones():
    """
    Calcula el estado de la ventana de inscripciones según:
    - fecha del servidor (date.today())
    - fechas configuradas en la DB (inscripciones_fecha_inicio / fin)
    - cierre manual del coordinador (inscripciones_cerrado_manual)

    Retorna dict con:
        abierto:    True/False
        motivo:     str (descripción del estado actual)
        fecha_inicio, fecha_fin:  date
        cerrado_manual:           bool
        motivo_cierre:            str
        hoy:                      date
    """
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("""
        SELECT clave, valor FROM configuracion
        WHERE clave IN ('inscripciones_fecha_inicio',
                        'inscripciones_fecha_fin',
                        'inscripciones_cerrado_manual',
                        'inscripciones_motivo_cierre')
    """)
    conf = {r[0]: r[1] for r in cur.fetchall()}
    cur.close()
    conn.close()

    hoy = date.today()

    try:
        finicio = date.fromisoformat(conf.get('inscripciones_fecha_inicio', '2026-04-28'))
        ffin    = date.fromisoformat(conf.get('inscripciones_fecha_fin',    '2026-05-31'))
    except Exception:
        finicio = date(hoy.year, 4, 28)
        ffin    = date(hoy.year, 5, 31)

    cerrado_manual = (conf.get('inscripciones_cerrado_manual', 'false').lower() == 'true')
    motivo_cierre  = conf.get('inscripciones_motivo_cierre', '') or ''

    if cerrado_manual:
        abierto = False
        motivo  = f'Cerrado manualmente por el coordinador'
        if motivo_cierre:
            # Limpiar puntos/espacios al final para evitar doble punto cuando se concatena
            motivo += f': {motivo_cierre.rstrip(". ")}'
    elif hoy < finicio:
        abierto = False
        motivo  = f'Las inscripciones abren el {finicio.strftime("%d/%m/%Y")}'
    elif hoy > ffin:
        abierto = False
        motivo  = f'Cerrado automáticamente — el período terminó el {ffin.strftime("%d/%m/%Y")}'
    else:
        abierto = True
        motivo  = f'Abierto hasta el {ffin.strftime("%d/%m/%Y")}'

    return {
        'abierto':        abierto,
        'motivo':         motivo,
        'fecha_inicio':   finicio.isoformat(),
        'fecha_fin':      ffin.isoformat(),
        'cerrado_manual': cerrado_manual,
        'motivo_cierre':  motivo_cierre,
        'hoy':            hoy.isoformat(),
    }


# ================================================================
# INDEX
# ================================================================

@auth.route('/')
def index():
    return redirect(url_for('auth.login'))


# ================================================================
# LOGIN
# ================================================================

@auth.route('/login', methods=['GET', 'POST'])
def login():
    if 'rol' in session:
        return redirect(url_for('auth.dashboard'))

    error = None

    if request.method == 'POST':
        usuario = limpiar_dni(request.form['usuario'].strip())
        password = request.form['password'].strip()

        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, nombre, apellido, rol, carrera_id, debe_cambiar_password, password_hash, dni
            FROM usuarios
            WHERE usuario = %s AND activo = TRUE
        """, (usuario,))
        user = cur.fetchone()
        cur.close()
        conn.close()

        if user and check_password_hash(user[6], password):
            session['user_id']    = user[0]
            session['nombre']     = user[1]
            session['apellido']   = user[2]
            session['rol']        = user[3]
            session['carrera_id'] = user[4]
            if user[3] == 'sys':
                return redirect(url_for('auth.dashboard'))
            if user[3] == 'admin' and not user[7]:
                # Cuenta admin sin DNI todavía → asistente de configuración inicial
                return redirect(url_for('auth.configurar_admin'))
            if user[5]:
                return redirect(url_for('auth.cambiar_password'))
            return redirect(url_for('auth.dashboard'))
        else:
            error = 'Usuario o contraseña incorrectos'

    return render_template('login.html', error=error)


# ================================================================
# CAMBIAR CONTRASEÑA (primer ingreso)
# ================================================================

@auth.route('/cambiar-password', methods=['GET', 'POST'])
def cambiar_password():
    if 'rol' not in session:
        return redirect(url_for('auth.login'))

    error = None

    if request.method == 'POST':
        nueva = request.form['nueva'].strip()
        confirmar = request.form['confirmar'].strip()

        error_validacion = validar_password_fuerte(nueva)
        if error_validacion:
            error = error_validacion
        elif nueva != confirmar:
            error = 'Las contraseñas no coinciden'
        else:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("""
                UPDATE usuarios
                SET password_hash = %s, debe_cambiar_password = FALSE
                WHERE id = %s
            """, (generate_password_hash(nueva), session['user_id']))
            conn.commit()
            cur.close()
            conn.close()
            return redirect(url_for('auth.dashboard'))

    return render_template('cambiar_password.html', error=error)


# ================================================================
# CONFIGURAR ADMINISTRADOR (asistente de primera configuración)
# ================================================================
# Se muestra una única vez, la primera vez que se ingresa con la cuenta
# bootstrap (usuario="admin", sin DNI todavía). Pide los datos reales del
# administrador y una contraseña con los mismos estándares que el resto
# de los roles. Al confirmar, la cuenta bootstrap pasa a ser la cuenta
# real del administrador (usuario = su DNI) y se cierra la sesión para
# que vuelva a entrar con sus credenciales definitivas.

@auth.route('/configurar-admin', methods=['GET', 'POST'])
def configurar_admin():
    if 'rol' not in session or session.get('rol') != 'admin':
        return redirect(url_for('auth.login'))

    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT dni FROM usuarios WHERE id = %s", (session['user_id'],))
    row = cur.fetchone()
    if row and row[0]:
        # Esta cuenta ya está configurada — no debería estar acá
        cur.close(); conn.close()
        return redirect(url_for('auth.dashboard'))

    error = None
    if request.method == 'POST':
        nombre    = request.form.get('nombre', '').strip()
        apellido  = request.form.get('apellido', '').strip()
        dni       = limpiar_dni(request.form.get('dni', ''))
        celular   = request.form.get('celular', '').strip() or None
        email     = request.form.get('email', '').strip() or None
        domicilio = request.form.get('domicilio', '').strip() or None
        nueva     = request.form.get('nueva', '').strip()
        confirmar = request.form.get('confirmar', '').strip()

        if not nombre or not apellido or not dni:
            error = 'Nombre, apellido y DNI son obligatorios'
        elif not dni.isdigit() or len(dni) < 7:
            error = 'DNI inválido'
        else:
            error_validacion = validar_password_fuerte(nueva)
            if error_validacion:
                error = error_validacion
            elif nueva != confirmar:
                error = 'Las contraseñas no coinciden'
            else:
                try:
                    cur.execute("""
                        UPDATE usuarios
                        SET usuario = %s, nombre = %s, apellido = %s, dni = %s,
                            celular = %s, email = %s, domicilio = %s,
                            password_hash = %s, debe_cambiar_password = FALSE
                        WHERE id = %s
                    """, (dni, nombre, apellido, dni, celular, email, domicilio,
                          generate_password_hash(nueva), session['user_id']))
                    conn.commit()
                    cur.close(); conn.close()
                    session.clear()
                    return redirect(url_for('auth.login'))
                except Exception as e:
                    conn.rollback()
                    if 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
                        error = 'Ya existe un usuario con ese DNI'
                    else:
                        error = 'Error al guardar: ' + str(e)

    cur.close()
    conn.close()
    return render_template('configurar_admin.html', error=error)


# ================================================================
# DASHBOARD
# ================================================================

@auth.route('/dashboard')
def dashboard():
    if 'rol' not in session:
        return redirect(url_for('auth.login'))

    rol        = session['rol']
    carrera_id = session.get('carrera_id')

    if rol == 'admin':
        conn_chk = get_db(); cur_chk = conn_chk.cursor()
        cur_chk.execute("SELECT dni FROM usuarios WHERE id = %s", (session['user_id'],))
        row_chk = cur_chk.fetchone()
        cur_chk.close(); conn_chk.close()
        if row_chk and not row_chk[0]:
            return redirect(url_for('auth.configurar_admin'))

    ciclo      = get_ciclo_lectivo()

    # Nombre de la carrera para mostrar en el header
    nombre_carrera = ''
    if carrera_id:
        conn_c = get_db(); cur_c = conn_c.cursor()
        cur_c.execute("SELECT COALESCE(nombre_corto, nombre) FROM carreras WHERE id = %s", (carrera_id,))
        row_c = cur_c.fetchone()
        cur_c.close(); conn_c.close()
        if row_c:
            nombre_carrera = row_c[0]

    # Calcular pendientes para coordinador y preceptora
    pendientes = []
    bloqueado  = False
    if rol in ('coordinador', 'preceptora') and carrera_id:
        pendientes = get_pendientes_libro_folio(carrera_id)
        # Bloqueo: si el ciclo venció Y hay pendientes
        if ciclo['vencido'] and pendientes:
            bloqueado = True

    return render_template(
        'dashboard.html',
        nombre=session['nombre'],
        apellido=session['apellido'],
        rol=rol,
        carrera_id=carrera_id,
        nombre_carrera=nombre_carrera,
        ciclo_label=ciclo['label'],
        pendientes_libro_folio=pendientes,
        bloqueado=bloqueado,
    )


# ================================================================
# LOGOUT
# ================================================================

@auth.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('auth.login'))


# ================================================================
# API — ADMINISTRACIÓN (solo admin)
# ================================================================

# ── CARRERAS ──────────────────────────────────────────────────

@auth.route('/api/carreras', methods=['GET'])
@login_requerido(['admin'])
def api_carreras_listar():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT c.id, c.nombre, c.nombre_corto, c.activa,
               COUNT(u.id) AS coordinadores
        FROM carreras c
        LEFT JOIN usuarios u ON u.carrera_id = c.id AND u.rol = 'coordinador' AND u.activo = TRUE
        GROUP BY c.id, c.nombre, c.nombre_corto, c.activa
        ORDER BY c.nombre
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify([{
        'id': r[0], 'nombre': r[1], 'nombre_corto': r[2], 'activa': r[3], 'coordinadores': r[4]
    } for r in rows])


@auth.route('/api/carreras', methods=['POST'])
@login_requerido(['admin'])
def api_carreras_crear():
    data = request.get_json()
    nombre = data.get('nombre', '').strip()
    nombre_corto = (data.get('nombre_corto') or '').strip() or None
    if not nombre:
        return jsonify({'error': 'El nombre es obligatorio'}), 400

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("INSERT INTO carreras (nombre, nombre_corto) VALUES (%s, %s) RETURNING id", (nombre, nombre_corto))
        nuevo_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({'ok': True, 'id': nuevo_id})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': 'Ya existe una carrera con ese nombre'}), 409
    finally:
        cur.close()
        conn.close()


@auth.route('/api/carreras/<int:cid>', methods=['PUT'])
@login_requerido(['admin'])
def api_carreras_editar(cid):
    data = request.get_json()
    nombre = data.get('nombre', '').strip()
    activa = data.get('activa', True)
    nombre_corto = (data.get('nombre_corto') or '').strip() or None
    if not nombre:
        return jsonify({'error': 'El nombre es obligatorio'}), 400

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("UPDATE carreras SET nombre = %s, nombre_corto = %s, activa = %s WHERE id = %s",
                    (nombre, nombre_corto, activa, cid))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': 'Ya existe una carrera con ese nombre'}), 409
    finally:
        cur.close()
        conn.close()


@auth.route('/api/carreras/<int:cid>/toggle', methods=['POST'])
@login_requerido(['admin'])
def api_carreras_toggle(cid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE carreras SET activa = NOT activa WHERE id = %s RETURNING activa", (cid,))
    resultado = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True, 'activa': resultado[0]})


# ── COORDINADORES ─────────────────────────────────────────────

@auth.route('/api/coordinadores', methods=['GET'])
@login_requerido(['admin'])
def api_coord_listar():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT u.id, u.usuario, u.nombre, u.apellido, u.dni,
               u.email, u.celular, u.activo, u.debe_cambiar_password,
               c.nombre AS carrera, u.carrera_id
        FROM usuarios u
        LEFT JOIN carreras c ON c.id = u.carrera_id
        WHERE u.rol = 'coordinador'
        ORDER BY u.apellido, u.nombre
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify([{
        'id': r[0], 'usuario': r[1], 'nombre': r[2], 'apellido': r[3],
        'dni': r[4], 'email': r[5], 'celular': r[6], 'activo': r[7],
        'debe_cambiar_password': r[8], 'carrera': r[9], 'carrera_id': r[10]
    } for r in rows])


@auth.route('/api/coordinadores', methods=['POST'])
@login_requerido(['admin'])
def api_coord_crear():
    data = request.get_json()
    nombre    = data.get('nombre', '').strip()
    apellido  = data.get('apellido', '').strip()
    dni       = data.get('dni', '').strip()
    email     = data.get('email', '').strip() or None
    celular   = data.get('celular', '').strip() or None
    carrera_id = data.get('carrera_id') or None

    if not nombre or not apellido or not dni:
        return jsonify({'error': 'Nombre, apellido y DNI son obligatorios'}), 400

    # El usuario es el DNI, contraseña inicial es el DNI también
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO usuarios (usuario, password_hash, rol, nombre, apellido, dni, email, celular, carrera_id, debe_cambiar_password)
            VALUES (%s, %s, 'coordinador', %s, %s, %s, %s, %s, %s, TRUE)
            RETURNING id
        """, (dni, generate_password_hash(dni), nombre, apellido, dni, email, celular, carrera_id))
        nuevo_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({'ok': True, 'id': nuevo_id})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': 'Ya existe un usuario con ese DNI'}), 409
    finally:
        cur.close()
        conn.close()


@auth.route('/api/coordinadores/<int:uid>', methods=['PUT'])
@login_requerido(['admin'])
def api_coord_editar(uid):
    data = request.get_json()
    nombre     = data.get('nombre', '').strip()
    apellido   = data.get('apellido', '').strip()
    email      = data.get('email', '').strip() or None
    celular    = data.get('celular', '').strip() or None
    carrera_id = data.get('carrera_id') or None

    if not nombre or not apellido:
        return jsonify({'error': 'Nombre y apellido son obligatorios'}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE usuarios SET nombre = %s, apellido = %s, email = %s, celular = %s, carrera_id = %s
        WHERE id = %s AND rol = 'coordinador'
    """, (nombre, apellido, email, celular, carrera_id, uid))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})


@auth.route('/api/coordinadores/<int:uid>/toggle', methods=['POST'])
@login_requerido(['admin'])
def api_coord_toggle(uid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE usuarios SET activo = NOT activo WHERE id = %s AND rol = 'coordinador' RETURNING activo", (uid,))
    resultado = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True, 'activo': resultado[0]})


@auth.route('/api/coordinadores/<int:uid>/reset', methods=['POST'])
@login_requerido(['admin'])
def api_coord_reset(uid):
    # Resetea la clave al DNI del coordinador
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT dni FROM usuarios WHERE id = %s AND rol = 'coordinador'", (uid,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        return jsonify({'error': 'Coordinador no encontrado'}), 404
    cur.execute("""
        UPDATE usuarios SET password_hash = %s, debe_cambiar_password = TRUE
        WHERE id = %s AND rol = 'coordinador'
    """, (generate_password_hash(row[0]), uid))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})


@auth.route('/api/coordinadores/<int:uid>', methods=['DELETE'])
@login_requerido(['admin'])
def api_coord_eliminar(uid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM usuarios WHERE id = %s AND rol = 'coordinador'", (uid,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})


# ================================================================
# API — CONFIGURACIÓN (solo admin)
# ================================================================

@auth.route('/api/config/anio-lectivo', methods=['GET'])
@login_requerido(['admin'])
def api_config_get():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    row = cur.fetchone()
    cur.close()
    conn.close()
    return jsonify({'anio': row[0] if row else '2026'})


@auth.route('/api/config/anio-lectivo', methods=['POST'])
@login_requerido(['admin'])
def api_config_set():
    data = request.get_json()
    anio = str(data.get('anio', '')).strip()
    if not anio.isdigit() or len(anio) != 4:
        return jsonify({'error': 'Año inválido'}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE configuracion SET valor = %s WHERE clave = 'anio_lectivo_actual'
    """, (anio,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})


# ================================================================
# API — STATS DASHBOARD ADMIN
# ================================================================
# SYS — GESTIÓN DE ADMINISTRADORES
# ================================================================

@auth.route('/api/sys/admins', methods=['GET'])
@login_requerido(['sys'])
def api_sys_admins():
    """Lista todos los usuarios con rol admin."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, apellido, nombre, dni, email, celular, domicilio, activo, debe_cambiar_password
        FROM usuarios WHERE rol = 'admin'
        ORDER BY apellido, nombre
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return jsonify([{
        'id': r[0], 'apellido': r[1] or '', 'nombre': r[2] or '',
        'dni': r[3] or '', 'email': r[4] or '', 'celular': r[5] or '',
        'domicilio': r[6] or '', 'activo': r[7],
        'debe_cambiar_password': r[8], 'configurado': bool(r[3])
    } for r in rows])


@auth.route('/api/sys/admins', methods=['POST'])
@login_requerido(['sys'])
def api_sys_crear_admin():
    """Crea un nuevo administrador."""
    d = request.json
    apellido  = (d.get('apellido') or '').strip()
    nombre    = (d.get('nombre') or '').strip()
    dni       = limpiar_dni(d.get('dni') or '')
    email     = (d.get('email') or '').strip() or None
    celular   = (d.get('celular') or '').strip() or None
    domicilio = (d.get('domicilio') or '').strip() or None

    if not apellido or not nombre or not dni:
        return jsonify({'error': 'Apellido, nombre y DNI son obligatorios'}), 400
    if not dni.isdigit() or len(dni) < 7:
        return jsonify({'error': 'DNI inválido'}), 400

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO usuarios (usuario, password_hash, rol, nombre, apellido, dni,
                                  email, celular, domicilio, debe_cambiar_password, activo)
            VALUES (%s, %s, 'admin', %s, %s, %s, %s, %s, %s, TRUE, TRUE)
            RETURNING id
        """, (dni, generate_password_hash(dni), nombre, apellido, dni, email, celular, domicilio))
        conn.commit()
        cur.close(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        cur.close(); conn.close()
        if 'unique' in str(e).lower():
            return jsonify({'error': 'Ya existe un usuario con ese DNI'}), 400
        return jsonify({'error': str(e)}), 500


@auth.route('/api/sys/admins/<int:uid>', methods=['PUT'])
@login_requerido(['sys'])
def api_sys_editar_admin(uid):
    """Edita los datos de un administrador."""
    d = request.json
    apellido  = (d.get('apellido') or '').strip()
    nombre    = (d.get('nombre') or '').strip()
    email     = (d.get('email') or '').strip() or None
    celular   = (d.get('celular') or '').strip() or None
    domicilio = (d.get('domicilio') or '').strip() or None

    if not apellido or not nombre:
        return jsonify({'error': 'Apellido y nombre son obligatorios'}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE usuarios SET apellido=%s, nombre=%s, email=%s, celular=%s, domicilio=%s
        WHERE id=%s AND rol='admin'
    """, (apellido, nombre, email, celular, domicilio, uid))
    conn.commit()
    cur.close(); conn.close()
    return jsonify({'ok': True})


@auth.route('/api/sys/admins/<int:uid>/reset', methods=['POST'])
@login_requerido(['sys'])
def api_sys_reset_admin(uid):
    """Resetea la contraseña del admin a su DNI."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT dni FROM usuarios WHERE id=%s AND rol='admin'", (uid,))
    row = cur.fetchone()
    if not row or not row[0]:
        cur.close(); conn.close()
        return jsonify({'error': 'Admin no encontrado o sin DNI configurado'}), 404
    cur.execute("""
        UPDATE usuarios SET password_hash=%s, debe_cambiar_password=TRUE
        WHERE id=%s AND rol='admin'
    """, (generate_password_hash(row[0]), uid))
    conn.commit()
    cur.close(); conn.close()
    return jsonify({'ok': True})


@auth.route('/api/sys/admins/<int:uid>/virgen', methods=['POST'])
@login_requerido(['sys'])
def api_sys_admin_virgen(uid):
    """Vuelve la cuenta admin a estado virgen (para cambio de persona)."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE usuarios
        SET usuario='admin', nombre=NULL, apellido=NULL, dni=NULL,
            celular=NULL, email=NULL, domicilio=NULL,
            password_hash=%s, debe_cambiar_password=TRUE
        WHERE id=%s AND rol='admin'
    """, (generate_password_hash('Sga-IES9#2026'), uid))
    conn.commit()
    cur.close(); conn.close()
    return jsonify({'ok': True})


@auth.route('/api/sys/admins/<int:uid>/toggle', methods=['POST'])
@login_requerido(['sys'])
def api_sys_toggle_admin(uid):
    """Activa o desactiva un administrador."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT activo FROM usuarios WHERE id=%s AND rol='admin'", (uid,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        return jsonify({'error': 'Admin no encontrado'}), 404
    cur.execute("UPDATE usuarios SET activo=%s WHERE id=%s", (not row[0], uid))
    conn.commit()
    cur.close(); conn.close()
    return jsonify({'ok': True})


@auth.route('/api/sys/admins/<int:uid>', methods=['DELETE'])
@login_requerido(['sys'])
def api_sys_eliminar_admin(uid):
    """Elimina permanentemente una cuenta de administrador."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM usuarios WHERE id=%s AND rol='admin'", (uid,))
    conn.commit()
    cur.close(); conn.close()
    return jsonify({'ok': True})


# ================================================================

def api_stats_admin():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM carreras WHERE activa = TRUE")
    carreras = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM usuarios WHERE rol = 'coordinador' AND activo = TRUE")
    coordinadores = cur.fetchone()[0]
    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = cur.fetchone()[0]
    cur.close()
    conn.close()
    return jsonify({'carreras': carreras, 'coordinadores': coordinadores, 'anio': anio})

# ================================================================
# DESCARGAR PLANTILLA PLAN DE ESTUDIOS
# ================================================================


@auth.route('/descargar-plantilla')
def descargar_plantilla():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan de Estudios"

    encabezados = [
        "Año",
        "Orden",
        "Espacios Curriculares",
        "Régimen",
        "Correlatividades - Regularizadas para cursar",
        "Correlatividades - Aprobadas para rendir",
        "Régimen de Aprobación",
    ]

    # Estilo encabezado
    fill    = PatternFill("solid", fgColor="1a4731")
    fuente  = Font(bold=True, color="FFFFFF", size=11)
    alin    = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, titulo in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.fill   = fill
        celda.font   = fuente
        celda.alignment = alin

    # Anchos de columna
    anchos = [8, 8, 35, 15, 38, 38, 22]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho

    ws.row_dimensions[1].height = 40

    # Fila de ejemplo
    ws.append([1, 1, "Matemática", "Cuatrimestral", "", "", "Promocional"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name="plantilla_plan_estudios.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ================================================================
# API — ESPACIOS CURRICULARES
# ================================================================

@auth.route('/api/materias', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_materias_listar():
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT m.id, m.nombre, m.anio, m.orden, m.regimen, m.regimen_aprobacion,
               STRING_AGG(CASE WHEN co.tipo = 'cursada' THEN r.orden::text END, '-' ORDER BY r.orden) AS correl_cursada,
               STRING_AGG(CASE WHEN co.tipo = 'aprobada' THEN r.orden::text END, '-' ORDER BY r.orden) AS correl_aprobada
        FROM materias m
        LEFT JOIN correlatividades co ON co.materia_id = m.id
        LEFT JOIN materias r ON r.id = co.requiere_materia_id
        WHERE m.carrera_id = %s AND m.activa = TRUE
        GROUP BY m.id, m.nombre, m.anio, m.orden, m.regimen, m.regimen_aprobacion
        ORDER BY m.anio, m.orden
    """, (carrera_id,))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify([{
        'id': r[0], 'nombre': r[1], 'anio': r[2], 'orden': r[3],
        'regimen': r[4], 'regimen_aprobacion': r[5],
        'correl_cursada': r[6], 'correl_aprobada': r[7]
    } for r in rows])

# ================================================================
# DESCARGAR PLAN DE ESTUDIOS EN PDF
# ================================================================

@auth.route('/api/materias/descargar-pdf')
@login_requerido(['coordinador', 'preceptora'])
def api_materias_descargar_pdf():
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT nombre FROM carreras WHERE id = %s", (carrera_id,))
    carrera = cur.fetchone()
    if not carrera:
        cur.close(); conn.close()
        return jsonify({'error': 'Carrera no encontrada'}), 404
    nombre_carrera = carrera[0]

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio_lectivo = cur.fetchone()[0]

    cur.execute("""
        SELECT m.anio, m.orden, m.nombre, m.regimen, m.regimen_aprobacion,
               STRING_AGG(CASE WHEN co.tipo = 'cursada' THEN r.orden::text END, ', ' ORDER BY r.orden) AS correl_cursada,
               STRING_AGG(CASE WHEN co.tipo = 'aprobada' THEN r.orden::text END, ', ' ORDER BY r.orden) AS correl_aprobada
        FROM materias m
        LEFT JOIN correlatividades co ON co.materia_id = m.id
        LEFT JOIN materias r ON r.id = co.requiere_materia_id
        WHERE m.carrera_id = %s AND m.activa = TRUE
        GROUP BY m.anio, m.orden, m.nombre, m.regimen, m.regimen_aprobacion
        ORDER BY m.anio, m.orden
    """, (carrera_id,))
    materias = cur.fetchall()
    cur.close()
    conn.close()

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle('titulo', parent=styles['Normal'],
        fontSize=13, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=4)
    estilo_sub = ParagraphStyle('sub', parent=styles['Normal'],
        fontSize=10, fontName='Helvetica', alignment=TA_CENTER, spaceAfter=2)
    estilo_celda = ParagraphStyle('celda', parent=styles['Normal'],
        fontSize=8, fontName='Helvetica', alignment=TA_LEFT)
    estilo_celda_centro = ParagraphStyle('celda_c', parent=styles['Normal'],
        fontSize=8, fontName='Helvetica', alignment=TA_CENTER)

    elementos = []

    # Logo con máscara circular
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo_ies9.png')
    if os.path.exists(logo_path):
        try:
            from PIL import Image as PILImage, ImageDraw
            pil_img = PILImage.open(logo_path).convert('RGBA')
            size = pil_img.size
            lado = min(size)
            left = (size[0] - lado) // 2
            top  = (size[1] - lado) // 2
            pil_img = pil_img.crop((left, top, left + lado, top + lado))
            margen = int(lado * 0.03)
            mascara = PILImage.new('L', (lado, lado), 0)
            draw = ImageDraw.Draw(mascara)
            draw.ellipse((margen, margen, lado - margen, lado - margen), fill=255)
            resultado = PILImage.new('RGBA', (lado, lado), (255, 255, 255, 0))
            resultado.paste(pil_img, mask=mascara)
            logo_buf = BytesIO()
            resultado.save(logo_buf, format='PNG')
            logo_buf.seek(0)
            logo = RLImage(logo_buf, width=2*cm, height=2*cm)
            logo.hAlign = 'CENTER'
            elementos.append(logo)
            elementos.append(Spacer(1, 0.2*cm))
        except Exception:
            pass

    # Encabezado
    elementos.append(Paragraph('Instituto de Educación Superior N° 9 "Juana Azurduy"', estilo_titulo))
    elementos.append(Paragraph('San Pedro de Jujuy — Jujuy', estilo_sub))
    elementos.append(Paragraph(f'{nombre_carrera}', estilo_sub))
    elementos.append(Paragraph(f'Plan de Estudios — Año lectivo {anio_lectivo}', estilo_sub))
    elementos.append(Spacer(1, 0.4*cm))

    # Encabezados de tabla
    encabezados = [
        Paragraph('<b>Año</b>', estilo_celda_centro),
        Paragraph('<b>Ord.</b>', estilo_celda_centro),
        Paragraph('<b>Espacio Curricular</b>', estilo_celda),
        Paragraph('<b>Régimen</b>', estilo_celda_centro),
        Paragraph('<b>Régimen de Aprobación</b>', estilo_celda_centro),
        Paragraph('<b>Correl. Regularizadas para cursar</b>', estilo_celda_centro),
        Paragraph('<b>Correl. Aprobadas para rendir</b>', estilo_celda_centro),
    ]
    filas = [encabezados]

    color_anio = {
        1: '#D6EAF8',
        2: '#FEF9E7',
        3: '#E9F7EF',
        4: '#FDEDEC',
        5: '#F4ECF7',
    }

    for m in materias:
        anio, orden, nombre, regimen, reg_aprobacion, correl_c, correl_a = m
        fila = [
            Paragraph(str(anio), estilo_celda_centro),
            Paragraph(str(orden), estilo_celda_centro),
            Paragraph(nombre or '', estilo_celda),
            Paragraph(regimen or '', estilo_celda_centro),
            Paragraph(reg_aprobacion or '', estilo_celda_centro),
            Paragraph(correl_c or '—', estilo_celda_centro),
            Paragraph(correl_a or '—', estilo_celda_centro),
        ]
        filas.append(fila)

    tabla = Table(filas, colWidths=[1.2*cm, 1.2*cm, 7*cm, 3.5*cm, 4*cm, 4.5*cm, 4.5*cm])

    estilo_tabla = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a4731')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
    ])

    for i, _ in enumerate(filas[1:], start=1):
        anio_fila = materias[i-1][0]
        hex_color = color_anio.get(anio_fila, '#FFFFFF')
        estilo_tabla.add('BACKGROUND', (0, i), (-1, i), colors.HexColor(hex_color))

    tabla.setStyle(estilo_tabla)
    elementos.append(tabla)

    doc.build(elementos)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name=f"plan_estudios_{anio_lectivo}.pdf",
        mimetype='application/pdf'
    )


# ================================================================
# API — IMPORTAR PLAN DE ESTUDIOS DESDE EXCEL
# ================================================================

@auth.route('/api/importar-plan', methods=['POST'])
@login_requerido(['coordinador'])
def api_importar_plan():
    carrera_id = session.get('carrera_id')
    if not carrera_id:
        return jsonify({'error': 'No tenés una carrera asignada'}), 400

    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename.endswith('.xlsx'):
        return jsonify({'error': 'Archivo inválido. Debe ser un .xlsx'}), 400

    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'No se pudo leer el archivo: {str(e)}'}), 400

    # ── Validar que el archivo realmente sea un plan de estudios ──
    # Se busca en la fila de encabezados al menos las columnas de "Año" y
    # "Espacios Curriculares" (o nombre/materia). Si no aparecen, el archivo
    # no tiene nada que ver con un plan de estudios y se rechaza de entrada,
    # sin tocar la base de datos.
    primera_fila = next(ws.iter_rows(min_row=1, max_row=1), [])
    header_vals = [str(c.value).strip().lower() if c.value else '' for c in primera_fila]
    tiene_anio   = any('año' in v or 'anio' in v for v in header_vals)
    tiene_nombre = any(p in v for v in header_vals for p in ['espacio', 'curricular', 'nombre', 'materia'])

    if not (tiene_anio and tiene_nombre):
        return jsonify({
            'error': 'Este archivo no parece ser un plan de estudios. No se encontraron '
                     'las columnas esperadas (Año, Espacios Curriculares, etc.). '
                     'Descargá la plantilla del sistema y usá esa estructura.'
        }), 400

    try:
        filas_nuevo = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            anio, orden, nombre, regimen, correl_cursada, correl_aprobada, regimen_aprobacion = (
                row[0], row[1], row[2], row[3], row[4], row[5], row[6]
            )
            if not nombre or not anio or not orden:
                continue
            try:
                anio_int  = int(anio)
                orden_int = int(orden)
            except (ValueError, TypeError):
                # Fila con datos que no coinciden con el formato esperado — se ignora
                # en vez de hacer fallar toda la importación por una fila suelta.
                continue
            filas_nuevo.append({
                'anio': anio_int,
                'orden': orden_int,
                'nombre': str(nombre).strip(),
                'regimen': str(regimen).strip() if regimen else None,
                'regimen_aprobacion': str(regimen_aprobacion).strip() if regimen_aprobacion else None,
                'correl_cursada': str(correl_cursada).strip() if correl_cursada else '',
                'correl_aprobada': str(correl_aprobada).strip() if correl_aprobada else '',
            })

        if not filas_nuevo:
            return jsonify({'error': 'El archivo no tiene datos válidos para importar.'}), 400

        conn = get_db()
        cur  = conn.cursor()

        # ── Comparar contra el plan actual (si no hay ninguno, queda vacío y
        #    todo cae naturalmente en "nuevas" — incluso la primera carga pasa
        #    siempre por la misma pantalla de revisión y confirmación) ──
        cur.execute("""
            SELECT id, nombre, anio, orden FROM materias
            WHERE carrera_id = %s ORDER BY anio, orden
        """, (carrera_id,))
        materias_actuales = {r[1].lower().strip(): {'id': r[0], 'anio': r[2], 'orden': r[3]}
                            for r in cur.fetchall()}
        cur.close(); conn.close()

        nombres_nuevos = {f['nombre'].lower().strip() for f in filas_nuevo}
        nombres_actuales = set(materias_actuales.keys())

        iguales    = nombres_actuales & nombres_nuevos
        eliminadas = nombres_actuales - nombres_nuevos
        nuevas     = nombres_nuevos   - nombres_actuales

        similares = []
        for nom_viejo in eliminadas:
            for nom_nuevo in nuevas:
                palabras_v = set(nom_viejo.split())
                palabras_n = set(nom_nuevo.split())
                if palabras_v and palabras_n:
                    interseccion = len(palabras_v & palabras_n)
                    union = len(palabras_v | palabras_n)
                    if union > 0 and interseccion / union >= 0.5:
                        similares.append({
                            'vieja': nom_viejo,
                            'nueva': nom_nuevo,
                            'id_vieja': materias_actuales[nom_viejo]['id']
                        })

        return jsonify({
            'ok': True,
            'hay_plan_actual': bool(materias_actuales),
            'comparacion': {
                'iguales':    list(iguales),
                'eliminadas': list(eliminadas),
                'nuevas':     list(nuevas),
                'similares':  similares,
            },
            'filas_nuevo': filas_nuevo,
            'total_nuevo': len(filas_nuevo),
        })

    except Exception as e:
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}'}), 500


def _ejecutar_importacion(cur, carrera_id, filas, plan_id=None):
    """Ejecuta la importación del plan de estudios."""
    cur.execute("DELETE FROM materias WHERE carrera_id = %s", (carrera_id,))
    orden_a_id = {}
    for f in filas:
        cur.execute("""
            INSERT INTO materias (carrera_id, nombre, anio, orden, regimen, regimen_aprobacion, plan_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id
        """, (carrera_id, f['nombre'], f['anio'], f['orden'],
              f['regimen'], f['regimen_aprobacion'], plan_id))
        orden_a_id[f['orden']] = cur.fetchone()[0]

    for f in filas:
        materia_id = orden_a_id[f['orden']]
        for num in f['correl_cursada'].split('-'):
            num = num.strip()
            if num.isdigit() and int(num) in orden_a_id:
                cur.execute("""
                    INSERT INTO correlatividades (materia_id, requiere_materia_id, tipo)
                    VALUES (%s, %s, 'cursada') ON CONFLICT DO NOTHING
                """, (materia_id, orden_a_id[int(num)]))
        for num in f['correl_aprobada'].split('-'):
            num = num.strip()
            if num.isdigit() and int(num) in orden_a_id:
                cur.execute("""
                    INSERT INTO correlatividades (materia_id, requiere_materia_id, tipo)
                    VALUES (%s, %s, 'aprobada') ON CONFLICT DO NOTHING
                """, (materia_id, orden_a_id[int(num)]))
    return orden_a_id


@auth.route('/api/confirmar-cambio-plan', methods=['POST'])
@login_requerido(['coordinador'])
def api_confirmar_cambio_plan():
    """
    Confirma el cambio de plan de estudios con:
    - nombre y resolución del nuevo plan
    - fecha de vigencia y fecha límite de transición
    - política de migración
    - equivalencias manuales definidas por el coordinador
    - filas del nuevo plan Excel
    """
    carrera_id = session.get('carrera_id')
    data = request.get_json()

    nombre_plan      = data.get('nombre_plan', '').strip()
    resolucion       = data.get('resolucion', '').strip()
    fecha_vigencia   = data.get('fecha_vigencia')
    fecha_cierre     = data.get('fecha_cierre')
    politica         = data.get('politica', 'exactas')
    equivalencias    = data.get('equivalencias', [])  # [{id_vieja, nombre_nueva}]
    filas_nuevo      = data.get('filas_nuevo', [])

    if not nombre_plan or not fecha_vigencia or not filas_nuevo:
        return jsonify({'error': 'Faltan datos obligatorios'}), 400

    conn = get_db()
    cur  = conn.cursor()
    try:
        # 1. Guardar materias actuales (plan viejo) antes de reemplazar
        cur.execute("""
            SELECT id, nombre FROM materias WHERE carrera_id = %s
        """, (carrera_id,))
        materias_viejas = {r[1].lower().strip(): r[0] for r in cur.fetchall()}

        # 2. Crear registro del nuevo plan
        cur.execute("""
            INSERT INTO planes_estudio
                (carrera_id, nombre, resolucion, fecha_vigencia, fecha_cierre,
                 politica_migracion, activo)
            VALUES (%s, %s, %s, %s, %s, %s, TRUE) RETURNING id
        """, (carrera_id, nombre_plan, resolucion or None,
              fecha_vigencia, fecha_cierre or None, politica))
        nuevo_plan_id = cur.fetchone()[0]

        # 3. Desactivar plan anterior
        cur.execute("""
            UPDATE planes_estudio SET activo = FALSE
            WHERE carrera_id = %s AND id != %s
        """, (carrera_id, nuevo_plan_id))

        # 4. Importar nuevo plan
        orden_a_id = _ejecutar_importacion(cur, carrera_id, filas_nuevo, nuevo_plan_id)
        nombres_nuevos = {f['nombre'].lower().strip(): orden_a_id[f['orden']]
                         for f in filas_nuevo}

        # 5. Registrar equivalencias
        # 5a. Automáticas (mismo nombre)
        for nom, id_vieja in materias_viejas.items():
            if nom in nombres_nuevos:
                id_nueva = nombres_nuevos[nom]
                cur.execute("""
                    INSERT INTO equivalencias_plan
                        (plan_nuevo_id, materia_nueva_id, materia_vieja_id, automatica)
                    VALUES (%s, %s, %s, TRUE)
                    ON CONFLICT DO NOTHING
                """, (nuevo_plan_id, id_nueva, id_vieja))

        # 5b. Manuales (definidas por el coordinador)
        for eq in equivalencias:
            id_vieja  = eq.get('id_vieja')
            nom_nueva = eq.get('nombre_nueva', '').lower().strip()
            if id_vieja and nom_nueva in nombres_nuevos:
                id_nueva = nombres_nuevos[nom_nueva]
                cur.execute("""
                    INSERT INTO equivalencias_plan
                        (plan_nuevo_id, materia_nueva_id, materia_vieja_id, automatica)
                    VALUES (%s, %s, %s, FALSE)
                    ON CONFLICT DO NOTHING
                """, (nuevo_plan_id, id_nueva, id_vieja))

        # 6. Migrar alumnos según política
        cur.execute("""
            SELECT id FROM alumnos WHERE carrera_id = %s
        """, (carrera_id,))
        alumnos = [r[0] for r in cur.fetchall()]

        for alumno_id in alumnos:
            cur.execute("""
                UPDATE alumnos SET plan_id = %s WHERE id = %s
            """, (nuevo_plan_id, alumno_id))
            cur.execute("""
                INSERT INTO historial_plan_alumno
                    (alumno_id, plan_nuevo_id, motivo, registrado_por)
                VALUES (%s, %s, 'automatico', %s)
            """, (alumno_id, nuevo_plan_id, session.get('user_id')))

        conn.commit()
        return jsonify({'ok': True, 'plan_id': nuevo_plan_id})

    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@auth.route('/api/plan-actual', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_plan_actual():
    """Devuelve info del plan activo de la carrera."""
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("""
        SELECT id, nombre, resolucion, fecha_vigencia, fecha_cierre,
               politica_migracion, activo
        FROM planes_estudio
        WHERE carrera_id = %s AND activo = TRUE
        ORDER BY fecha_vigencia DESC LIMIT 1
    """, (carrera_id,))
    r = cur.fetchone()
    cur.close(); conn.close()
    if not r:
        return jsonify({'plan': None})
    return jsonify({'plan': {
        'id': r[0], 'nombre': r[1], 'resolucion': r[2],
        'fecha_vigencia': str(r[3]) if r[3] else None,
        'fecha_cierre':   str(r[4]) if r[4] else None,
        'politica':       r[5], 'activo': r[6]
    }})


# ================================================================
# API — AGREGAR MATERIA MANUAL
# ================================================================

@auth.route('/api/materias/agregar', methods=['POST'])
@login_requerido(['coordinador'])
def api_materia_agregar():
    carrera_id = session.get('carrera_id')
    if not carrera_id:
        return jsonify({'error': 'No tenés una carrera asignada'}), 400

    data           = request.get_json()
    nombre         = data.get('nombre', '').strip()
    anio           = data.get('anio')
    orden          = data.get('orden')
    regimen        = data.get('regimen', '').strip() or None
    reg_aprobacion = data.get('regimen_aprobacion', '').strip() or None

    if not nombre:
        return jsonify({'error': 'El nombre es obligatorio'}), 400
    try:
        anio  = int(anio)
        orden = int(orden)
        if not (1 <= anio <= 6):
            raise ValueError
        if orden < 1:
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({'error': 'Año (1-6) y orden (≥1) son obligatorios y deben ser números válidos'}), 400

    conn = get_db()
    cur  = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO materias (carrera_id, nombre, anio, orden, regimen, regimen_aprobacion)
            VALUES (%s, %s, %s, %s, %s, %s) RETURNING id
        """, (carrera_id, nombre, anio, orden, regimen, reg_aprobacion))
        nuevo_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({'ok': True, 'id': nuevo_id})
    except Exception as e:
        conn.rollback()
        if 'unique' in str(e).lower():
            return jsonify({'error': f'Ya existe una materia con orden {orden} en {anio}° año'}), 409
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ================================================================
# API — STATS DASHBOARD COORDINADOR / PRECEPTORA
# ================================================================

@auth.route('/api/stats/carrera', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_stats_carrera():
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM alumnos WHERE carrera_id = %s AND activo = TRUE", (carrera_id,))
    total_alumnos = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM materias WHERE carrera_id = %s AND activa = TRUE", (carrera_id,))
    total_materias = cur.fetchone()[0]

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    row = cur.fetchone()
    anio = row[0] if row else '2026'

    cur.execute("""
        SELECT COUNT(DISTINCT c.id)
        FROM cursadas c
        JOIN inscripciones i ON i.id = c.inscripcion_id
        WHERE i.anio_lectivo = %s
          AND i.materia_id IN (SELECT id FROM materias WHERE carrera_id = %s)
          AND c.nota_cursada IS NOT NULL
    """, (anio, carrera_id))
    cursadas_cerradas = cur.fetchone()[0]

    # Últimos 5 alumnos
    cur.execute("""
        SELECT id, apellido, nombre, dni
        FROM alumnos
        WHERE carrera_id = %s AND activo = TRUE
        ORDER BY id DESC
        LIMIT 5
    """, (carrera_id,))
    ultimos = [{'id': r[0], 'apellido': r[1], 'nombre': r[2], 'dni': formatear_dni(r[3])} for r in cur.fetchall()]

    cur.close()
    conn.close()
    return jsonify({
        'alumnos': total_alumnos,
        'materias': total_materias,
        'cursadas': cursadas_cerradas,
        'anio': anio,
        'ultimos_alumnos': ultimos
    })


# ================================================================
# API — ALUMNOS (coordinador + preceptora)
# ================================================================

@auth.route('/api/alumnos', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_alumnos_listar():
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, apellido, nombre, dni, email, celular,
               fecha_nacimiento, direccion, localidad,
               contacto_emergencia_nombre, contacto_emergencia_telefono,
               activo, anio_ingreso, tipo_documento, cuil, provincia
        FROM alumnos
        WHERE carrera_id = %s
        ORDER BY apellido, nombre
    """, (carrera_id,))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify([{
        'id': r[0], 'apellido': r[1], 'nombre': r[2],
        'dni':             formatear_documento(r[13], r[3]),  # formato según tipo
        'dni_raw':         r[3],
        'tipo_documento':  r[13] or 'DNI',
        'tipo_doc_label':  etiqueta_documento(r[13]),
        'cuil':            r[14],
        'provincia':       r[15],
        'email': r[4], 'celular': r[5],
        'fecha_nacimiento': str(r[6]) if r[6] else None,
        'direccion': r[7], 'localidad': r[8],
        'contacto_emergencia_nombre': r[9],
        'contacto_emergencia_telefono': r[10],
        'activo': r[11], 'anio_ingreso': r[12]
    } for r in rows])


@auth.route('/api/alumnos', methods=['POST'])
@login_requerido(['coordinador', 'preceptora'])
def api_alumnos_crear():
    carrera_id = session.get('carrera_id')
    data = request.get_json()

    apellido       = data.get('apellido', '').strip()
    nombre         = data.get('nombre', '').strip()
    tipo_documento = (data.get('tipo_documento') or 'DNI').strip().upper()
    dni            = limpiar_documento(data.get('dni', ''))
    cuil           = limpiar_cuil(data.get('cuil', ''))
    provincia      = (data.get('provincia') or '').strip() or None
    fecha_nac      = data.get('fecha_nacimiento') or None

    if not apellido or not nombre:
        return jsonify({'error': 'Apellido y nombre son obligatorios'}), 400

    # Validar documento según tipo
    err_doc = validar_documento(tipo_documento, dni)
    if err_doc:
        return jsonify({'error': err_doc}), 400

    # Validar CUIL (si se cargó)
    err_cuil = validar_cuil(cuil)
    if err_cuil:
        return jsonify({'error': err_cuil}), 400

    # Validar provincia (si se cargó)
    err_prov = validar_provincia(provincia)
    if err_prov:
        return jsonify({'error': err_prov}), 400

    # Validar fecha de nacimiento (si se cargó)
    err_fnac = validar_fecha_nacimiento(fecha_nac)
    if err_fnac:
        return jsonify({'error': err_fnac}), 400

    # Validar ventana de inscripciones (misma que para inscripciones a materias)
    estado_ventana = get_estado_inscripciones()
    if not estado_ventana['abierto']:
        return jsonify({
            'error': f'No se pueden cargar alumnos nuevos. {estado_ventana["motivo"]}.',
            'ventana_cerrada': True
        }), 403

    conn = get_db()
    cur = conn.cursor()

    # Obtener año lectivo actual como fallback
    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    row = cur.fetchone()
    anio_lectivo = int(row[0]) if row else 2026

    anio_ingreso = data.get('anio_ingreso')
    try:
        anio_ingreso = int(anio_ingreso) if anio_ingreso else anio_lectivo
    except (ValueError, TypeError):
        anio_ingreso = anio_lectivo

    try:
        cur.execute("""
            INSERT INTO alumnos (
                carrera_id, apellido, nombre, dni, tipo_documento, cuil,
                email, celular, fecha_nacimiento, direccion, localidad, provincia,
                contacto_emergencia_nombre, contacto_emergencia_telefono,
                anio_ingreso
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            carrera_id, apellido, nombre, dni, tipo_documento, cuil,
            data.get('email', '').strip() or None,
            data.get('celular', '').strip() or None,
            fecha_nac,
            data.get('direccion', '').strip() or None,
            data.get('localidad', '').strip() or None,
            provincia,
            data.get('contacto_emergencia_nombre', '').strip() or None,
            data.get('contacto_emergencia_telefono', '').strip() or None,
            anio_ingreso,
        ))
        nuevo_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({'ok': True, 'id': nuevo_id})
    except Exception as e:
        conn.rollback()
        if 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
            return jsonify({'error': 'Ya existe un alumno con ese documento'}), 409
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@auth.route('/api/alumnos/<int:aid>', methods=['PUT'])
@login_requerido(['coordinador', 'preceptora'])
def api_alumnos_editar(aid):
    carrera_id = session.get('carrera_id')
    rol        = session.get('rol')
    data       = request.get_json()

    apellido       = data.get('apellido', '').strip()
    nombre         = data.get('nombre', '').strip()
    tipo_documento = (data.get('tipo_documento') or 'DNI').strip().upper()
    dni            = limpiar_documento(data.get('dni', ''))
    cuil           = limpiar_cuil(data.get('cuil', ''))
    provincia      = (data.get('provincia') or '').strip() or None
    fecha_nac      = data.get('fecha_nacimiento') or None

    if not apellido or not nombre:
        return jsonify({'error': 'Apellido y nombre son obligatorios'}), 400

    # Validar documento según tipo
    err_doc = validar_documento(tipo_documento, dni)
    if err_doc:
        return jsonify({'error': err_doc}), 400

    # Validar CUIL (si se cargó)
    err_cuil = validar_cuil(cuil)
    if err_cuil:
        return jsonify({'error': err_cuil}), 400

    # Validar provincia (si se cargó)
    err_prov = validar_provincia(provincia)
    if err_prov:
        return jsonify({'error': err_prov}), 400

    # Validar fecha de nacimiento (si se cargó)
    err_fnac = validar_fecha_nacimiento(fecha_nac)
    if err_fnac:
        return jsonify({'error': err_fnac}), 400

    conn = get_db()
    cur = conn.cursor()

    # Obtener documento actual del alumno
    cur.execute("SELECT dni, tipo_documento FROM alumnos WHERE id = %s AND carrera_id = %s", (aid, carrera_id))
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return jsonify({'error': 'Alumno no encontrado'}), 404

    dni_actual           = row[0]
    tipo_documento_actual = row[1] or 'DNI'
    documento_cambia      = (dni != dni_actual) or (tipo_documento != tipo_documento_actual)

    # Solo el coordinador puede cambiar el documento (número o tipo)
    if documento_cambia and rol != 'coordinador':
        cur.close()
        conn.close()
        return jsonify({'error': 'Solo el coordinador puede modificar el documento de un alumno'}), 403

    # Si el coordinador cambia el documento, verificar confirmación del número
    if documento_cambia and rol == 'coordinador':
        dni_confirmacion = limpiar_documento(data.get('dni_confirmacion', ''))
        if dni != dni_confirmacion:
            cur.close()
            conn.close()
            return jsonify({'error': 'Los documentos ingresados no coinciden'}), 400

    anio_ingreso = data.get('anio_ingreso')
    try:
        anio_ingreso = int(anio_ingreso) if anio_ingreso else None
    except (ValueError, TypeError):
        anio_ingreso = None

    try:
        cur.execute("""
            UPDATE alumnos SET
                apellido = %s, nombre = %s, dni = %s, tipo_documento = %s, cuil = %s,
                email = %s, celular = %s, fecha_nacimiento = %s,
                direccion = %s, localidad = %s, provincia = %s,
                contacto_emergencia_nombre = %s,
                contacto_emergencia_telefono = %s,
                anio_ingreso = %s
            WHERE id = %s AND carrera_id = %s
        """, (
            apellido, nombre, dni, tipo_documento, cuil,
            data.get('email', '').strip() or None,
            data.get('celular', '').strip() or None,
            fecha_nac,
            data.get('direccion', '').strip() or None,
            data.get('localidad', '').strip() or None,
            provincia,
            data.get('contacto_emergencia_nombre', '').strip() or None,
            data.get('contacto_emergencia_telefono', '').strip() or None,
            anio_ingreso,
            aid, carrera_id
        ))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        if 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
            return jsonify({'error': 'Ya existe un alumno con ese documento'}), 409
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@auth.route('/api/alumnos/<int:aid>/toggle', methods=['POST'])
@login_requerido(['coordinador', 'preceptora'])
def api_alumnos_toggle(aid):
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE alumnos SET activo = NOT activo
        WHERE id = %s AND carrera_id = %s
        RETURNING activo
    """, (aid, carrera_id))
    resultado = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True, 'activo': resultado[0]})


@auth.route('/api/alumnos/<int:aid>', methods=['DELETE'])
@login_requerido(['coordinador', 'preceptora'])
def api_alumnos_eliminar(aid):
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()
    # Verificar que el alumno no tenga inscripciones/cursadas antes de eliminar
    cur.execute("""
        SELECT COUNT(*) FROM inscripciones i
        JOIN materias m ON m.id = i.materia_id
        WHERE i.alumno_id = %s AND m.carrera_id = %s
    """, (aid, carrera_id))
    tiene_inscripciones = cur.fetchone()[0]
    if tiene_inscripciones:
        cur.close()
        conn.close()
        return jsonify({'error': 'No se puede eliminar: el alumno tiene inscripciones registradas. Podés desactivarlo en su lugar.'}), 409
    cur.execute("DELETE FROM alumnos WHERE id = %s AND carrera_id = %s", (aid, carrera_id))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})


# ================================================================
# API — HISTORIAL ACADÉMICO (coordinador + preceptora)
# ================================================================

@auth.route('/api/historial/alumno/<int:aid>', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_historial_alumno(aid):
    """Devuelve todas las cursadas del alumno agrupadas por año lectivo."""
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()

    # Verificar que el alumno pertenece a esta carrera
    cur.execute("""
        SELECT id, apellido, nombre, dni, anio_ingreso
        FROM alumnos WHERE id = %s AND carrera_id = %s
    """, (aid, carrera_id))
    alumno = cur.fetchone()
    if not alumno:
        cur.close(); conn.close()
        return jsonify({'error': 'Alumno no encontrado'}), 404

    # Todas las cursadas del alumno
    cur.execute("""
        SELECT
            i.anio_lectivo,
            m.nombre AS materia,
            m.anio AS anio_materia,
            m.orden,
            m.regimen,
            cu.id AS cursada_id,
            cu.nota_cursada,
            cu.porcentaje_asistencia,
            cu.porcentaje_tp,
            cu.condicion,
            cu.cerrada,
            cu.observaciones,
            i.id AS inscripcion_id,
            COALESCE(cu.promocion_provisoria, FALSE),
            cu.libro,
            cu.folio,
            (SELECT e.resultado FROM examenes e
              WHERE e.alumno_id = i.alumno_id AND e.materia_id = i.materia_id
              ORDER BY e.fecha_mesa DESC LIMIT 1),
            (SELECT e.nota FROM examenes e
              WHERE e.alumno_id = i.alumno_id AND e.materia_id = i.materia_id
              ORDER BY e.fecha_mesa DESC LIMIT 1)
        FROM inscripciones i
        JOIN materias m ON m.id = i.materia_id
        LEFT JOIN cursadas cu ON cu.inscripcion_id = i.id
        WHERE i.alumno_id = %s AND m.carrera_id = %s
        ORDER BY i.anio_lectivo DESC, m.anio, m.orden
    """, (aid, carrera_id))

    rows = cur.fetchall()
    cur.close()
    conn.close()

    # Agrupar por año lectivo
    anios = {}
    for r in rows:
        anio_lec = r[0]
        nota = float(r[6]) if r[6] is not None else None
        if anio_lec not in anios:
            anios[anio_lec] = []
        anios[anio_lec].append({
            'materia':               r[1],
            'anio_materia':          r[2],
            'orden':                 r[3],
            'regimen':               r[4],
            'cursada_id':            r[5],
            'nota_cursada':          nota,
            'nota_letras':           nota_en_letras(nota),
            'porcentaje_asistencia': float(r[7]) if r[7] is not None else None,
            'porcentaje_tp':         float(r[8]) if r[8] is not None else None,
            'condicion':             r[9],
            'cerrada':               r[10] or False,
            'observaciones':         r[11] or '',
            'inscripcion_id':        r[12],
            'promocion_provisoria':  bool(r[13]),
            'libro':                 r[14],
            'folio':                 r[15],
            'resultado_mesa':        r[16],
            'nota_mesa':             float(r[17]) if r[17] is not None else None,
        })

    return jsonify({
        'alumno': {
            'id': alumno[0], 'apellido': alumno[1],
            'nombre': alumno[2], 'dni': formatear_dni(alumno[3]),
            'anio_ingreso': alumno[4]
        },
        'anios': anios
    })


@auth.route('/api/historial/editar/<int:cursada_id>', methods=['PUT'])
@login_requerido(['coordinador'])
def api_historial_editar(cursada_id):
    """Permite al coordinador editar una cursada solo si NO está cerrada."""
    carrera_id = session.get('carrera_id')
    data = request.get_json()

    nota      = data.get('nota_cursada')
    asist     = data.get('porcentaje_asistencia')
    tp        = data.get('porcentaje_tp')
    condicion = data.get('condicion')
    obs       = data.get('observaciones', '')

    nota      = float(nota) if nota is not None and str(nota).strip() != '' else None
    asist     = float(asist) if asist is not None and str(asist).strip() != '' else None
    tp        = float(tp) if tp is not None and str(tp).strip() != '' else None
    condicion = condicion if condicion in ['regular', 'libre', 'promocionado', 'ausente'] else None

    conn = get_db()
    cur = conn.cursor()

    # Verificar que la cursada pertenece a la carrera y NO está cerrada
    cur.execute("""
        SELECT cu.id, cu.cerrada FROM cursadas cu
        JOIN inscripciones i ON i.id = cu.inscripcion_id
        JOIN materias m ON m.id = i.materia_id
        WHERE cu.id = %s AND m.carrera_id = %s
    """, (cursada_id, carrera_id))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        return jsonify({'error': 'Cursada no encontrada'}), 404
    if row[1]:  # cerrada = TRUE
        cur.close(); conn.close()
        return jsonify({'error': 'La cursada está cerrada y no puede modificarse.'}), 403

    cur.execute("""
        UPDATE cursadas SET
            nota_cursada = %s,
            porcentaje_asistencia = %s,
            porcentaje_tp = %s,
            condicion = %s,
            observaciones = %s,
            cerrada = TRUE
        WHERE id = %s
    """, (nota, asist, tp, condicion, obs or None, cursada_id))

    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})


# ================================================================
# HISTORIAL PDF — por año lectivo
# ================================================================

@auth.route('/api/historial/descargar-pdf/<int:aid>/<int:anio_lectivo>', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_historial_descargar_pdf(aid, anio_lectivo):
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur  = conn.cursor()

    cur.execute("""
        SELECT apellido, nombre, dni, tipo_documento FROM alumnos
        WHERE id = %s AND carrera_id = %s
    """, (aid, carrera_id))
    alumno = cur.fetchone()
    if not alumno:
        cur.close(); conn.close()
        return jsonify({'error': 'Alumno no encontrado'}), 404
    apellido, nombre, dni, tipo_doc = alumno

    cur.execute("SELECT nombre FROM carreras WHERE id = %s", (carrera_id,))
    nombre_carrera = cur.fetchone()[0]

    cur.execute("""
        SELECT m.nombre, m.anio, m.regimen,
               cu.nota_cursada, cu.porcentaje_asistencia,
               cu.porcentaje_tp, cu.condicion
        FROM inscripciones i
        JOIN materias m ON m.id = i.materia_id
        LEFT JOIN cursadas cu ON cu.inscripcion_id = i.id
        WHERE i.alumno_id = %s AND i.anio_lectivo = %s AND m.carrera_id = %s
        ORDER BY m.anio, m.orden
    """, (aid, anio_lectivo, carrera_id))
    cursadas = cur.fetchall()
    cur.close()
    conn.close()

    VERDE      = colors.HexColor('#1a4731')
    GRIS_PAR   = colors.HexColor('#F5F5F5')
    GRIS_BORDE = colors.HexColor('#CCCCCC')

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm,   bottomMargin=1.5*cm)

    styles  = getSampleStyleSheet()
    st_tit  = ParagraphStyle('t', parent=styles['Normal'], fontSize=12, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=3)
    st_sub  = ParagraphStyle('s', parent=styles['Normal'], fontSize=9,  fontName='Helvetica',      alignment=TA_CENTER, spaceAfter=2)
    st_head = ParagraphStyle('h', parent=styles['Normal'], fontSize=8,  fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=colors.white)
    st_cel  = ParagraphStyle('c', parent=styles['Normal'], fontSize=8,  fontName='Helvetica',      alignment=TA_LEFT)
    st_num  = ParagraphStyle('n', parent=styles['Normal'], fontSize=8,  fontName='Helvetica',      alignment=TA_CENTER)

    cond_label = {'promocionado': 'Promocionado', 'regular': 'Regular', 'libre': 'Libre', 'ausente': 'Ausente'}
    cond_color = {'promocionado': colors.HexColor('#1a6e3c'), 'regular': colors.HexColor('#1a3a6e'),
                  'libre': colors.HexColor('#9e2a2a'), 'ausente': colors.HexColor('#666666')}

    elementos = []

    # Logo
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo_ies9.png')
    if os.path.exists(logo_path):
        try:
            from PIL import Image as PILImage, ImageDraw
            pil_img = PILImage.open(logo_path).convert('RGBA')
            size = pil_img.size; lado = min(size)
            pil_img = pil_img.crop(((size[0]-lado)//2, (size[1]-lado)//2, (size[0]+lado)//2, (size[1]+lado)//2))
            margen = int(lado * 0.03)
            mascara = PILImage.new('L', (lado, lado), 0)
            ImageDraw.Draw(mascara).ellipse((margen, margen, lado-margen, lado-margen), fill=255)
            resultado = PILImage.new('RGBA', (lado, lado), (255, 255, 255, 0))
            resultado.paste(pil_img, mask=mascara)
            lb = BytesIO(); resultado.save(lb, format='PNG'); lb.seek(0)
            logo = RLImage(lb, width=1.8*cm, height=1.8*cm); logo.hAlign = 'CENTER'
            elementos.append(logo); elementos.append(Spacer(1, 0.2*cm))
        except Exception:
            pass

    elementos.append(Paragraph('Instituto de Educación Superior N° 9 "Juana Azurduy"', st_tit))
    elementos.append(Paragraph('San Pedro de Jujuy — Jujuy', st_sub))
    elementos.append(Paragraph(nombre_carrera, st_sub))
    elementos.append(Paragraph(f'Historial Académico — Año lectivo {anio_lectivo}', st_sub))
    elementos.append(Spacer(1, 0.2*cm))
    elementos.append(Paragraph(f'Alumno/a: <b>{apellido}, {nombre}</b>   —   {etiqueta_documento(tipo_doc)}: {formatear_documento(tipo_doc, dni)}', ParagraphStyle('m', parent=styles['Normal'], fontSize=9, fontName='Helvetica', alignment=TA_LEFT)))
    elementos.append(Spacer(1, 0.4*cm))

    enc = [Paragraph(t, st_head) for t in ['Espacio Curricular', 'Año', 'Régimen', 'Nota', 'Asist.', 'TP', 'Condición']]
    filas = [enc]
    for i, c in enumerate(cursadas):
        mat, anio_m, reg, nota, asist, tp, cond = c
        nota_f = f"{float(nota):.2f}".rstrip('0').rstrip('.') if nota is not None else '—'
        st_cond = ParagraphStyle('cd', parent=st_num, textColor=cond_color.get(cond, colors.HexColor('#333')), fontName='Helvetica-Bold')
        filas.append([
            Paragraph(mat, st_cel),
            Paragraph(f'{anio_m}°', st_num),
            Paragraph(reg or '—', st_num),
            Paragraph(nota_f, st_num),
            Paragraph(f'{float(asist):.0f}%' if asist is not None else '—', st_num),
            Paragraph(f'{float(tp):.0f}%' if tp is not None else '—', st_num),
            Paragraph(cond_label.get(cond, '—'), st_cond),
        ])

    tabla = Table(filas, colWidths=[7*cm, 1.2*cm, 3*cm, 1.5*cm, 1.5*cm, 1.5*cm, 3*cm])
    estilo = TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), VERDE),
        ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
        ('ALIGN',         (0,0), (-1,0), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('GRID',          (0,0), (-1,-1), 0.5, GRIS_BORDE),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING',   (0,0), (-1,-1), 5),
        ('RIGHTPADDING',  (0,0), (-1,-1), 5),
    ])
    for i in range(1, len(filas)):
        if i % 2 == 0:
            estilo.add('BACKGROUND', (0,i), (-1,i), GRIS_PAR)
    tabla.setStyle(estilo)
    elementos.append(tabla)

    doc.build(elementos)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f'historial_{apellido}_{anio_lectivo}.pdf')


# ================================================================
# HISTORIAL PDF — estado académico actual (por año del plan)
# ================================================================

@auth.route('/api/historial/descargar-estado/<int:aid>', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_historial_descargar_estado(aid):
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur  = conn.cursor()

    cur.execute("""
        SELECT apellido, nombre, dni, tipo_documento FROM alumnos
        WHERE id = %s AND carrera_id = %s
    """, (aid, carrera_id))
    alumno = cur.fetchone()
    if not alumno:
        cur.close(); conn.close()
        return jsonify({'error': 'Alumno no encontrado'}), 404
    apellido, nombre, dni, tipo_doc = alumno

    cur.execute("SELECT nombre FROM carreras WHERE id = %s", (carrera_id,))
    nombre_carrera = cur.fetchone()[0]

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio_actual = cur.fetchone()[0]

    # Todas las materias del plan con el mejor estado del alumno
    cur.execute("""
        SELECT
            m.id, m.nombre, m.anio, m.orden, m.regimen,
            (
                SELECT cu.condicion
                FROM inscripciones i2
                JOIN cursadas cu ON cu.inscripcion_id = i2.id
                WHERE i2.alumno_id = %s AND i2.materia_id = m.id
                  AND cu.condicion IS NOT NULL
                ORDER BY
                    CASE cu.condicion
                        WHEN 'promocionado' THEN 1
                        WHEN 'regular'      THEN 2
                        WHEN 'libre'        THEN 3
                        ELSE 4
                    END,
                    i2.anio_lectivo DESC
                LIMIT 1
            ) AS mejor_condicion
        FROM materias m
        WHERE m.carrera_id = %s AND m.activa = TRUE
        ORDER BY m.anio, m.orden
    """, (aid, carrera_id))
    materias = cur.fetchall()
    cur.close()
    conn.close()

    VERDE      = colors.HexColor('#1a4731')
    GRIS_BORDE = colors.HexColor('#CCCCCC')
    C_PROMO    = colors.HexColor('#1a6e3c')
    C_REG      = colors.HexColor('#1a3a6e')
    C_LIBRE    = colors.HexColor('#9e2a2a')
    C_PEND     = colors.HexColor('#888888')

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm,   bottomMargin=1.5*cm)

    styles  = getSampleStyleSheet()
    st_tit  = ParagraphStyle('t', parent=styles['Normal'], fontSize=12, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=3)
    st_sub  = ParagraphStyle('s', parent=styles['Normal'], fontSize=9,  fontName='Helvetica',      alignment=TA_CENTER, spaceAfter=2)
    st_head = ParagraphStyle('h', parent=styles['Normal'], fontSize=8,  fontName='Helvetica-Bold', alignment=TA_CENTER, textColor=colors.white)
    st_cel  = ParagraphStyle('c', parent=styles['Normal'], fontSize=8,  fontName='Helvetica',      alignment=TA_LEFT)
    st_num  = ParagraphStyle('n', parent=styles['Normal'], fontSize=8,  fontName='Helvetica',      alignment=TA_CENTER)
    st_ley  = ParagraphStyle('l', parent=styles['Normal'], fontSize=7,  fontName='Helvetica-Oblique', textColor=colors.HexColor('#555'), alignment=TA_LEFT)

    elementos = []

    # Logo
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo_ies9.png')
    if os.path.exists(logo_path):
        try:
            from PIL import Image as PILImage, ImageDraw
            pil_img = PILImage.open(logo_path).convert('RGBA')
            size = pil_img.size; lado = min(size)
            pil_img = pil_img.crop(((size[0]-lado)//2, (size[1]-lado)//2, (size[0]+lado)//2, (size[1]+lado)//2))
            margen = int(lado * 0.03)
            mascara = PILImage.new('L', (lado, lado), 0)
            ImageDraw.Draw(mascara).ellipse((margen, margen, lado-margen, lado-margen), fill=255)
            resultado = PILImage.new('RGBA', (lado, lado), (255, 255, 255, 0))
            resultado.paste(pil_img, mask=mascara)
            lb = BytesIO(); resultado.save(lb, format='PNG'); lb.seek(0)
            logo = RLImage(lb, width=1.8*cm, height=1.8*cm); logo.hAlign = 'CENTER'
            elementos.append(logo); elementos.append(Spacer(1, 0.2*cm))
        except Exception:
            pass

    elementos.append(Paragraph('Instituto de Educación Superior N° 9 "Juana Azurduy"', st_tit))
    elementos.append(Paragraph('San Pedro de Jujuy — Jujuy', st_sub))
    elementos.append(Paragraph(nombre_carrera, st_sub))
    elementos.append(Paragraph(f'Estado Académico — Año lectivo {anio_actual}', st_sub))
    elementos.append(Spacer(1, 0.2*cm))
    elementos.append(Paragraph(f'Alumno/a: <b>{apellido}, {nombre}</b>   —   {etiqueta_documento(tipo_doc)}: {formatear_documento(tipo_doc, dni)}',
        ParagraphStyle('m', parent=styles['Normal'], fontSize=9, fontName='Helvetica', alignment=TA_LEFT)))
    elementos.append(Spacer(1, 0.4*cm))

    # Agrupar por año del plan
    por_anio = {}
    for m in materias:
        a = m[2]
        if a not in por_anio:
            por_anio[a] = []
        por_anio[a].append(m)

    cond_label = {'promocionado': '✔ Aprobada', 'regular': '~ Regular (adeuda final)',
                  'libre': '✘ Libre (adeuda)', None: '— Sin cursar'}
    cond_col   = {'promocionado': C_PROMO, 'regular': C_REG, 'libre': C_LIBRE, None: C_PEND}

    for anio_plan in sorted(por_anio.keys()):
        mats = por_anio[anio_plan]
        elementos.append(Paragraph(f'{anio_plan}° Año del Plan de Estudios',
            ParagraphStyle('ap', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold',
                           textColor=VERDE, spaceBefore=10, spaceAfter=4)))

        enc = [Paragraph(t, st_head) for t in ['Espacio Curricular', 'Régimen', 'Estado']]
        filas = [enc]
        for i, m in enumerate(mats):
            mid, mat, anio_m, orden, reg, cond = m
            st_est = ParagraphStyle('e', parent=st_num, textColor=cond_col.get(cond, C_PEND),
                                    fontName='Helvetica-Bold' if cond == 'promocionado' else 'Helvetica')
            filas.append([
                Paragraph(mat, st_cel),
                Paragraph(reg or '—', st_num),
                Paragraph(cond_label.get(cond, '— Sin cursar'), st_est),
            ])

        tabla = Table(filas, colWidths=[8*cm, 3*cm, 6.5*cm])
        estilo = TableStyle([
            ('BACKGROUND',    (0,0), (-1,0), VERDE),
            ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
            ('GRID',          (0,0), (-1,-1), 0.5, GRIS_BORDE),
            ('TOPPADDING',    (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING',   (0,0), (-1,-1), 5),
            ('RIGHTPADDING',  (0,0), (-1,-1), 5),
        ])
        for i in range(1, len(filas)):
            if i % 2 == 0:
                estilo.add('BACKGROUND', (0,i), (-1,i), colors.HexColor('#F5F5F5'))
        tabla.setStyle(estilo)
        elementos.append(tabla)

    elementos.append(Spacer(1, 0.4*cm))
    elementos.append(Paragraph(
        '✔ Aprobada = Promocionó o aprobó el final   |   ~ Regular = Adeuda examen final   |   ✘ Libre = Debe recursar   |   — Sin cursar = No inscripta aún',
        st_ley))

    doc.build(elementos)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f'estado_academico_{apellido}.pdf')


# ================================================================
# API — CONSTANCIA DE ALUMNO REGULAR (ítem G)
# ================================================================

@auth.route('/api/alumnos/<int:aid>/constancia-pdf', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_constancia_alumno(aid):
    """
    Genera la constancia de alumno regular en PDF.
    Valida las condiciones según el ciclo lectivo vigente:
    - 1° año ingresante  → puede emitirse sin condición de materias
    - 1° año recursante  → NO puede emitirse
    - 2° año en adelante → necesita al menos 2 materias aprobadas o promocionadas
                          en el ciclo lectivo vigente
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib import colors

    carrera_id = session.get('carrera_id')
    ciclo      = get_ciclo_lectivo()

    conn = get_db()
    cur  = conn.cursor()

    # Datos del alumno
    cur.execute("""
        SELECT a.apellido, a.nombre, a.dni, a.carrera_id,
               c.nombre AS carrera
        FROM alumnos a
        JOIN carreras c ON c.id = a.carrera_id
        WHERE a.id = %s AND a.carrera_id = %s
    """, (aid, carrera_id))
    alumno = cur.fetchone()
    if not alumno:
        cur.close(); conn.close()
        return jsonify({'error': 'Alumno no encontrado'}), 404

    apellido, nombre, dni, _, nombre_carrera = alumno

    # ── Verificar condición para emitir ──
    # Anio lectivo actual
    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio_actual = int(cur.fetchone()[0])

    # ¿Tiene inscripciones este año?
    cur.execute("""
        SELECT COUNT(*) FROM inscripciones i
        JOIN materias m ON m.id = i.materia_id
        WHERE i.alumno_id = %s AND m.carrera_id = %s
          AND i.anio_lectivo = %s
    """, (aid, carrera_id, anio_actual))
    inscripto_este_anio = cur.fetchone()[0] > 0

    # anio_ingreso del alumno — define si es ingresante
    cur.execute("SELECT anio_ingreso FROM alumnos WHERE id = %s", (aid,))
    anio_ingreso = cur.fetchone()[0]
    es_ingresante = (anio_ingreso == anio_actual)

    # ¿Qué año del plan está cursando?
    cur.execute("""
        SELECT COALESCE(MAX(m.anio), 0) FROM inscripciones i
        JOIN materias m ON m.id = i.materia_id
        WHERE i.alumno_id = %s AND m.carrera_id = %s AND i.anio_lectivo = %s
    """, (aid, carrera_id, anio_actual))
    anio_carrera = cur.fetchone()[0]
    es_recursante = (anio_carrera == 1 and not es_ingresante)

    # ¿Cuántas materias aprobadas/promocionadas en el ciclo vigente?
    cur.execute("""
        SELECT COUNT(*)
        FROM cursadas cu
        JOIN inscripciones i ON i.id = cu.inscripcion_id
        JOIN materias m ON m.id = i.materia_id
        WHERE i.alumno_id = %s AND m.carrera_id = %s
          AND i.anio_lectivo = %s
          AND cu.condicion IN ('aprobado', 'promocionado')
    """, (aid, carrera_id, anio_actual))
    cant_aprobadas = cur.fetchone()[0]

    cur.close(); conn.close()

    # ── Validación ──
    puede_emitir = False
    motivo_no    = ''

    if not inscripto_este_anio:
        motivo_no = f'El alumno no tiene inscripciones en el año lectivo {anio_actual}.'
    elif anio_carrera == 1 and es_ingresante:
        # 1° año ingresante sin aprobadas → puede igual
        puede_emitir = True
    elif cant_aprobadas >= 2:
        # Tiene 2+ aprobadas → puede (sin importar si es recursante)
        puede_emitir = True
    elif es_recursante:
        materia_txt = '1 materia aprobada' if cant_aprobadas == 1 else f'{cant_aprobadas} materias aprobadas'
        motivo_no = (f'Recursante de 1° año con {materia_txt}. '
                    f'Necesita al menos 2 para emitir la constancia.')
    else:
        motivo_no = (f'Necesita al menos 2 materias aprobadas o promocionadas '
                    f'en el ciclo lectivo {ciclo["label"]}. '
                    f'Actualmente tiene {cant_aprobadas}.')

    if not puede_emitir:
        return jsonify({'error': motivo_no}), 400

    # ── Generar PDF ──
    PAGE_W, PAGE_H = A4
    LM = 2.0*cm; RM = 2.0*cm; TM = 1.5*cm
    W  = PAGE_W - LM - RM
    NEGRO = colors.black

    def linea_punteada(c, x1, x2, y):
        c.setDash(1, 3)
        c.setLineWidth(0.5)
        c.line(x1, y - 2, x2, y - 2)
        c.setDash()

    def txt(c, t, x, y, size=11, bold=False):
        fn = 'Helvetica-Bold' if bold else 'Helvetica'
        c.setFont(fn, size)
        c.drawString(x, y, t)
        return c.stringWidth(t, fn, size)

    buf = BytesIO()
    c   = rl_canvas.Canvas(buf, pagesize=A4)

    y = PAGE_H - TM

    # Logo placeholder (en producción reemplazar con imagen real)
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo_ies9.png')
    if os.path.exists(logo_path):
        try:
            from PIL import Image as PILImage, ImageDraw
            from reportlab.lib.utils import ImageReader
            pil_img = PILImage.open(logo_path).convert('RGBA')
            size    = pil_img.size
            lado    = min(size)
            left    = (size[0] - lado) // 2
            top     = (size[1] - lado) // 2
            pil_img = pil_img.crop((left, top, left + lado, top + lado))
            margen  = int(lado * 0.03)
            mascara = PILImage.new('L', (lado, lado), 0)
            draw    = ImageDraw.Draw(mascara)
            draw.ellipse((margen, margen, lado - margen, lado - margen), fill=255)
            resultado = PILImage.new('RGBA', (lado, lado), (255, 255, 255, 0))
            resultado.paste(pil_img, mask=mascara)
            logo_buf = BytesIO()
            resultado.save(logo_buf, format='PNG')
            logo_buf.seek(0)
            c.drawImage(ImageReader(logo_buf), LM, y - 2.8*cm,
                        width=2.6*cm, height=2.6*cm, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass
    else:
        c.setStrokeColor(NEGRO)
        c.setLineWidth(1)
        c.circle(LM + 1.3*cm, y - 1.4*cm, 1.2*cm)
        c.setFont('Helvetica', 6)
        c.drawCentredString(LM + 1.3*cm, y - 1.4*cm, 'IES N°9')

    # Título
    c.setFont('Helvetica-Bold', 16)
    c.drawCentredString(PAGE_W/2 + 1*cm, y - 0.8*cm, 'CONSTANCIA DE ALUMNO/A REGULAR')
    titulo_w = c.stringWidth('CONSTANCIA DE ALUMNO/A REGULAR', 'Helvetica-Bold', 16)
    tx = (PAGE_W/2 + 1*cm) - titulo_w/2
    c.setLineWidth(1)
    c.line(tx, y - 0.97*cm, tx + titulo_w, y - 0.97*cm)

    y -= 3.2*cm

    # Línea 1: Establecimiento
    x = LM
    x += txt(c, 'Establecimiento:', x, y, 11, bold=True); x += 0.2*cm
    linea_punteada(c, x, LM + W, y)
    y -= 0.9*cm

    # Línea 2: Se hace constar que
    x = LM
    x += txt(c, 'Se hace constar que', x, y, 11, bold=True); x += 0.2*cm
    linea_punteada(c, x, LM + W, y)
    y -= 0.9*cm

    # Línea 3: DNI + es alumn_ regular del + curso + división
    x = LM
    x += txt(c, 'D.N.I. N°', x, y, 11, bold=True); x += 0.2*cm
    linea_punteada(c, x, x + 3.5*cm, y); x += 3.6*cm
    x += txt(c, 'es alumn', x, y, 11, bold=True); x += 0.1*cm
    linea_punteada(c, x, x + 0.6*cm, y); x += 0.7*cm
    x += txt(c, 'regular del', x, y, 11, bold=True); x += 0.2*cm
    linea_punteada(c, x, x + 1.5*cm, y); x += 1.6*cm
    x += txt(c, 'curso', x, y, 11, bold=True); x += 0.2*cm
    linea_punteada(c, x, x + 1.5*cm, y); x += 1.6*cm
    x += txt(c, 'división', x, y, 11, bold=True); x += 0.2*cm
    linea_punteada(c, x, LM + W, y)
    y -= 0.9*cm

    # Línea 4: turno
    x = LM
    x += txt(c, 'turno', x, y, 11, bold=True); x += 0.2*cm
    linea_punteada(c, x, LM + W, y)
    y -= 0.9*cm

    # Línea 5: (*)
    x = LM
    x += txt(c, '(*)', x, y, 11, bold=True); x += 0.2*cm
    linea_punteada(c, x, LM + W, y)
    y -= 0.9*cm

    # Línea 6: extra
    linea_punteada(c, LM, LM + W, y)
    y -= 0.9*cm

    # Línea 7: Al pedido de + interesad
    x = LM
    x += txt(c, 'Al pedido de', x, y, 11, bold=True); x += 0.2*cm
    linea_punteada(c, x, x + 0.8*cm, y); x += 0.9*cm
    x += txt(c, 'interesad', x, y, 11, bold=True); x += 0.1*cm
    linea_punteada(c, x, x + 0.6*cm, y); x += 0.7*cm
    txt(c, 'y al solo efecto de su presentación ante las autoridades de:', x, y, 11, bold=True)
    y -= 0.9*cm

    # Línea 8: autoridades
    linea_punteada(c, LM, LM + W, y)
    y -= 0.9*cm

    # Línea 9: se le extiende
    x = LM
    x += txt(c, 'se le extiende la presente constancia en', x, y, 11, bold=True); x += 0.2*cm
    linea_punteada(c, x, LM + W, y)
    y -= 0.9*cm

    # Línea 10: a los + días del mes + año
    x = LM
    x += txt(c, 'a los', x, y, 11, bold=True); x += 0.2*cm
    linea_punteada(c, x, x + 2*cm, y); x += 2.1*cm
    x += txt(c, 'días del mes de', x, y, 11, bold=True); x += 0.2*cm
    linea_punteada(c, x, x + 3.5*cm, y); x += 3.6*cm
    x += txt(c, 'del año 20', x, y, 11, bold=True); x += 0.1*cm
    linea_punteada(c, x, x + 1.2*cm, y)
    y -= 0.7*cm

    # Nota al pie
    c.setFont('Helvetica', 7.5)
    c.drawString(LM, y, '(*) Se debe asentar si asiste a clase regularmente, finalizó curso, año lectivo, etc.')
    y -= 1.8*cm

    # Sello y firma
    firma_x = LM + W - 6*cm
    c.setLineWidth(0.5)
    c.line(firma_x, y - 1.5*cm, firma_x + 6*cm, y - 1.5*cm)
    c.setFont('Helvetica', 10)
    c.drawString(LM + 1*cm, y - 1.8*cm, 'SELLO')
    c.setFont('Helvetica', 9)
    c.drawCentredString(firma_x + 3*cm, y - 1.8*cm, 'Firma Autorizada — Sello y Aclaración')

    c.save()
    buf.seek(0)
    nombre_archivo = f'constancia_{apellido}_{nombre}.pdf'.replace(' ', '_')
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=nombre_archivo)


@auth.route('/api/alumnos/<int:aid>/constancia-validar', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_constancia_validar(aid):
    """Valida si el alumno puede recibir constancia, sin generar el PDF."""
    carrera_id = session.get('carrera_id')
    ciclo      = get_ciclo_lectivo()
    conn = get_db()
    cur  = conn.cursor()

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio_actual = int(cur.fetchone()[0])

    # Datos del alumno — anio_ingreso es el año en que ingresó a la carrera
    cur.execute("SELECT anio_ingreso FROM alumnos WHERE id = %s", (aid,))
    row = cur.fetchone()
    anio_ingreso = row[0] if row else anio_actual

    # ¿Tiene inscripciones este año?
    cur.execute("""
        SELECT COUNT(*) FROM inscripciones i
        JOIN materias m ON m.id = i.materia_id
        WHERE i.alumno_id = %s AND m.carrera_id = %s AND i.anio_lectivo = %s
    """, (aid, carrera_id, anio_actual))
    inscripto = cur.fetchone()[0] > 0

    # ¿Qué año del plan de estudios está cursando?
    cur.execute("""
        SELECT COALESCE(MAX(m.anio), 0) FROM inscripciones i
        JOIN materias m ON m.id = i.materia_id
        WHERE i.alumno_id = %s AND m.carrera_id = %s AND i.anio_lectivo = %s
    """, (aid, carrera_id, anio_actual))
    anio_carrera = cur.fetchone()[0]

    # Ingresante = primer año lectivo en el sistema es el actual
    es_ingresante = (anio_ingreso == anio_actual)
    es_recursante = (anio_carrera == 1 and not es_ingresante)

    # Materias aprobadas/promocionadas en el ciclo vigente
    cur.execute("""
        SELECT COUNT(*) FROM cursadas cu
        JOIN inscripciones i ON i.id = cu.inscripcion_id
        JOIN materias m ON m.id = i.materia_id
        WHERE i.alumno_id = %s AND m.carrera_id = %s
          AND i.anio_lectivo = %s AND cu.condicion IN ('aprobado','promocionado')
          -- Una promoción provisoria todavía puede caerse: no habilita
          -- la emisión de un documento oficial.
          AND NOT COALESCE(cu.promocion_provisoria, FALSE)
    """, (aid, carrera_id, anio_actual))
    cant_aprobadas = cur.fetchone()[0]

    cur.close(); conn.close()

    if not inscripto:
        return jsonify({'puede': False, 'motivo': f'Sin inscripciones en el año lectivo {anio_actual}'})
    # 1° año ingresante sin materias aprobadas → puede igual
    if anio_carrera == 1 and es_ingresante:
        return jsonify({'puede': True, 'motivo': 'Ingresante de 1° año'})
    # Tiene 2 o más aprobadas/promocionadas → puede (sin importar si es recursante)
    if cant_aprobadas >= 2:
        return jsonify({'puede': True, 'motivo': f'{cant_aprobadas} materias aprobadas/promocionadas'})
    # Recursante sin suficientes aprobadas → no puede
    if es_recursante:
        materia_txt = '1 materia aprobada' if cant_aprobadas == 1 else f'{cant_aprobadas} materias aprobadas'
        return jsonify({'puede': False,
                        'motivo': f'Recursante de 1° año con {materia_txt}. Necesita al menos 2.'})
    return jsonify({'puede': False,
                    'motivo': f'Necesita al menos 2 materias aprobadas o promocionadas. Tiene {cant_aprobadas}.'})


# ================================================================
# API — INSCRIPCIONES (coordinador + preceptora)
# ================================================================

@auth.route('/api/inscripciones/alumno/<int:aid>', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_inscripciones_alumno(aid):
    """
    Devuelve todas las materias del plan de la carrera,
    indicando para cada una:
      - si el alumno está inscripto este año lectivo
      - si cumple correlatividades (y si no, por qué)
    """
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()

    # Año lectivo actual
    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])

    # Verificar que el alumno pertenece a esta carrera
    cur.execute("SELECT id, apellido, nombre, dni, tipo_documento FROM alumnos WHERE id = %s AND carrera_id = %s", (aid, carrera_id))
    alumno = cur.fetchone()
    if not alumno:
        cur.close(); conn.close()
        return jsonify({'error': 'Alumno no encontrado'}), 404

    # Todas las materias activas de la carrera
    cur.execute("""
        SELECT id, nombre, anio, orden, regimen, regimen_aprobacion
        FROM materias
        WHERE carrera_id = %s AND activa = TRUE
        ORDER BY anio, orden
    """, (carrera_id,))
    materias = cur.fetchall()

    # Inscripciones actuales del alumno este año
    cur.execute("""
        SELECT materia_id FROM inscripciones
        WHERE alumno_id = %s AND anio_lectivo = %s
    """, (aid, anio))
    inscriptas = {r[0] for r in cur.fetchall()}

    # Historial académico del alumno para evaluar correlatividades.
    # REGLA INSTITUCIONAL: solo cuentan materias regularizadas/aprobadas
    # en ciclos lectivos ANTERIORES al actual — no el ciclo en curso.
    # Ejemplo: si estamos en 2026, solo cuentan cursadas de 2025 o antes.
    # Esto impide que un alumno se inscriba en 2° año dentro del mismo
    # ciclo en que está cursando 1° año, que es el comportamiento correcto.

    # Materias REGULARIZADAS en ciclos anteriores
    cur.execute("""
        SELECT DISTINCT i.materia_id
        FROM inscripciones i
        JOIN cursadas cu ON cu.inscripcion_id = i.id
        WHERE i.alumno_id = %s
          AND i.anio_lectivo < %s
          AND cu.condicion IN ('regular', 'promocionado', 'aprobado')
    """, (aid, anio))
    cursadas_ok = {r[0] for r in cur.fetchall()}

    # Materias APROBADAS en ciclos anteriores
    cur.execute("""
        SELECT DISTINCT materia_id FROM (
            SELECT i.materia_id
            FROM inscripciones i
            JOIN cursadas cu ON cu.inscripcion_id = i.id
            WHERE i.alumno_id = %s
              AND i.anio_lectivo < %s
              AND cu.condicion = 'promocionado'
            UNION
            SELECT materia_id
            FROM examenes
            WHERE alumno_id = %s AND resultado = 'aprobado'
        ) sub
    """, (aid, anio, aid))
    aprobadas_ok = {r[0] for r in cur.fetchall()}

    # Correlatividades de todas las materias
    cur.execute("""
        SELECT materia_id, requiere_materia_id, tipo
        FROM correlatividades
        WHERE materia_id IN (
            SELECT id FROM materias WHERE carrera_id = %s
        )
    """, (carrera_id,))
    correlatividades = cur.fetchall()

    # Mapear correlatividades por materia
    correl_map = {}
    for materia_id, requiere_id, tipo in correlatividades:
        if materia_id not in correl_map:
            correl_map[materia_id] = []
        correl_map[materia_id].append((requiere_id, tipo))

    # Mapa id → nombre para mostrar en mensajes
    cur.execute("SELECT id, nombre, orden, anio FROM materias WHERE carrera_id = %s", (carrera_id,))
    rows_mat = cur.fetchall()
    materia_nombres = {r[0]: f"({r[2]}) {r[1]}" for r in rows_mat}
    # Mapa id → año de la materia (para saber a qué año pertenece cada materia del historial)
    materia_anio = {r[0]: r[3] for r in rows_mat}

    cur.close()
    conn.close()

    # Años desbloqueados para este alumno (Opción C):
    # - 1° año: siempre disponible
    # - Año X (X > 1): disponible si el alumno tiene al menos 1 materia
    #   regularizada o aprobada del año X-1
    historial_ids = cursadas_ok | aprobadas_ok  # union de regularizadas + aprobadas
    anios_con_historial = {materia_anio[mid] for mid in historial_ids if mid in materia_anio}

    def anio_habilitado(anio_materia):
        if anio_materia == 1:
            return True
        return (anio_materia - 1) in anios_con_historial

    # Evaluar cada materia
    resultado = []
    for m in materias:
        mid, nombre, anio_m, orden, regimen, reg_aprobacion = m
        inscripta = mid in inscriptas

        puede = True
        bloqueada_por = []

        # Primero verificar si el año está habilitado (Opción C)
        if not anio_habilitado(anio_m) and not inscripta:
            puede = False
            anio_prev = anio_m - 1
            bloqueada_por.append(f"Necesitás regularizar al menos una materia de {anio_prev}° año")
        else:
            # Si el año está habilitado, verificar correlatividades específicas.
            #
            # IMPORTANTE — las dos columnas del plan de estudios (Res. 3003-E)
            # son requisitos de MOMENTOS DISTINTOS:
            #   tipo 'cursada'  = "Regularizadas para cursar"  → bloquea INSCRIBIRSE
            #   tipo 'aprobada' = "Aprobadas para rendir"      → bloquea la MESA DE EXAMEN
            #
            # Por eso acá solo se controlan las de tipo 'cursada'. Las de tipo
            # 'aprobada' se verifican al inscribir a mesa, no al inscribir a
            # la cursada. Exigirlas acá impedía que un alumno cursara una
            # materia adeudando finales, que es un caso normal del régimen.
            for req_id, tipo in correl_map.get(mid, []):
                req_nombre = materia_nombres.get(req_id, f'Materia {req_id}')
                if tipo == 'cursada' and req_id not in cursadas_ok:
                    puede = False
                    bloqueada_por.append(f"Regularizar: {req_nombre}")

        resultado.append({
            'id': mid,
            'nombre': nombre,
            'anio': anio_m,
            'orden': orden,
            'regimen': regimen,
            'regimen_aprobacion': reg_aprobacion,
            'inscripta': inscripta,
            'puede_inscribirse': puede,
            'bloqueada_por': bloqueada_por,
        })

    # ── Estado de bloqueo (post-guardado) ──
    # ¿Ya tiene inscripciones guardadas este ciclo lectivo?
    tiene_inscripciones_guardadas = any(m['inscripta'] for m in resultado)

    # ¿El coordinador autorizó una reapertura vigente (no consumida)?
    # Una autorización se "consume" cuando se guardan cambios — luego vuelve a estar bloqueado.
    rol = session.get('rol')
    autorizado_hoy = False
    if tiene_inscripciones_guardadas:
        conn2 = get_db()
        c2 = conn2.cursor()
        c2.execute("""
            SELECT id FROM inscripciones_auditoria
            WHERE alumno_id = %s
              AND anio_lectivo = %s
              AND accion = 'reapertura_alumno'
              AND fecha::date = CURRENT_DATE
              AND detalle IS NOT NULL AND detalle->>'consumida_en' IS NULL
            ORDER BY fecha DESC LIMIT 1
        """, (aid, anio))
        autorizado_hoy = c2.fetchone() is not None
        c2.close(); conn2.close()

    # ── Estado de la ventana de inscripciones ──
    ventana = get_estado_inscripciones()

    # Determinar si el panel debe estar en modo solo-lectura
    # Modo edición permitido si:
    #  - la ventana está abierta
    #  - Y (el alumno no tiene inscripciones aún, O el coordinador autorizó reapertura hoy)
    puede_editar = ventana['abierto'] and (not tiene_inscripciones_guardadas or autorizado_hoy)

    if not ventana['abierto']:
        motivo_bloqueo = f'Inscripciones cerradas: {ventana["motivo"]}.'
    elif tiene_inscripciones_guardadas and not autorizado_hoy:
        motivo_bloqueo = (f'El alumno ya tiene inscripciones guardadas para el ciclo lectivo '
                          f'{anio}. Solo el coordinador puede autorizar una modificación '
                          f'registrando el motivo.')
    else:
        motivo_bloqueo = ''

    return jsonify({
        'alumno': {'id': alumno[0], 'apellido': alumno[1], 'nombre': alumno[2],
                   'dni': formatear_documento(alumno[4], alumno[3]),
                   'tipo_documento': alumno[4] or 'DNI',
                   'tipo_doc_label':  etiqueta_documento(alumno[4])},
        'anio_lectivo': anio,
        'materias': resultado,
        'tiene_inscripciones_guardadas': tiene_inscripciones_guardadas,
        'puede_editar':   puede_editar,
        'motivo_bloqueo': motivo_bloqueo,
        'ventana':        ventana,
        'es_coordinador': (rol == 'coordinador'),
    })


@auth.route('/api/inscripciones/guardar/<int:aid>', methods=['POST'])
@login_requerido(['coordinador', 'preceptora'])
def api_inscripciones_guardar(aid):
    """
    Recibe lista de materia_ids seleccionados.

    REGLAS DE NEGOCIO (Fase 1):
    - Solo se puede guardar si la ventana de inscripciones está abierta
      (validada con date.today() del servidor + configuración en DB)
    - Si el alumno ya tiene inscripciones guardadas para el ciclo lectivo,
      la modificación queda BLOQUEADA, salvo que el coordinador haya
      autorizado previamente la reapertura (registro en inscripciones_auditoria
      con accion='reapertura_alumno' del día actual).
    - Las inscripciones con cursada cargada nunca se borran (integridad
      del historial académico).
    """
    rol        = session.get('rol')
    carrera_id = session.get('carrera_id')
    user_id    = session.get('user_id')
    data       = request.get_json()
    ids_nuevos = set(data.get('materia_ids', []))

    # ── 1. Validar ventana de inscripciones (excepto coordinador con autorización) ──
    estado_ventana = get_estado_inscripciones()
    if not estado_ventana['abierto']:
        return jsonify({
            'error': f'Inscripciones cerradas. {estado_ventana["motivo"]}.',
            'ventana_cerrada': True
        }), 403

    conn = get_db()
    cur = conn.cursor()

    # ── 2. Verificar alumno ──
    cur.execute("SELECT id FROM alumnos WHERE id = %s AND carrera_id = %s AND activo = TRUE",
                (aid, carrera_id))
    if not cur.fetchone():
        cur.close(); conn.close()
        return jsonify({'error': 'Alumno no encontrado o inactivo'}), 404

    # ── 3. Año lectivo actual ──
    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])

    # ── 4. Inscripciones actuales del alumno en este ciclo ──
    cur.execute("""
        SELECT id, materia_id FROM inscripciones
        WHERE alumno_id = %s AND anio_lectivo = %s
    """, (aid, anio))
    actuales = {r[1]: r[0] for r in cur.fetchall()}  # materia_id → inscripcion_id
    ids_actuales = set(actuales.keys())

    # ── 5. Bloqueo post-guardado ──
    # Si ya tiene inscripciones en este ciclo lectivo, requiere autorización
    # del coordinador (registrada en inscripciones_auditoria) para modificar.
    # La autorización es de UN SOLO USO: se consume al guardar.
    autorizacion_id = None
    if ids_actuales:
        cur.execute("""
            SELECT id FROM inscripciones_auditoria
            WHERE alumno_id = %s
              AND anio_lectivo = %s
              AND accion = 'reapertura_alumno'
              AND fecha::date = CURRENT_DATE
              AND detalle IS NOT NULL AND detalle->>'consumida_en' IS NULL
            ORDER BY fecha DESC LIMIT 1
        """, (aid, anio))
        row = cur.fetchone()
        autorizacion_id = row[0] if row else None

        if autorizacion_id is None:
            cur.close(); conn.close()
            return jsonify({
                'error': 'Este alumno ya tiene inscripciones guardadas para el ciclo lectivo '
                         f'{anio}. Para modificarlas, el coordinador debe autorizar la '
                         'reapertura desde su panel registrando un motivo.',
                'requiere_autorizacion': True
            }), 403

    # ── 6. Inscribir las nuevas ──
    inscriptas = 0
    for mid in ids_nuevos - ids_actuales:
        cur.execute("""
            INSERT INTO inscripciones (alumno_id, materia_id, anio_lectivo)
            VALUES (%s, %s, %s) ON CONFLICT DO NOTHING
        """, (aid, mid, anio))
        inscriptas += 1

    # ── 7. Desinscribir las quitadas (solo si no tienen cursada cargada) ──
    desinscriptas = 0
    no_borradas = 0
    for mid in ids_actuales - ids_nuevos:
        insc_id = actuales[mid]
        cur.execute("SELECT id FROM cursadas WHERE inscripcion_id = %s", (insc_id,))
        tiene_cursada = cur.fetchone()
        if tiene_cursada:
            no_borradas += 1
        else:
            cur.execute("DELETE FROM inscripciones WHERE id = %s", (insc_id,))
            desinscriptas += 1

    # ── 8. Consumir la autorización (si se usó una) ──
    # Marca la autorización como ya usada SIEMPRE que se haya usado para guardar,
    # incluso si no hubo cambios efectivos. La autorización es de un solo uso:
    # apretar Guardar la consume sí o sí.
    if autorizacion_id is not None:
        cur.execute("""
            UPDATE inscripciones_auditoria
            SET detalle = COALESCE(detalle, '{}'::jsonb) ||
                          jsonb_build_object(
                              'consumida_en',         NOW()::text,
                              'inscriptas_agregadas', %s,
                              'inscriptas_removidas', %s
                          )
            WHERE id = %s
        """, (inscriptas, desinscriptas, autorizacion_id))

    conn.commit()
    cur.close()
    conn.close()

    # Singular/plural
    def plural(n, sing, plur):
        return f"{n} {sing if n == 1 else plur}"

    partes = []
    if inscriptas:
        partes.append(plural(inscriptas, 'inscripción nueva', 'inscripciones nuevas'))
    if desinscriptas:
        partes.append(plural(desinscriptas, 'inscripción removida', 'inscripciones removidas'))
    if not partes:
        msg = "No hubo cambios en la inscripción."
    else:
        msg = "Guardado: " + ", ".join(partes) + "."
    if no_borradas:
        msg += f" {plural(no_borradas, 'materia no se pudo quitar', 'materias no se pudieron quitar')} porque ya tiene notas cargadas."

    return jsonify({'ok': True, 'mensaje': msg, 'advertencia': no_borradas > 0})


# ================================================================
# API — VENTANA DE INSCRIPCIONES (configuración + cierre/apertura manual)
# ================================================================

@auth.route('/api/inscripciones/ventana', methods=['GET'])
@login_requerido(['coordinador', 'preceptora', 'admin'])
def api_inscripciones_ventana_get():
    """Devuelve el estado actual de la ventana de inscripciones."""
    return jsonify(get_estado_inscripciones())


@auth.route('/api/inscripciones/ventana', methods=['POST'])
@login_requerido(['coordinador'])
def api_inscripciones_ventana_set():
    """
    Actualiza las fechas de la ventana de inscripciones.
    Solo coordinador. Queda registrado en auditoría con motivo.
    """
    carrera_id = session.get('carrera_id')
    user_id    = session.get('user_id')
    data       = request.get_json()

    fecha_inicio = (data.get('fecha_inicio') or '').strip()
    fecha_fin    = (data.get('fecha_fin')    or '').strip()
    motivo       = (data.get('motivo')       or '').strip()

    if not fecha_inicio or not fecha_fin:
        return jsonify({'error': 'Las fechas de apertura y cierre son obligatorias'}), 400

    try:
        finicio = date.fromisoformat(fecha_inicio)
        ffin    = date.fromisoformat(fecha_fin)
    except Exception:
        return jsonify({'error': 'Formato de fecha inválido (debe ser AAAA-MM-DD)'}), 400

    if ffin < finicio:
        return jsonify({'error': 'La fecha de cierre no puede ser anterior a la de apertura'}), 400

    if not motivo or len(motivo) < 5:
        return jsonify({'error': 'El motivo es obligatorio (mínimo 5 caracteres)'}), 400

    conn = get_db()
    cur  = conn.cursor()
    try:
        cur.execute("""INSERT INTO configuracion (clave, valor) VALUES ('inscripciones_fecha_inicio', %s)
                       ON CONFLICT (clave) DO UPDATE SET valor = EXCLUDED.valor""",
                    (finicio.isoformat(),))
        cur.execute("""INSERT INTO configuracion (clave, valor) VALUES ('inscripciones_fecha_fin', %s)
                       ON CONFLICT (clave) DO UPDATE SET valor = EXCLUDED.valor""",
                    (ffin.isoformat(),))

        cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
        anio = int(cur.fetchone()[0])

        cur.execute("""
            INSERT INTO inscripciones_auditoria
                (alumno_id, coordinador_id, accion, motivo, detalle, anio_lectivo)
            VALUES (NULL, %s, 'modificacion_fechas', %s, %s::jsonb, %s)
        """, (user_id, motivo,
              json.dumps({'fecha_inicio': finicio.isoformat(), 'fecha_fin': ffin.isoformat()}),
              anio))
        # Nota: alumno_id NULL para este tipo de acción (es global, no para un alumno)

        conn.commit()
    except Exception as e:
        conn.rollback()
        # Si falla por NOT NULL en alumno_id, reintentamos sin auditoría individual
        # (la tabla tiene NOT NULL en alumno_id por diseño)
        cur.close(); conn.close()
        return jsonify({'error': f'Error al guardar: {str(e)}'}), 500
    finally:
        if not conn.closed:
            cur.close()
            conn.close()

    return jsonify({'ok': True, 'mensaje': 'Fechas de inscripciones actualizadas correctamente'})


@auth.route('/api/inscripciones/cerrar-manual', methods=['POST'])
@login_requerido(['coordinador'])
def api_inscripciones_cerrar_manual():
    """Cierra manualmente las inscripciones antes de la fecha de cierre."""
    user_id = session.get('user_id')
    data    = request.get_json()
    motivo  = (data.get('motivo') or '').strip()

    if not motivo or len(motivo) < 5:
        return jsonify({'error': 'El motivo es obligatorio (mínimo 5 caracteres)'}), 400

    conn = get_db()
    cur  = conn.cursor()
    cur.execute("UPDATE configuracion SET valor = 'true' WHERE clave = 'inscripciones_cerrado_manual'")
    cur.execute("UPDATE configuracion SET valor = %s WHERE clave = 'inscripciones_motivo_cierre'",
                (motivo,))
    conn.commit()
    cur.close(); conn.close()

    return jsonify({'ok': True, 'mensaje': 'Inscripciones cerradas manualmente'})


@auth.route('/api/inscripciones/reabrir-manual', methods=['POST'])
@login_requerido(['coordinador'])
def api_inscripciones_reabrir_manual():
    """Reabre las inscripciones (saca el cierre manual)."""
    user_id = session.get('user_id')
    data    = request.get_json()
    motivo  = (data.get('motivo') or '').strip()

    if not motivo or len(motivo) < 5:
        return jsonify({'error': 'El motivo es obligatorio (mínimo 5 caracteres)'}), 400

    conn = get_db()
    cur  = conn.cursor()
    cur.execute("UPDATE configuracion SET valor = 'false' WHERE clave = 'inscripciones_cerrado_manual'")
    cur.execute("UPDATE configuracion SET valor = %s WHERE clave = 'inscripciones_motivo_cierre'",
                (f'(Reapertura) {motivo}',))
    conn.commit()
    cur.close(); conn.close()

    return jsonify({'ok': True, 'mensaje': 'Inscripciones reabiertas'})


@auth.route('/api/inscripciones/autorizar-reapertura/<int:aid>', methods=['POST'])
@login_requerido(['coordinador'])
def api_inscripciones_autorizar_reapertura(aid):
    """
    Autoriza al coordinador a modificar la inscripción de un alumno específico
    que ya tiene inscripciones guardadas para el ciclo lectivo actual.
    Registra el motivo en inscripciones_auditoria.
    La autorización vale por el día calendario actual.
    """
    user_id    = session.get('user_id')
    carrera_id = session.get('carrera_id')
    data       = request.get_json()
    motivo     = (data.get('motivo') or '').strip()

    if not motivo or len(motivo) < 5:
        return jsonify({'error': 'El motivo es obligatorio (mínimo 5 caracteres)'}), 400

    conn = get_db()
    cur  = conn.cursor()

    cur.execute("SELECT id FROM alumnos WHERE id = %s AND carrera_id = %s", (aid, carrera_id))
    if not cur.fetchone():
        cur.close(); conn.close()
        return jsonify({'error': 'Alumno no encontrado'}), 404

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])

    cur.execute("""
        INSERT INTO inscripciones_auditoria
            (alumno_id, coordinador_id, accion, motivo, detalle, anio_lectivo)
        VALUES (%s, %s, 'reapertura_alumno', %s, '{}'::jsonb, %s)
    """, (aid, user_id, motivo, anio))
    conn.commit()
    cur.close(); conn.close()

    return jsonify({'ok': True, 'mensaje': 'Reapertura autorizada. Ya podés modificar las inscripciones de este alumno (una sola vez).'})


@auth.route('/api/inscripciones/cancelar-autorizacion/<int:aid>', methods=['POST'])
@login_requerido(['coordinador'])
def api_inscripciones_cancelar_autorizacion(aid):
    """
    Consume cualquier autorización vigente (no consumida) del alumno para HOY.
    La autorización queda marcada con origen='cancelada' (no se usó para modificar).
    Usado por:
    - Botón "✖ Cancelar modificación" del panel
    - Modal de guardia al salir sin guardar
    - Beacon de beforeunload (cierre de pestaña)
    Idempotente: si no hay autorización vigente, devuelve ok sin error.
    """
    user_id    = session.get('user_id')
    carrera_id = session.get('carrera_id')
    data       = request.get_json(silent=True) or {}
    origen     = (data.get('origen') or 'cancelada').strip()

    conn = get_db()
    cur  = conn.cursor()

    cur.execute("SELECT id FROM alumnos WHERE id = %s AND carrera_id = %s", (aid, carrera_id))
    if not cur.fetchone():
        cur.close(); conn.close()
        return jsonify({'error': 'Alumno no encontrado'}), 404

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])

    cur.execute("""
        UPDATE inscripciones_auditoria
        SET detalle = COALESCE(detalle, '{}'::jsonb) ||
                      jsonb_build_object(
                          'consumida_en', NOW()::text,
                          'origen',       %s
                      )
        WHERE alumno_id = %s
          AND anio_lectivo = %s
          AND accion = 'reapertura_alumno'
          AND fecha::date = CURRENT_DATE
          AND detalle IS NOT NULL
          AND detalle->>'consumida_en' IS NULL
        RETURNING id
    """, (origen, aid, anio))
    afectadas = cur.fetchall()
    conn.commit()
    cur.close(); conn.close()

    return jsonify({
        'ok': True,
        'mensaje': 'Modificación cancelada' if afectadas else 'No había autorización vigente',
        'consumidas': len(afectadas)
    })


# ================================================================
# API — LISTA DE ALUMNOS POR AÑO DEL PLAN (para módulo Inscripciones)
# ================================================================

@auth.route('/api/inscripciones/anios-plan', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_inscripciones_anios_plan():
    """
    Devuelve los años distintos del plan de estudios vigente de la carrera.
    Esto alimenta los tabs dinámicos del frontend.
    """
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("""
        SELECT DISTINCT anio FROM materias
        WHERE carrera_id = %s AND activa = TRUE
        ORDER BY anio
    """, (carrera_id,))
    anios = [r[0] for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify({'anios': anios})


@auth.route('/api/inscripciones/alumnos-por-anio/<int:anio_plan>', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_inscripciones_alumnos_por_anio(anio_plan):
    """
    Devuelve los alumnos activos de la carrera que están EN CONDICIONES de
    inscribirse a materias del año `anio_plan` según las reglas:

    - 1° año: todos los alumnos activos
    - Año X > 1: alumnos con al menos una materia regularizada/aprobada de X-1
      (misma lógica de habilitación que ya usa /api/inscripciones/alumno)

    Para cada alumno indica también si ya tiene inscripciones guardadas en
    el ciclo lectivo actual (badge 'Inscripto' vs 'Pendiente').
    """
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur  = conn.cursor()

    # Año lectivo actual
    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio_lectivo = int(cur.fetchone()[0])

    # Traer todos los alumnos activos
    cur.execute("""
        SELECT a.id, a.apellido, a.nombre, a.dni, a.tipo_documento
        FROM alumnos a
        WHERE a.carrera_id = %s AND a.activo = TRUE
        ORDER BY a.apellido, a.nombre
    """, (carrera_id,))
    alumnos = cur.fetchall()

    if not alumnos:
        cur.close(); conn.close()
        return jsonify({'alumnos': [], 'anio_plan': anio_plan, 'anio_lectivo': anio_lectivo})

    # Si es 1° año → todos están habilitados
    if anio_plan == 1:
        habilitados_ids = {a[0] for a in alumnos}
    else:
        # Habilitados: alumnos con al menos una materia del año anio_plan-1
        # regularizada o aprobada
        cur.execute("""
            SELECT DISTINCT a.id
            FROM alumnos a
            WHERE a.carrera_id = %s AND a.activo = TRUE
              AND (
                EXISTS (
                    SELECT 1 FROM inscripciones i
                    JOIN cursadas cu ON cu.inscripcion_id = i.id
                    JOIN materias m  ON m.id = i.materia_id
                    WHERE i.alumno_id = a.id
                      AND m.carrera_id = %s
                      AND m.anio = %s
                      AND cu.condicion IN ('regular', 'promocionado')
                )
                OR EXISTS (
                    SELECT 1 FROM inscripciones i
                    JOIN examenes ex ON ex.inscripcion_id = i.id
                    JOIN materias m  ON m.id = i.materia_id
                    WHERE i.alumno_id = a.id
                      AND m.carrera_id = %s
                      AND m.anio = %s
                      AND ex.aprobado = TRUE
                )
              )
        """, (carrera_id, carrera_id, anio_plan - 1, carrera_id, anio_plan - 1))
        habilitados_ids = {r[0] for r in cur.fetchall()}

    # ¿Quiénes ya tienen alguna inscripción guardada este ciclo lectivo
    # para una materia del año `anio_plan`?
    cur.execute("""
        SELECT DISTINCT i.alumno_id
        FROM inscripciones i
        JOIN materias m ON m.id = i.materia_id
        WHERE m.carrera_id = %s AND m.anio = %s AND i.anio_lectivo = %s
    """, (carrera_id, anio_plan, anio_lectivo))
    inscriptos_ids = {r[0] for r in cur.fetchall()}

    # ¿Quiénes ya tienen TODAS las materias del año aprobadas/regularizadas?
    # (esos también los marcamos para diferenciarlos)
    cur.execute("""
        SELECT COUNT(*) FROM materias
        WHERE carrera_id = %s AND anio = %s AND activa = TRUE
    """, (carrera_id, anio_plan))
    total_materias_anio = cur.fetchone()[0]

    cur.close(); conn.close()

    resultado = []
    for row in alumnos:
        aid, apellido, nombre, dni, tipo_doc = row
        if aid not in habilitados_ids:
            continue
        resultado.append({
            'id':        aid,
            'apellido':  apellido,
            'nombre':    nombre,
            'dni':       formatear_documento(tipo_doc, dni),
            'tipo_doc_label': etiqueta_documento(tipo_doc),
            'inscripto': aid in inscriptos_ids,
        })

    return jsonify({
        'alumnos':              resultado,
        'anio_plan':            anio_plan,
        'anio_lectivo':         anio_lectivo,
        'total_materias_anio':  total_materias_anio,
    })


# ================================================================
# API — NOTAS (coordinador + preceptora)
# ================================================================

# Diccionario de autocorrección para condiciones
DICCIONARIO_CONDICION = {
    'regular': 'regular', 'reguler': 'regular', 'regualar': 'regular',
    'regularr': 'regular', 'rregular': 'regular', 'regualr': 'regular',
    'regularizo': 'regular', 'regularizó': 'regular',
    'promocionado': 'promocionado', 'promocionada': 'promocionado',
    'promociondo': 'promocionado', 'promociono': 'promocionado',
    'promovido': 'promocionado', 'promocion': 'promocionado', 'promo': 'promocionado',
    'promoción': 'promocionado', 'promocionó': 'promocionado',
    'libre': 'libre', 'libres': 'libre',
    'ausente': 'libre', 'ausentes': 'libre', 'aus': 'libre',
}

DICCIONARIO_NOTA = {
    'uno': 1, 'dos': 2, 'tres': 3, 'cuatro': 4, 'cinco': 5,
    'seis': 6, 'siete': 7, 'ocho': 8, 'nueve': 9, 'diez': 10,
}

def normalizar_condicion(valor):
    if not valor:
        return None
    v = str(valor).strip().lower()
    return DICCIONARIO_CONDICION.get(v, None)

def normalizar_nota(valor):
    if valor is None or str(valor).strip() in ('', '-'):
        return None
    v = str(valor).strip().lower()
    if v in DICCIONARIO_NOTA:
        return float(DICCIONARIO_NOTA[v])
    v = v.replace('%', '').strip()
    try:
        n = float(v)
        if 1 <= n <= 10:
            return n
        return None
    except:
        return None

def normalizar_porcentaje(valor):
    if valor is None or str(valor).strip() in ('', '-'):
        return None
    v = str(valor).strip().replace('%', '').strip()
    try:
        n = float(v)
        if 0 < n <= 1:
            n = n * 100
        if 0 <= n <= 100:
            return round(n, 2)
        return None
    except:
        return None

def admite_promocion(regimen_aprobacion):
    """
    ¿La materia se puede aprobar por promoción?
    Lee el texto del plan de estudios ("Promoción / Examen Final",
    "Examen Final", "Promoción"). Si el dato falta, se asume que sí,
    para no bloquear materias cargadas antes de tener este campo.
    """
    if not regimen_aprobacion:
        return True
    return 'promoc' in str(regimen_aprobacion).lower()


def admite_examen_final(regimen_aprobacion):
    """
    ¿La materia se rinde en mesa de examen?
    Las de régimen sólo "Promoción" (prácticas profesionalizantes,
    EDI) no van a mesa: si el alumno no promociona, recursa.
    """
    if not regimen_aprobacion:
        return True
    return 'final' in str(regimen_aprobacion).lower()


def sugerir_condicion(nota, regimen_aprobacion=None):
    """
    Sugerencia automática de condición según la nota.
    Si la materia no admite promoción (régimen sólo Examen Final),
    el tope sugerido es 'regular' por más alta que sea la nota.
    La condición final siempre la decide el profesor.
    """
    if nota is None:
        return None
    if nota >= 7:
        return 'promocionado' if admite_promocion(regimen_aprobacion) else 'regular'
    elif nota >= 4:
        return 'regular'
    else:
        return 'libre'

def nota_en_letras(nota):
    letras = {1:'UNO',2:'DOS',3:'TRES',4:'CUATRO',5:'CINCO',
              6:'SEIS',7:'SIETE',8:'OCHO',9:'NUEVE',10:'DIEZ'}
    if nota is None:
        return '-'
    try:
        return letras.get(int(nota), str(int(nota)))
    except:
        return '-'


@auth.route('/api/notas/materias', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_notas_materias():
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])

    cur.execute("""
        SELECT
            m.id, m.nombre, m.anio, m.orden, m.regimen,
            COUNT(DISTINCT i.id) AS inscriptos,
            COUNT(DISTINCT cu.id) AS con_notas,
            BOOL_OR(cu.cerrada) AS cerrada
        FROM materias m
        LEFT JOIN inscripciones i ON i.materia_id = m.id AND i.anio_lectivo = %s
        LEFT JOIN cursadas cu ON cu.inscripcion_id = i.id
        WHERE m.carrera_id = %s AND m.activa = TRUE
        GROUP BY m.id, m.nombre, m.anio, m.orden, m.regimen
        ORDER BY m.anio, m.orden
    """, (anio, carrera_id))

    rows = cur.fetchall()
    cur.close()
    conn.close()

    anios = {}
    for r in rows:
        anio_m = r[2]
        if anio_m not in anios:
            anios[anio_m] = []
        anios[anio_m].append({
            'id': r[0], 'nombre': r[1], 'anio': r[2], 'orden': r[3],
            'regimen': r[4], 'inscriptos': r[5], 'con_notas': r[6],
            'cerrada': r[7] or False
        })

    return jsonify({'anio_lectivo': anio, 'anios': anios})


@auth.route('/api/notas/materia/<int:mid>', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_notas_materia_detalle(mid):
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])

    cur.execute("""
        SELECT id, nombre, anio, orden, regimen, regimen_aprobacion
        FROM materias WHERE id = %s AND carrera_id = %s
    """, (mid, carrera_id))
    materia = cur.fetchone()
    if not materia:
        cur.close(); conn.close()
        return jsonify({'error': 'Materia no encontrada'}), 404

    cur.execute("""
        SELECT
            a.id, a.apellido, a.nombre, a.dni,
            i.id AS inscripcion_id,
            cu.id AS cursada_id,
            cu.nota_cursada, cu.porcentaje_asistencia, cu.porcentaje_tp,
            cu.condicion, cu.cerrada, cu.observaciones,
            COALESCE(cu.promocion_provisoria, FALSE)
        FROM inscripciones i
        JOIN alumnos a ON a.id = i.alumno_id
        LEFT JOIN cursadas cu ON cu.inscripcion_id = i.id
        WHERE i.materia_id = %s AND i.anio_lectivo = %s AND a.carrera_id = %s
        ORDER BY a.apellido, a.nombre
    """, (mid, anio, carrera_id))

    alumnos = []
    cerrada = False
    for r in cur.fetchall():
        nota = float(r[6]) if r[6] is not None else None
        if r[10]:
            cerrada = True
        alumnos.append({
            'alumno_id': r[0], 'apellido': r[1], 'nombre': r[2],
            'dni': formatear_dni(r[3]),
            'inscripcion_id': r[4], 'cursada_id': r[5],
            'nota_cursada': nota,
            'nota_letras': nota_en_letras(nota),
            'porcentaje_asistencia': float(r[7]) if r[7] is not None else None,
            'porcentaje_tp': float(r[8]) if r[8] is not None else None,
            'condicion': r[9], 'cerrada': r[10] or False,
            'observaciones': r[11],
            'promocion_provisoria': bool(r[12]),
            'condicion_sugerida': sugerir_condicion(nota, materia[5]),
        })

    cur.close()
    conn.close()

    return jsonify({
        'materia': {
            'id': materia[0], 'nombre': materia[1], 'anio': materia[2],
            'orden': materia[3], 'regimen': materia[4], 'regimen_aprobacion': materia[5]
        },
        'anio_lectivo': anio,
        'cerrada': cerrada,
        'alumnos': alumnos,
        'totales': {
            'inscriptos':    len(alumnos),
            'promocionados': sum(1 for a in alumnos if a['condicion'] == 'promocionado'),
            'regulares':     sum(1 for a in alumnos if a['condicion'] == 'regular'),
            # "Libre" agrupa: condición libre explícita, el viejo valor "ausente" (ya no
            # existe como categoría propia) y los que quedaron sin condición registrada
            # al momento del cierre — así Promocionados+Regulares+Libres = inscriptos.
            'libres':        sum(1 for a in alumnos if a['condicion'] not in ('promocionado', 'regular')),
        }
    })


@auth.route('/api/notas/guardar/<int:mid>', methods=['POST'])
@login_requerido(['coordinador', 'preceptora'])
def api_notas_guardar(mid):
    carrera_id = session.get('carrera_id')
    data = request.get_json()
    filas = data.get('filas', [])

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""SELECT id, regimen_aprobacion FROM materias
                   WHERE id = %s AND carrera_id = %s""", (mid, carrera_id))
    materia_row = cur.fetchone()
    if not materia_row:
        cur.close(); conn.close()
        return jsonify({'error': 'Materia no encontrada'}), 404

    regimen_aprob = materia_row[1]
    puede_promocionar = admite_promocion(regimen_aprob)

    # ── Correlativas que hay que tener APROBADAS para que la promoción
    #    de esta materia sea válida (art. del plan de estudios) ──
    cur.execute("""
        SELECT co.requiere_materia_id, m.nombre, m.orden
        FROM correlatividades co
        JOIN materias m ON m.id = co.requiere_materia_id
        WHERE co.materia_id = %s AND co.tipo = 'aprobada'
    """, (mid,))
    requeridas_aprob = cur.fetchall()

    # Fecha límite del ciclo lectivo para confirmar promociones.
    _hoy = date.today()
    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    _row_anio = cur.fetchone()
    _anio_cfg = int(_row_anio[0]) if _row_anio else _hoy.year
    _limite_promo = _fecha_limite_promocion(cur, _anio_cfg)

    guardados = 0
    avisos    = []
    for fila in filas:
        insc_id   = fila.get('inscripcion_id')
        nota      = fila.get('nota_cursada')
        asist     = fila.get('porcentaje_asistencia')
        tp        = fila.get('porcentaje_tp')
        condicion = fila.get('condicion')
        obs       = fila.get('observaciones', '')

        nota      = float(nota) if nota is not None and str(nota).strip() != '' else None
        asist     = float(asist) if asist is not None and str(asist).strip() != '' else None
        tp        = float(tp) if tp is not None and str(tp).strip() != '' else None
        condicion = condicion if condicion in ['regular', 'libre', 'promocionado', 'ausente'] else None

        # Se descartan las marcas automaticas que hayan quedado de un
        # guardado anterior. Si la regla sigue aplicando, se vuelve a
        # escribir mas abajo; si no, desaparece sola.
        obs = _limpiar_marcas_automaticas(obs)

        # ── REGLA 1: el plan manda sobre la promoción ──
        # Si la materia se aprueba sólo con examen final, no se puede
        # promocionar por más alta que sea la nota.
        if condicion == 'promocionado' and not puede_promocionar:
            condicion = 'regular'
            avisos.append(
                f"La materia no admite promoción según el plan "
                f"({regimen_aprob}). Se guardó como regular."
            )

        # ── REGLA 2: la promoción cae si adeuda la correlativa ──
        # Regla institucional: el alumno que promociona una materia y
        # adeuda el final de su correlativa del año anterior tiene hasta
        # la fecha límite del ciclo (por defecto 31/12) para aprobarlo.
        #
        #   • Antes de la fecha límite → la promoción se guarda, pero
        #     marcada como PROVISORIA. Todavía le quedan mesas.
        #   • Pasada la fecha límite   → la promoción se cae y la nota
        #     se asienta directamente como 'regular'.
        #
        # La nota numérica del profesor NO se toca en ningún caso; lo
        # único que cambia es la condición.
        _promo_prov = False
        if condicion == 'promocionado' and requeridas_aprob:
            cur.execute("""
                SELECT alumno_id FROM inscripciones WHERE id = %s
            """, (insc_id,))
            _row_al = cur.fetchone()
            if _row_al:
                _alumno_id = _row_al[0]
                # Una promoción provisoria NO cuenta como materia aprobada:
                # si contara, serviría para validar la promoción de la
                # materia siguiente y el error se propagaría en cascada.
                cur.execute("""
                    SELECT materia_id FROM (
                        SELECT i.materia_id
                        FROM inscripciones i
                        JOIN cursadas cu ON cu.inscripcion_id = i.id
                        WHERE i.alumno_id = %s
                          AND cu.condicion IN ('promocionado', 'aprobado')
                          AND NOT cu.promocion_provisoria
                        UNION
                        SELECT materia_id FROM examenes
                        WHERE alumno_id = %s AND resultado = 'aprobado'
                    ) sub
                """, (_alumno_id, _alumno_id))
                _aprobadas = {r[0] for r in cur.fetchall()}
                _faltan = [f"({o}) {n}" for rid, n, o in requeridas_aprob
                           if rid not in _aprobadas]
                if _faltan:
                    if _hoy <= _limite_promo:
                        _promo_prov = True
                        _nota_cond = (
                            "PROMOCIÓN PROVISORIA — se confirma solo si aprueba "
                            "el final de: " + ", ".join(_faltan) +
                            f" (plazo: {_limite_promo.strftime('%d/%m/%Y')})"
                        )
                    else:
                        condicion = 'regular'
                        _nota_cond = (
                            "PROMOCIÓN CAÍDA — venció el plazo del "
                            f"{_limite_promo.strftime('%d/%m/%Y')} sin aprobar "
                            "el final de: " + ", ".join(_faltan) +
                            ". Se asienta como regular."
                        )
                    obs = (obs + " | " + _nota_cond).strip(" |") if obs else _nota_cond
                    avisos.append(_nota_cond)

        cur.execute("""
            SELECT cu.id, cu.cerrada FROM inscripciones i
            LEFT JOIN cursadas cu ON cu.inscripcion_id = i.id
            WHERE i.id = %s
        """, (insc_id,))
        row = cur.fetchone()
        if not row:
            continue

        cursada_id, cerrada = row
        if cerrada:
            continue

        if cursada_id:
            # Verificar si el alumno ya aprobó en mesa de examen para esta materia
            cur.execute("""
                SELECT 1 FROM examenes e
                JOIN inscripciones i ON i.alumno_id = e.alumno_id
                WHERE i.id = %s AND e.materia_id = i.materia_id
                  AND e.resultado = 'aprobado'
                LIMIT 1
            """, (insc_id,))
            if cur.fetchone():
                # Alumno ya aprobó en mesa — intocable, saltar
                continue

            # Verificar si tiene condicion aprobado/promocionado con libro y folio
            cur.execute("""
                SELECT 1 FROM cursadas
                WHERE id = %s
                  AND condicion IN ('aprobado','promocionado')
                  AND libro IS NOT NULL AND folio IS NOT NULL
            """, (cursada_id,))
            if cur.fetchone():
                # Datos oficiales completos — intocable, saltar
                continue

            cur.execute("""
                UPDATE cursadas SET
                    nota_cursada = %s, porcentaje_asistencia = %s,
                    porcentaje_tp = %s, condicion = %s, observaciones = %s,
                    promocion_provisoria = %s
                WHERE id = %s
            """, (nota, asist, tp, condicion, obs or None, _promo_prov, cursada_id))
        else:
            cur.execute("""
                INSERT INTO cursadas
                    (inscripcion_id, nota_cursada, porcentaje_asistencia,
                     porcentaje_tp, condicion, observaciones, promocion_provisoria)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (insc_id, nota, asist, tp, condicion, obs or None, _promo_prov))
        guardados += 1

    conn.commit()
    cur.close()
    conn.close()
    # avisos: mensajes para la preceptora sobre reglas del plan que se
    # aplicaron al guardar (promociones bajadas a regular, condicionadas, etc.)
    return jsonify({
        'ok': True,
        'guardados': guardados,
        'avisos': sorted(set(avisos))
    })


@auth.route('/api/notas/cerrar/<int:mid>', methods=['POST'])
@login_requerido(['coordinador', 'preceptora'])
def api_notas_cerrar(mid):
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])

    # ── Bloqueo: promociones provisorias sin resolver ──
    # Una cursada cerrada es inmodificable, así que si se cierra antes
    # de la fecha límite las promociones provisorias quedarían firmes
    # por accidente. Se cierra recién pasado el plazo, cuando el sistema
    # ya las resolvió.
    limite = _fecha_limite_promocion(cur, anio)
    if date.today() <= limite:
        cur.execute("""
            SELECT COUNT(*)
            FROM cursadas cu
            JOIN inscripciones i ON i.id = cu.inscripcion_id
            JOIN materias m ON m.id = i.materia_id
            WHERE i.materia_id = %s AND i.anio_lectivo = %s
              AND m.carrera_id = %s
              AND cu.promocion_provisoria
        """, (mid, anio, carrera_id))
        provisorias = cur.fetchone()[0]
        if provisorias:
            cur.close(); conn.close()
            return jsonify({'error':
                f'No se puede cerrar todavía: hay {provisorias} '
                f'{"promoción provisoria" if provisorias == 1 else "promociones provisorias"} '
                f'a la espera de que el alumno apruebe el final de su correlativa. '
                f'El cierre se habilita después del {limite.strftime("%d/%m/%Y")}.'}), 400

    cur.execute("""
        UPDATE cursadas SET cerrada = TRUE
        WHERE inscripcion_id IN (
            SELECT i.id FROM inscripciones i
            JOIN materias m ON m.id = i.materia_id
            WHERE i.materia_id = %s AND i.anio_lectivo = %s AND m.carrera_id = %s
        )
    """, (mid, anio, carrera_id))

    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})


# NOTA: La ruta /api/notas/reabrir fue eliminada por integridad del sistema.
# Las cursadas cerradas NO se pueden reabrir. Solo el coordinador puede
# hacer correcciones puntuales mediante /api/notas/corregir/<cursada_id>


@auth.route('/api/notas/corregir/<int:cursada_id>', methods=['POST'])
@login_requerido(['coordinador'])
def api_notas_corregir(cursada_id):
    """
    Corrección puntual de una nota en cursada cerrada.
    Solo el coordinador puede hacerlo.
    NO se permite si el alumno ya aprobó en mesa con libro/folio.
    Registra auditoría del cambio.
    """
    carrera_id = session.get('carrera_id')
    data = request.get_json()
    campo     = data.get('campo')    # 'nota_cursada' | 'condicion' | 'observaciones'
    valor     = data.get('valor')
    motivo    = data.get('motivo', '').strip()

    if not motivo:
        return jsonify({'error': 'Debe ingresar un motivo para la corrección'}), 400

    campos_permitidos = ['nota_cursada', 'porcentaje_asistencia', 'porcentaje_tp',
                         'condicion', 'observaciones']
    if campo not in campos_permitidos:
        return jsonify({'error': 'Campo no permitido'}), 400

    conn = get_db()
    cur  = conn.cursor()
    try:
        # Verificar que la cursada pertenece a la carrera del coordinador
        cur.execute("""
            SELECT cu.id, cu.cerrada, cu.condicion, cu.libro, cu.folio,
                   i.alumno_id, i.materia_id
            FROM cursadas cu
            JOIN inscripciones i ON i.id = cu.inscripcion_id
            JOIN materias m ON m.id = i.materia_id
            WHERE cu.id = %s AND m.carrera_id = %s
        """, (cursada_id, carrera_id))
        cu = cur.fetchone()
        if not cu:
            return jsonify({'error': 'Cursada no encontrada'}), 404

        _, cerrada, condicion_actual, libro, folio, alumno_id, materia_id = cu

        if not cerrada:
            return jsonify({'error': 'La cursada está abierta, editá directamente en Notas'}), 400

        # Bloqueo: alumno ya aprobó en mesa de examen
        cur.execute("""
            SELECT 1 FROM examenes
            WHERE alumno_id = %s AND materia_id = %s AND resultado = 'aprobado'
            LIMIT 1
        """, (alumno_id, materia_id))
        if cur.fetchone():
            return jsonify({'error': 'No se puede corregir: el alumno ya aprobó esta materia en mesa de examen'}), 400

        # Bloqueo: condicion aprobado/promocionado + libro + folio
        if condicion_actual in ('aprobado', 'promocionado') and libro and folio:
            return jsonify({'error': 'No se puede corregir: el alumno tiene datos oficiales completos (libro y folio registrados)'}), 400

        # Obtener valor anterior para auditoría
        cur.execute(f'SELECT {campo} FROM cursadas WHERE id = %s', (cursada_id,))
        valor_anterior = str(cur.fetchone()[0])

        # Aplicar corrección
        cur.execute(f'UPDATE cursadas SET {campo} = %s WHERE id = %s', (valor, cursada_id))

        # Registrar auditoría
        cur.execute("""
            INSERT INTO cursadas_auditoria
                (cursada_id, campo, valor_anterior, valor_nuevo,
                 modificado_por_id, motivo)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (cursada_id, campo, valor_anterior, str(valor),
                session['user_id'], motivo))

        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()



# ================================================================
# API — CICLO LECTIVO
# ================================================================

@auth.route('/api/ciclo-lectivo', methods=['GET'])
@login_requerido(['admin', 'coordinador', 'preceptora'])
def api_ciclo_lectivo():
    ciclo = get_ciclo_lectivo()
    return jsonify({
        'label':       ciclo['label'],
        'anio_inicio': ciclo['anio_inicio'],
        'anio_fin':    ciclo['anio_fin'],
        'vencido':     ciclo['vencido'],
    })


# ================================================================
# API — PENDIENTES LIBRO/FOLIO
# ================================================================

@auth.route('/api/pendientes-libro-folio', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_pendientes_libro_folio():
    carrera_id = session.get('carrera_id')
    pendientes = get_pendientes_libro_folio(carrera_id)
    return jsonify({'pendientes': pendientes, 'total': len(pendientes)})


@auth.route('/api/pendientes-libro-folio', methods=['POST'])
@login_requerido(['coordinador', 'preceptora'])
def api_guardar_libro_folio():
    """Guarda libro/folio de uno o varios alumnos promocionados."""
    data   = request.get_json()
    items  = data.get('items', [])  # [{cursada_id, libro, folio}, ...]

    conn = get_db()
    cur  = conn.cursor()
    try:
        for item in items:
            cursada_id = item.get('cursada_id')
            libro      = item.get('libro', '').strip() or None
            folio      = item.get('folio', '').strip() or None
            if not cursada_id:
                continue
            # Una promoción provisoria no puede recibir libro y folio:
            # todavía puede caerse. Asentarla sería un registro oficial
            # sobre una condición que no está firme.
            cur.execute("""
                SELECT 1 FROM cursadas
                WHERE id = %s AND promocion_provisoria
            """, (cursada_id,))
            if cur.fetchone():
                continue
            cur.execute("""
                UPDATE cursadas SET libro = %s, folio = %s
                WHERE id = %s
            """, (libro, folio, cursada_id))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@auth.route('/api/notas/descargar-plantilla/<int:mid>', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_notas_descargar_plantilla(mid):
    import os
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Border, Side

    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])

    cur.execute("""
        SELECT m.nombre, m.anio, m.orden, m.regimen, c.nombre AS carrera
        FROM materias m
        JOIN carreras c ON c.id = m.carrera_id
        WHERE m.id = %s AND m.carrera_id = %s
    """, (mid, carrera_id))
    materia = cur.fetchone()
    if not materia:
        cur.close(); conn.close()
        return jsonify({'error': 'Materia no encontrada'}), 404

    cur.execute("""
        SELECT p.nombre, p.apellido
        FROM materia_profesor mp
        JOIN profesores p ON p.id = mp.profesor_id
        WHERE mp.materia_id = %s AND mp.anio_lectivo = %s
    """, (mid, anio))
    profe_row = cur.fetchone()
    nombre_profe = f"{profe_row[1]}, {profe_row[0]}" if profe_row else ''

    cur.execute("""
        SELECT a.apellido, a.nombre, a.dni, i.id
        FROM inscripciones i
        JOIN alumnos a ON a.id = i.alumno_id
        WHERE i.materia_id = %s AND i.anio_lectivo = %s AND a.carrera_id = %s
        ORDER BY a.apellido, a.nombre
    """, (mid, anio, carrera_id))
    alumnos = cur.fetchall()
    cur.close()
    conn.close()

    nombre_materia = materia[0]
    regimen        = materia[3] or 'Anual'
    nombre_carrera = materia[4]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Notas'

    # ── ESTILOS ──
    font_titulo  = Font(name='Arial', bold=True, size=12)
    font_inst    = Font(name='Arial', bold=True, size=10)
    font_dir     = Font(name='Arial', size=9)
    font_bold    = Font(name='Arial', bold=True, size=11)
    font_normal  = Font(name='Arial', size=11)
    font_header  = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    font_warning = Font(name='Arial', italic=True, size=9, color='CC0000')
    fill_header  = PatternFill('solid', fgColor='1a4731')
    fill_gris    = PatternFill('solid', fgColor='F2F2F2')
    fill_amarillo= PatternFill('solid', fgColor='FFF9C4')
    fill_totales = PatternFill('solid', fgColor='E8F5E9')
    al_center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al_left      = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    al_right     = Alignment(horizontal='right',  vertical='center')
    borde = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    # ── ANCHOS DE COLUMNA ──
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 22

    # ── LOGO (fila 1-3, columna A) — recortado en su circunferencia ──
    from flask import current_app
    from io import BytesIO as _BytesIO
    logo_path = os.path.join(current_app.root_path, 'static', 'logo_ies9.png')
    if os.path.exists(logo_path):
        try:
            from PIL import Image as PILImage, ImageDraw
            pil_img = PILImage.open(logo_path).convert('RGBA')
            size = pil_img.size
            lado = min(size)
            left = (size[0] - lado) // 2
            top  = (size[1] - lado) // 2
            pil_img = pil_img.crop((left, top, left + lado, top + lado))
            margen = int(lado * 0.03)
            mascara = PILImage.new('L', (lado, lado), 0)
            ImageDraw.Draw(mascara).ellipse((margen, margen, lado - margen, lado - margen), fill=255)
            resultado = PILImage.new('RGBA', (lado, lado), (255, 255, 255, 0))
            resultado.paste(pil_img, mask=mascara)
            logo_buf = _BytesIO()
            resultado.save(logo_buf, format='PNG')
            logo_buf.seek(0)
            img = XLImage(logo_buf)
            img.width  = 85
            img.height = 85
            ws.add_image(img, 'A1')
        except Exception as e:
            print(f"[LOGO ERROR] {e}")

    # Filas 1-3: encabezado institucional (centrado en B:H)
    ws.merge_cells('B1:H1')
    ws['B1'] = 'INSTITUTO DE EDUCACIÓN SUPERIOR Nº 9 "JUANA AZURDUY"'
    ws['B1'].font = font_inst
    ws['B1'].alignment = al_center

    ws.merge_cells('B2:H2')
    ws['B2'] = 'Undiano esquina Silvester barrio Bernachi'
    ws['B2'].font = font_dir
    ws['B2'].alignment = al_center

    ws.merge_cells('B3:H3')
    ws['B3'] = 'San Pedro de Jujuy (4500) – Argentina'
    ws['B3'].font = font_dir
    ws['B3'].alignment = al_center

    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    # Fila 4 — Título planilla
    ws.merge_cells('A4:H4')
    ws['A4'] = f'Planilla de Calificaciones-Periodo Escolar {anio}'
    ws['A4'].font      = font_titulo
    ws['A4'].alignment = al_center
    ws.row_dimensions[4].height = 22

    # Fila 5 — Carrera + Año
    ws.merge_cells('A5:F5')
    ws['A5'] = f'Carrera: {nombre_carrera}'
    ws['A5'].font = font_bold; ws['A5'].alignment = al_left
    ws.merge_cells('G5:H5')
    ws['G5'] = f'Año: {anio}'
    ws['G5'].font = font_bold; ws['G5'].alignment = al_center

    # Fila 6 — Unidad Curricular + Régimen
    ws.merge_cells('A6:F6')
    ws['A6'] = f'Unidad Curricular: {nombre_materia}'
    ws['A6'].font = font_bold; ws['A6'].alignment = al_left
    ws.merge_cells('G6:H6')
    ws['G6'] = regimen
    ws['G6'].font = font_bold; ws['G6'].alignment = al_center

    # Fila 7 — Profesor/a + Curso y División (vacío para completar a mano)
    ws.merge_cells('A7:F7')
    ws['A7'] = f'Profesor/a: {nombre_profe}'
    ws['A7'].font = font_bold; ws['A7'].alignment = al_left
    ws.merge_cells('G7:H7')
    ws['G7'] = 'Curso y División:'
    ws['G7'].font = font_bold; ws['G7'].alignment = al_left

    # Fila 8 — Libro / Folio (vacíos) + ALUMNO (vacío para completar a mano)
    ws['A8'] = 'Libro:'
    ws['A8'].font = font_normal
    ws['B8'] = ''   # completar a mano
    ws['C8'] = 'Folio:'
    ws['C8'].font = font_normal
    ws['D8'] = ''   # completar a mano
    ws.merge_cells('G8:H8')
    ws['G8'] = 'Alumno:'
    ws['G8'].font = font_bold; ws['G8'].alignment = al_left

    # Fila 9 — Aviso
    ws.merge_cells('A9:H9')
    ws['A9'] = '⚠ NO modificar columnas A, B, C. Completar solo D (% Asist.), E (% TP), F (Nota Nº), G (Letras), H (Condición).'
    ws['A9'].font      = font_warning
    ws['A9'].fill      = PatternFill('solid', fgColor='FFF3CD')
    ws['A9'].alignment = al_left

    # ── HEADERS DOBLES (filas 10 y 11) ──
    # Fila 10: N° | DNI | Apellido y Nombre | % Asistencia | %TP-Aprob. | Calificación (merge F-G) | Promocionó/Regularizó/Libre
    ws.merge_cells('A10:A11')
    ws['A10'] = 'N°'
    ws['A10'].font = font_header; ws['A10'].fill = fill_header; ws['A10'].alignment = al_center

    ws.merge_cells('B10:B11')
    ws['B10'] = 'DNI'
    ws['B10'].font = font_header; ws['B10'].fill = fill_header; ws['B10'].alignment = al_center

    ws.merge_cells('C10:C11')
    ws['C10'] = 'Apellido y Nombre'
    ws['C10'].font = font_header; ws['C10'].fill = fill_header; ws['C10'].alignment = al_center

    ws.merge_cells('D10:D11')
    ws['D10'] = '% Asistencia'
    ws['D10'].font = font_header; ws['D10'].fill = fill_header; ws['D10'].alignment = al_center

    ws.merge_cells('E10:E11')
    ws['E10'] = '%TP-\nAprob.'
    ws['E10'].font = font_header; ws['E10'].fill = fill_header; ws['E10'].alignment = al_center

    ws.merge_cells('F10:G10')
    ws['F10'] = 'Calificación'
    ws['F10'].font = font_header; ws['F10'].fill = fill_header; ws['F10'].alignment = al_center
    ws['F11'] = 'N°'
    ws['F11'].font = font_header; ws['F11'].fill = fill_header; ws['F11'].alignment = al_center
    ws['G11'] = 'Letras'
    ws['G11'].font = font_header; ws['G11'].fill = fill_header; ws['G11'].alignment = al_center

    ws.merge_cells('H10:H11')
    ws['H10'] = 'Promocionó/\nRegularizó/Libre'
    ws['H10'].font = font_header; ws['H10'].fill = fill_header; ws['H10'].alignment = al_center

    for col in ['A','B','C','D','E','F','G','H']:
        for r in [10, 11]:
            ws[f'{col}{r}'].border = borde

    ws.row_dimensions[10].height = 30
    ws.row_dimensions[11].height = 20

    # ── DATOS DE ALUMNOS (desde fila 12) ──
    for idx, (apellido, nombre_al, dni, insc_id) in enumerate(alumnos, start=1):
        row  = idx + 11
        fill = fill_gris if idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')

        ws[f'A{row}'] = idx
        ws[f'A{row}'].font = font_normal; ws[f'A{row}'].fill = fill
        ws[f'A{row}'].alignment = al_center; ws[f'A{row}'].border = borde

        ws[f'B{row}'] = formatear_dni(dni)
        ws[f'B{row}'].font = font_normal; ws[f'B{row}'].fill = fill
        ws[f'B{row}'].alignment = al_center; ws[f'B{row}'].border = borde

        ws[f'C{row}'] = f"{apellido}, {nombre_al}"
        ws[f'C{row}'].font = font_normal; ws[f'C{row}'].fill = fill
        ws[f'C{row}'].alignment = al_left; ws[f'C{row}'].border = borde

        for col in ['D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{row}'].fill      = fill_amarillo
            ws[f'{col}{row}'].alignment = al_center
            ws[f'{col}{row}'].font      = font_normal
            ws[f'{col}{row}'].border    = borde

    # Lista desplegable columna H (Condición) para todas las filas de alumnos
    from openpyxl.worksheet.datavalidation import DataValidation
    if alumnos:
        fila_inicio = 12
        fila_fin = 11 + len(alumnos)
        dv = DataValidation(
            type='list',
            formula1='"Promocionó,Regularizó,Libre"',
            allow_blank=True,
            showDropDown=False
        )
        dv.sqref = f'H{fila_inicio}:H{fila_fin}'
        ws.add_data_validation(dv)

    # ── TOTALES (5 filas de margen) ──
    total_row = len(alumnos) + 12 + 5

    ws.merge_cells(f'A{total_row}:E{total_row}')
    ws[f'A{total_row}'] = 'Situacion'
    ws[f'A{total_row}'].font = font_bold; ws[f'A{total_row}'].alignment = al_right

    ws[f'F{total_row}'] = 'N°'
    ws[f'F{total_row}'].font = font_bold; ws[f'F{total_row}'].alignment = al_center
    ws[f'F{total_row}'].border = borde

    ws[f'G{total_row}'] = 'En letras'
    ws[f'G{total_row}'].font = font_bold; ws[f'G{total_row}'].alignment = al_center
    ws.merge_cells(f'G{total_row}:H{total_row}')
    ws[f'G{total_row}'].border = borde

    totales_labels = [
        ('Total de Alumnos Inscriptos:', len(alumnos)),
        ('Total de Alumnos Promocionados:', ''),
        ('Total de Alumnos Regularizados:', ''),
        ('Total de Alumnos Libres:', ''),
    ]

    for i, (label, val) in enumerate(totales_labels):
        r = total_row + 1 + i
        ws.merge_cells(f'A{r}:E{r}')
        ws[f'A{r}'] = label
        ws[f'A{r}'].font = font_normal; ws[f'A{r}'].alignment = al_right
        ws[f'A{r}'].border = borde
        ws[f'F{r}'] = val if val != '' else ''
        ws[f'F{r}'].font = font_normal; ws[f'F{r}'].alignment = al_center
        ws[f'F{r}'].fill = fill_totales; ws[f'F{r}'].border = borde
        ws.merge_cells(f'G{r}:H{r}')
        ws[f'G{r}'].fill = fill_totales; ws[f'G{r}'].border = borde
        ws[f'G{r}'].alignment = al_center

    # ── FIRMA ──
    firma_row = total_row + 6
    ws.merge_cells(f'A{firma_row}:D{firma_row}')
    ws[f'A{firma_row}'] = 'Fecha de Presentación: ___/___/______'
    ws[f'A{firma_row}'].font = font_normal

    ws.merge_cells(f'E{firma_row}:H{firma_row}')
    ws[f'E{firma_row}'] = 'Verificó: Firma: _________________ Aclaración: _________________'
    ws[f'E{firma_row}'].font = font_normal

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f"notas_{nombre_materia.replace(' ', '_')}_{anio}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre_archivo
    )


# ================================================================
# PLANILLA DE ASISTENCIAS
# ================================================================

@auth.route('/api/notas/descargar-asistencia/<int:mid>', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_notas_descargar_asistencia(mid):
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur  = conn.cursor()

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])

    cur.execute("""
        SELECT m.nombre, m.anio, m.regimen, c.nombre AS carrera
        FROM materias m
        JOIN carreras c ON c.id = m.carrera_id
        WHERE m.id = %s AND m.carrera_id = %s
    """, (mid, carrera_id))
    materia = cur.fetchone()
    if not materia:
        cur.close(); conn.close()
        return jsonify({'error': 'Materia no encontrada'}), 404

    nombre_materia, anio_materia, regimen, nombre_carrera = materia

    cur.execute("""
        SELECT p.nombre, p.apellido
        FROM materia_profesor mp
        JOIN profesores p ON p.id = mp.profesor_id
        WHERE mp.materia_id = %s AND mp.anio_lectivo = %s
    """, (mid, anio))
    profe_row = cur.fetchone()
    nombre_profe = f"{profe_row[1]}, {profe_row[0]}" if profe_row else '—'

    cur.execute("""
        SELECT a.apellido, a.nombre, a.dni
        FROM inscripciones i
        JOIN alumnos a ON a.id = i.alumno_id
        WHERE i.materia_id = %s AND i.anio_lectivo = %s AND a.carrera_id = %s
        ORDER BY a.apellido, a.nombre
    """, (mid, anio, carrera_id))
    alumnos = cur.fetchall()
    cur.close()
    conn.close()

    # ── Colores y estilos ────────────────────────────────────────
    VERDE        = colors.HexColor('#1a4731')
    AZUL_FECHA   = colors.HexColor('#D6EAF8')
    GRIS_PAR     = colors.HexColor('#F5F5F5')
    GRIS_BORDE   = colors.HexColor('#CCCCCC')

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1.2*cm, leftMargin=1.2*cm,
        topMargin=1.2*cm,   bottomMargin=1.2*cm
    )

    styles   = getSampleStyleSheet()
    st_inst  = ParagraphStyle('inst',  parent=styles['Normal'],
                               fontSize=11, fontName='Helvetica-Bold',
                               alignment=TA_CENTER, textColor=colors.white)
    st_carr  = ParagraphStyle('carr',  parent=styles['Normal'],
                               fontSize=9,  fontName='Helvetica',
                               alignment=TA_CENTER, textColor=colors.white)
    st_meta  = ParagraphStyle('meta',  parent=styles['Normal'],
                               fontSize=8,  fontName='Helvetica',
                               alignment=TA_LEFT)
    st_meta_b= ParagraphStyle('metab', parent=styles['Normal'],
                               fontSize=8,  fontName='Helvetica-Bold',
                               alignment=TA_LEFT)
    st_head  = ParagraphStyle('head',  parent=styles['Normal'],
                               fontSize=8,  fontName='Helvetica-Bold',
                               alignment=TA_CENTER, textColor=colors.white)
    st_fecha = ParagraphStyle('fecha', parent=styles['Normal'],
                               fontSize=7,  fontName='Helvetica-Oblique',
                               alignment=TA_LEFT,
                               textColor=colors.HexColor('#1a4731'))
    st_celda = ParagraphStyle('celda', parent=styles['Normal'],
                               fontSize=8,  fontName='Helvetica',
                               alignment=TA_LEFT)
    st_num   = ParagraphStyle('num',   parent=styles['Normal'],
                               fontSize=8,  fontName='Helvetica',
                               alignment=TA_CENTER)
    st_ley   = ParagraphStyle('ley',   parent=styles['Normal'],
                               fontSize=7,  fontName='Helvetica-Oblique',
                               textColor=colors.HexColor('#555555'),
                               alignment=TA_LEFT)

    elementos = []

    # ── Logo circular (igual que plan de estudios) ───────────────
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo_ies9.png')
    logo_img  = None
    if os.path.exists(logo_path):
        try:
            from PIL import Image as PILImage, ImageDraw
            pil_img = PILImage.open(logo_path).convert('RGBA')
            size    = pil_img.size
            lado    = min(size)
            left    = (size[0] - lado) // 2
            top_c   = (size[1] - lado) // 2
            pil_img = pil_img.crop((left, top_c, left + lado, top_c + lado))
            margen  = int(lado * 0.03)
            mascara = PILImage.new('L', (lado, lado), 0)
            draw    = ImageDraw.Draw(mascara)
            draw.ellipse((margen, margen, lado - margen, lado - margen), fill=255)
            resultado = PILImage.new('RGBA', (lado, lado), (255, 255, 255, 0))
            resultado.paste(pil_img, mask=mascara)
            logo_buf = BytesIO()
            resultado.save(logo_buf, format='PNG')
            logo_buf.seek(0)
            logo_img = RLImage(logo_buf, width=1.6*cm, height=1.6*cm)
        except Exception:
            pass

    # ── Encabezado verde con logo ────────────────────────────────
    encabezado_data = [[
        logo_img if logo_img else '',
        Paragraph(f'Instituto de Educación Superior N° 9 "Juana Azurduy"<br/>San Pedro de Jujuy — Jujuy', st_inst),
        ''
    ]]
    encabezado_tabla = Table(encabezado_data, colWidths=[2*cm, None, 2*cm])
    encabezado_tabla.setStyle(TableStyle([
        ('BACKGROUND',  (0, 0), (-1, 0), VERDE),
        ('VALIGN',      (0, 0), (-1, 0), 'MIDDLE'),
        ('ALIGN',       (0, 0), (0,  0), 'CENTER'),
        ('TOPPADDING',  (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING',(0,0), (-1, 0), 6),
    ]))
    elementos.append(encabezado_tabla)
    elementos.append(Spacer(1, 0.15*cm))

    # ── Carrera ──────────────────────────────────────────────────
    carr_data = [[Paragraph(nombre_carrera, st_carr)]]
    carr_tabla = Table(carr_data, colWidths=[None])
    carr_tabla.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,-1), VERDE),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 4),
    ]))
    elementos.append(carr_tabla)
    elementos.append(Spacer(1, 0.2*cm))

    # ── Datos de la materia ──────────────────────────────────────
    meta_data = [[
        Paragraph(f'<b>Espacio Curricular:</b> {nombre_materia}', st_meta),
        Paragraph(f'<b>Profesor/a:</b> {nombre_profe}',           st_meta),
        Paragraph(f'<b>Régimen:</b> {regimen or "—"}',            st_meta),
        Paragraph(f'<b>Año:</b> {anio_materia}°  —  <b>Año lectivo:</b> {anio}', st_meta),
    ]]
    meta_tabla = Table(meta_data, colWidths=[7*cm, 6*cm, 4*cm, None])
    meta_tabla.setStyle(TableStyle([
        ('FONTNAME',     (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE',     (0,0), (-1,-1), 8),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',   (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',(0,0), (-1,-1), 3),
        ('LINEBELOW',    (0,0), (-1,-1), 0.5, GRIS_BORDE),
    ]))
    elementos.append(meta_tabla)
    elementos.append(Spacer(1, 0.25*cm))

    # ── Tabla de asistencia ──────────────────────────────────────
    # Anchos: N°, DNI, Nombre, C1..C10
    ancho_pagina = landscape(A4)[0] - 2.4*cm
    ancho_c      = 1.5*cm
    ancho_nom    = ancho_pagina - 1*cm - 2.8*cm - 10*ancho_c
    col_widths   = [1*cm, 2.8*cm, ancho_nom] + [ancho_c]*10

    # Fila encabezados
    fila_head = [
        Paragraph('N°',                st_head),
        Paragraph('DNI',               st_head),
        Paragraph('Apellido y Nombre', st_head),
    ] + [Paragraph('Clase', st_head) for i in range(10)]

    # Fila de fechas (vacía, el profe completa)
    fila_fechas = [
        Paragraph('', st_fecha),
        Paragraph('', st_fecha),
        Paragraph('Fecha →', st_fecha),
    ] + [Paragraph('', st_fecha) for _ in range(10)]

    filas = [fila_head, fila_fechas]

    for idx, alumno in enumerate(alumnos):
        apellido, nombre, dni = alumno
        fila = [
            Paragraph(str(idx + 1),          st_num),
            Paragraph(formatear_dni(dni),     st_num),
            Paragraph(f'{apellido}, {nombre}',st_celda),
        ] + [Paragraph('', st_num) for _ in range(10)]
        filas.append(fila)

    tabla = Table(filas, colWidths=col_widths, repeatRows=2)

    estilo = TableStyle([
        # Encabezado verde
        ('BACKGROUND',   (0, 0), (-1, 0), VERDE),
        ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
        ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, 0), 8),
        ('ALIGN',        (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN',       (0, 0), (-1,-1), 'MIDDLE'),
        # Fila fechas azul
        ('BACKGROUND',   (0, 1), (-1, 1), AZUL_FECHA),
        ('FONTNAME',     (0, 1), (-1, 1), 'Helvetica-Oblique'),
        ('FONTSIZE',     (0, 1), (-1, 1), 7),
        # Grid
        ('GRID',         (0, 0), (-1,-1), 0.5, GRIS_BORDE),
        ('TOPPADDING',   (0, 0), (-1,-1), 4),
        ('BOTTOMPADDING',(0, 0), (-1,-1), 4),
        ('LEFTPADDING',  (0, 0), (-1,-1), 4),
        ('RIGHTPADDING', (0, 0), (-1,-1), 4),
        # Altura fila fechas y alumnos
        ('ROWHEIGHT',    (0, 1), (-1, 1), 14),
    ])

    # Filas alternadas para alumnos
    for i in range(len(alumnos)):
        if i % 2 == 0:
            estilo.add('BACKGROUND', (0, i + 2), (-1, i + 2), GRIS_PAR)

    tabla.setStyle(estilo)
    elementos.append(tabla)

    # ── Leyenda ──────────────────────────────────────────────────
    elementos.append(Spacer(1, 0.2*cm))
    elementos.append(Paragraph(
        'P = Presente &nbsp;&nbsp;|&nbsp;&nbsp; A = Ausente &nbsp;&nbsp;|&nbsp;&nbsp; T = Tardanza',
        st_ley
    ))

    doc.build(elementos)
    buf.seek(0)

    nombre_archivo = f"asistencia_{nombre_materia.replace(' ', '_')}_{anio}.pdf"
    return send_file(
        buf,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=nombre_archivo
    )


@auth.route('/api/notas/importar/<int:mid>', methods=['POST'])
@login_requerido(['coordinador', 'preceptora'])
def api_notas_importar(mid):
    carrera_id = session.get('carrera_id')

    if 'archivo' not in request.files:
        return jsonify({'error': 'No se recibió ningún archivo'}), 400

    archivo = request.files['archivo']
    if not archivo.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'El archivo debe ser .xlsx o .xls'}), 400

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])

    cur.execute("""SELECT id, nombre, regimen_aprobacion
                   FROM materias WHERE id = %s AND carrera_id = %s""", (mid, carrera_id))
    materia = cur.fetchone()
    if not materia:
        cur.close(); conn.close()
        return jsonify({'error': 'Materia no encontrada'}), 404

    cur.execute("""
        SELECT a.dni, i.id FROM inscripciones i
        JOIN alumnos a ON a.id = i.alumno_id
        WHERE i.materia_id = %s AND i.anio_lectivo = %s AND a.carrera_id = %s
    """, (mid, anio, carrera_id))
    dni_map = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("""
        SELECT a.dni, a.apellido || ', ' || a.nombre
        FROM alumnos a WHERE a.carrera_id = %s AND a.activo = TRUE
    """, (carrera_id,))
    todos_dni = {r[0]: r[1] for r in cur.fetchall()}

    cur.close()
    conn.close()

    try:
        wb = openpyxl.load_workbook(BytesIO(archivo.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'No se pudo leer el archivo: {str(e)}'}), 400

    # Detectar fila de headers
    header_row = None
    col_map = {}
    NOMBRES_COLUMNAS = {
        'dni':                   ['dni', 'documento', 'doc'],
        'porcentaje_asistencia': ['% asistencia', '%asistencia', 'asistencia', '% asist'],
        'porcentaje_tp':         ['% tp aprobado', '% tp', '%tp', 'tp aprobado', '%tp-', 'tp-'],
        'nota_cursada':          ['nota (1-10)', 'nota', 'calificacion', 'calificación', 'calificacion'],
        'condicion':             ['condicion', 'condición', 'situacion', 'situación',
                                  'promocionó/regularizó/libre', 'promociono/regularizo/libre',
                                  'promocionó/', 'regularizó/libre'],
    }

    for row in ws.iter_rows(min_row=1, max_row=12):
        fila_vals = [str(c.value).strip().lower() if c.value else '' for c in row]
        matches = 0
        temp_map = {}
        for campo, posibles in NOMBRES_COLUMNAS.items():
            for idx_c, val in enumerate(fila_vals):
                if not val:
                    continue
                if campo == 'nota_cursada' and val in ['n°', 'nº', 'numero', 'número']:
                    continue
                if any(p == val or (len(p) > 2 and p in val) for p in posibles):
                    temp_map[campo] = idx_c
                    matches += 1
                    break
        if matches >= 1 and 'dni' in temp_map:
            header_row = row[0].row
            col_map = temp_map
            # Verificar si la fila siguiente es sub-header (ej: "N°", "Letras")
            next_row_vals = []
            for r in ws.iter_rows(min_row=header_row+1, max_row=header_row+1):
                next_row_vals = [str(c.value).strip().lower() if c.value else '' for c in r]
            sub_header_words = ['n°', 'nº', 'letras', 'en letras', 'numero', 'número']
            if any(v in sub_header_words for v in next_row_vals):
                # La sub-fila puede tener la nota (N°) y condición
                for campo, posibles in NOMBRES_COLUMNAS.items():
                    if campo not in col_map:
                        for idx_c, val in enumerate(next_row_vals):
                            if not val:
                                continue
                            if any(p == val or (len(p) > 2 and p in val) for p in posibles):
                                col_map[campo] = idx_c
                                break
                header_row += 1
            break

    if header_row is None:
        return jsonify({
            'error': 'No se encontraron las columnas esperadas. '
                     'Descargá la plantilla del sistema o verificá que el Excel tenga '
                     'columnas de DNI, Nota y Condición.'
        }), 400

    filas_ok = []
    filas_error = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None or str(v).strip() == '' for v in row):
            continue
        primera = str(row[0]).lower().strip() if row[0] else ''
        segunda = str(row[1]).lower().strip() if len(row) > 1 and row[1] else ''
        # Ignorar filas de totales, firma, situación y filas sin DNI numérico
        if any(p in primera for p in ['total', 'situacion', 'situación', 'fecha', 'verificó', 'verifico']):
            continue
        if any(p in segunda for p in ['total', 'situacion', 'situación', 'fecha', 'verificó', 'verifico']):
            continue
        # Ignorar filas donde la columna DNI claramente no es un DNI
        dni_tentativo = ''
        if 'dni' in col_map and row[col_map['dni']]:
            dni_tentativo = limpiar_dni(str(row[col_map['dni']]))
        if not dni_tentativo and not any(
            row[col_map[c]] for c in ['nota_cursada', 'porcentaje_asistencia']
            if c in col_map and len(row) > col_map[c] and row[col_map[c]]
        ):
            continue

        row = list(row)

        # Extraer DNI primero — si no tiene DNI válido (7-8 dígitos), ignorar la fila
        dni_raw    = str(row[col_map['dni']]).strip() if 'dni' in col_map and len(row) > col_map['dni'] and row[col_map['dni']] else ''
        dni_limpio = limpiar_dni(dni_raw)

        # DNI válido argentino: entre 7 y 8 dígitos numéricos
        if not dni_limpio.isdigit() or not (7 <= len(dni_limpio) <= 8):
            continue

        nota_raw     = row[col_map['nota_cursada']]          if 'nota_cursada'          in col_map else None
        asist_raw    = row[col_map['porcentaje_asistencia']] if 'porcentaje_asistencia' in col_map else None
        tp_raw       = row[col_map['porcentaje_tp']]         if 'porcentaje_tp'         in col_map else None
        condicion_raw= row[col_map['condicion']]             if 'condicion'             in col_map else None

        nota      = normalizar_nota(nota_raw)
        asistencia= normalizar_porcentaje(asist_raw)
        tp        = normalizar_porcentaje(tp_raw)
        condicion = normalizar_condicion(condicion_raw)

        if condicion is None and nota is not None:
            condicion = sugerir_condicion(nota, materia[2])

        # El plan de estudios manda: si la materia no admite promoción,
        # una condición 'promocionado' que venga en la planilla se baja a
        # 'regular' y el alumno queda obligado a rendir el final.
        if condicion == 'promocionado' and not admite_promocion(materia[2]):
            condicion = 'regular'

        # Es la planilla final del profesor (no carga manual progresiva): una fila
        # sin nota y sin condición especificada significa que el alumno no cursó/
        # no se presentó → se clasifica como libre, no queda sin clasificar.
        if condicion is None and nota is None:
            condicion = 'libre'

        if not dni_limpio or not dni_limpio.isdigit():
            filas_error.append({
                'dni_raw': dni_raw, 'nota': nota, 'asistencia': asistencia,
                'tp': tp, 'condicion': condicion,
                'errores': ['DNI inválido o vacío'], 'sugerencias': []
            })
            continue

        insc_id = dni_map.get(dni_limpio)
        if insc_id is None:
            sugerencias = []
            for dni_db, nombre_alumno in todos_dni.items():
                dif = sum(1 for a, b in zip(dni_limpio.zfill(8), dni_db.zfill(8)) if a != b)
                if dif <= 1 and abs(len(dni_limpio) - len(dni_db)) <= 1:
                    sugerencias.append({'dni': formatear_dni(dni_db), 'nombre': nombre_alumno})
            filas_error.append({
                'dni_raw': formatear_dni(dni_limpio), 'nota': nota,
                'asistencia': asistencia, 'tp': tp, 'condicion': condicion,
                'errores': [f'DNI {formatear_dni(dni_limpio)} no está inscripto en esta materia'],
                'sugerencias': sugerencias[:3]
            })
            continue

        filas_ok.append({
            'inscripcion_id': insc_id,
            'dni': formatear_dni(dni_limpio),
            'nota_cursada': nota,
            'porcentaje_asistencia': asistencia,
            'porcentaje_tp': tp,
            'condicion': condicion,
            'condicion_sugerida': sugerir_condicion(nota, materia[2]),
            'nota_letras': nota_en_letras(nota),
        })

    return jsonify({
        'ok': True,
        'filas_ok': filas_ok,
        'filas_error': filas_error,
        'resumen': {
            'total':   len(filas_ok) + len(filas_error),
            'ok':      len(filas_ok),
            'errores': len(filas_error),
        }
    })

# ================================================================
# API — USUARIOS / PRECEPTORAS (solo coordinador)
# ================================================================

@auth.route('/api/usuarios/preceptoras', methods=['GET'])
@login_requerido(['coordinador'])
def api_preceptoras_listar():
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, usuario, nombre, apellido, dni, email, celular,
               activo, debe_cambiar_password
        FROM usuarios
        WHERE rol = 'preceptora' AND carrera_id = %s
        ORDER BY apellido, nombre
    """, (carrera_id,))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify([{
        'id': r[0], 'usuario': r[1], 'nombre': r[2], 'apellido': r[3],
        'dni': formatear_dni(r[4]) if r[4] else '',
        'dni_raw': r[4] or '',
        'email': r[5] or '', 'celular': r[6] or '',
        'activo': r[7], 'debe_cambiar_password': r[8]
    } for r in rows])


@auth.route('/api/usuarios/preceptoras', methods=['POST'])
@login_requerido(['coordinador'])
def api_preceptoras_crear():
    carrera_id = session.get('carrera_id')
    data = request.get_json()
    nombre   = data.get('nombre', '').strip()
    apellido = data.get('apellido', '').strip()
    dni      = limpiar_dni(data.get('dni', ''))
    email    = data.get('email', '').strip() or None
    celular  = data.get('celular', '').strip() or None

    if not nombre or not apellido or not dni:
        return jsonify({'error': 'Nombre, apellido y DNI son obligatorios'}), 400
    if not dni.isdigit() or len(dni) < 7:
        return jsonify({'error': 'DNI inválido'}), 400

    # usuario = DNI, contraseña inicial = DNI
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO usuarios
                (usuario, password_hash, rol, nombre, apellido, dni,
                 email, celular, carrera_id, debe_cambiar_password)
            VALUES (%s, %s, 'preceptora', %s, %s, %s, %s, %s, %s, TRUE)
            RETURNING id
        """, (dni, generate_password_hash(dni), nombre, apellido, dni, email, celular, carrera_id))
        nuevo_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({'ok': True, 'id': nuevo_id})
    except Exception as e:
        conn.rollback()
        if 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
            return jsonify({'error': 'Ya existe un usuario con ese DNI'}), 409
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@auth.route('/api/usuarios/preceptoras/<int:uid>', methods=['PUT'])
@login_requerido(['coordinador'])
def api_preceptoras_editar(uid):
    carrera_id = session.get('carrera_id')
    data = request.get_json()
    nombre   = data.get('nombre', '').strip()
    apellido = data.get('apellido', '').strip()
    email    = data.get('email', '').strip() or None
    celular  = data.get('celular', '').strip() or None

    if not nombre or not apellido:
        return jsonify({'error': 'Nombre y apellido son obligatorios'}), 400

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE usuarios
        SET nombre = %s, apellido = %s, email = %s, celular = %s
        WHERE id = %s AND rol = 'preceptora' AND carrera_id = %s
    """, (nombre, apellido, email, celular, uid, carrera_id))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})


@auth.route('/api/usuarios/preceptoras/<int:uid>/toggle', methods=['POST'])
@login_requerido(['coordinador'])
def api_preceptoras_toggle(uid):
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE usuarios SET activo = NOT activo
        WHERE id = %s AND rol = 'preceptora' AND carrera_id = %s
        RETURNING activo
    """, (uid, carrera_id))
    resultado = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    if not resultado:
        return jsonify({'error': 'Preceptora no encontrada'}), 404
    return jsonify({'ok': True, 'activo': resultado[0]})


@auth.route('/api/usuarios/preceptoras/<int:uid>/reset', methods=['POST'])
@login_requerido(['coordinador'])
def api_preceptoras_reset(uid):
    """Resetea la contraseña al DNI de la preceptora."""
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT dni FROM usuarios WHERE id = %s AND rol = 'preceptora' AND carrera_id = %s", (uid, carrera_id))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        return jsonify({'error': 'Preceptora no encontrada'}), 404
    cur.execute("""
        UPDATE usuarios
        SET password_hash = %s, debe_cambiar_password = TRUE
        WHERE id = %s AND rol = 'preceptora' AND carrera_id = %s
    """, (generate_password_hash(row[0]), uid, carrera_id))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})


# ================================================================
# API — PROFESORES (coordinador + preceptora)
# ================================================================

@auth.route('/api/profesores', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_profesores_listar():
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])

    cur.execute("""
        SELECT p.id, p.nombre, p.apellido, p.dni, p.email, p.celular, p.titulo, p.activo,
               m.id, m.nombre, m.anio
        FROM profesores p
        LEFT JOIN materia_profesor mp ON mp.profesor_id = p.id AND mp.anio_lectivo = %s
        LEFT JOIN materias m ON m.id = mp.materia_id AND m.carrera_id = %s
        ORDER BY p.apellido, p.nombre, m.anio, m.nombre
    """, (anio, carrera_id))

    rows = cur.fetchall()
    cur.close()
    conn.close()

    profesores = {}
    for r in rows:
        pid = r[0]
        if pid not in profesores:
            profesores[pid] = {
                'id': r[0], 'nombre': r[1], 'apellido': r[2],
                'dni': formatear_dni(r[3]) if r[3] else '',
                'dni_raw': r[3] or '',
                'email': r[4] or '', 'celular': r[5] or '',
                'titulo': r[6] or '', 'activo': r[7],
                'materias': []
            }
        if r[8]:  # tiene materia asignada en esta fila
            profesores[pid]['materias'].append({
                'id': r[8], 'nombre': r[9], 'anio': r[10]
            })

    return jsonify(list(profesores.values()))


@auth.route('/api/profesores', methods=['POST'])
@login_requerido(['coordinador'])
def api_profesores_crear():
    data = request.get_json()
    nombre   = data.get('nombre', '').strip()
    apellido = data.get('apellido', '').strip()
    dni      = limpiar_dni(data.get('dni', ''))
    email    = data.get('email', '').strip() or None
    celular  = data.get('celular', '').strip() or None
    titulo   = data.get('titulo', '').strip() or None

    if not nombre or not apellido:
        return jsonify({'error': 'Nombre y apellido son obligatorios'}), 400

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO profesores (nombre, apellido, dni, email, celular, titulo)
            VALUES (%s, %s, %s, %s, %s, %s) RETURNING id
        """, (nombre, apellido, dni or None, email, celular, titulo))
        nuevo_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({'ok': True, 'id': nuevo_id})
    except Exception as e:
        conn.rollback()
        if 'unique' in str(e).lower():
            return jsonify({'error': 'Ya existe un profesor con ese DNI'}), 409
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@auth.route('/api/profesores/<int:pid>', methods=['PUT'])
@login_requerido(['coordinador'])
def api_profesores_editar(pid):
    data = request.get_json()
    nombre   = data.get('nombre', '').strip()
    apellido = data.get('apellido', '').strip()
    dni      = limpiar_dni(data.get('dni', ''))
    email    = data.get('email', '').strip() or None
    celular  = data.get('celular', '').strip() or None
    titulo   = data.get('titulo', '').strip() or None

    if not nombre or not apellido:
        return jsonify({'error': 'Nombre y apellido son obligatorios'}), 400

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE profesores SET nombre=%s, apellido=%s, dni=%s,
                email=%s, celular=%s, titulo=%s
            WHERE id=%s
        """, (nombre, apellido, dni or None, email, celular, titulo, pid))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        if 'unique' in str(e).lower():
            return jsonify({'error': 'Ya existe un profesor con ese DNI'}), 409
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@auth.route('/api/profesores/<int:pid>', methods=['DELETE'])
@login_requerido(['coordinador'])
def api_profesores_eliminar(pid):
    conn = get_db()
    cur = conn.cursor()
    # Desvincular de materias primero
    cur.execute("DELETE FROM materia_profesor WHERE profesor_id = %s", (pid,))
    cur.execute("DELETE FROM profesores WHERE id = %s", (pid,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})


@auth.route('/api/profesores/<int:pid>/asignar', methods=['POST'])
@login_requerido(['coordinador'])
def api_profesores_asignar(pid):
    """Asigna una materia más al profesor en el año lectivo actual.
    Un profesor puede tener varias materias asignadas a la vez —
    esta ruta agrega, no reemplaza. Para quitar una asignación puntual
    se usa /api/profesores/<id>/desasignar/<materia_id>."""
    carrera_id = session.get('carrera_id')
    data = request.get_json()
    materia_id = data.get('materia_id')

    if not materia_id:
        return jsonify({'error': 'Falta materia_id'}), 400

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])

    try:
        # Verificar que la materia pertenece a la carrera
        cur.execute("SELECT id FROM materias WHERE id = %s AND carrera_id = %s", (materia_id, carrera_id))
        if not cur.fetchone():
            cur.close(); conn.close()
            return jsonify({'error': 'Materia no encontrada'}), 404

        cur.execute("""
            INSERT INTO materia_profesor (materia_id, profesor_id, anio_lectivo)
            VALUES (%s, %s, %s)
            ON CONFLICT (materia_id, profesor_id, anio_lectivo) DO NOTHING
        """, (materia_id, pid, anio))

        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@auth.route('/api/profesores/<int:pid>/desasignar/<int:materia_id>', methods=['DELETE'])
@login_requerido(['coordinador'])
def api_profesores_desasignar(pid, materia_id):
    """Quita UNA materia puntual de las asignadas al profesor, sin tocar las demás."""
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])

    cur.execute("""
        DELETE FROM materia_profesor mp
        USING materias m
        WHERE mp.materia_id = m.id
          AND mp.profesor_id = %s AND mp.materia_id = %s
          AND mp.anio_lectivo = %s AND m.carrera_id = %s
    """, (pid, materia_id, anio, carrera_id))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})


@auth.route('/api/profesores/materia/<int:mid>', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_profesor_de_materia(mid):
    """Devuelve el profesor asignado a una materia en el año lectivo actual."""
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])

    cur.execute("""
        SELECT p.id, p.nombre, p.apellido, p.titulo
        FROM materia_profesor mp
        JOIN profesores p ON p.id = mp.profesor_id
        JOIN materias m ON m.id = mp.materia_id
        WHERE mp.materia_id = %s AND mp.anio_lectivo = %s AND m.carrera_id = %s
    """, (mid, anio, carrera_id))

    row = cur.fetchone()
    cur.close()
    conn.close()

    if row:
        return jsonify({
            'id': row[0], 'nombre': row[1], 'apellido': row[2], 'titulo': row[3]
        })
    return jsonify(None)


# ================================================================
# API — REPORTES
# ================================================================

@auth.route('/api/reportes/mejores-promedios', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_reportes_mejores_promedios():
    carrera_id  = session.get('carrera_id')
    conn = get_db()
    cur  = conn.cursor()

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio_actual = int(cur.fetchone()[0])

    anio_lectivo = request.args.get('anio_lectivo', anio_actual, type=int)

    # Años de cursada disponibles en la carrera
    cur.execute("""
        SELECT DISTINCT anio FROM materias
        WHERE carrera_id = %s AND activa = TRUE
        ORDER BY anio
    """, (carrera_id,))
    anios_carrera = [r[0] for r in cur.fetchall()]

    # Años lectivos con datos para el selector
    cur.execute("""
        SELECT DISTINCT i.anio_lectivo
        FROM inscripciones i
        JOIN materias m ON m.id = i.materia_id
        WHERE m.carrera_id = %s
        ORDER BY i.anio_lectivo DESC
    """, (carrera_id,))
    anios_disponibles = [r[0] for r in cur.fetchall()]
    if anio_actual not in anios_disponibles:
        anios_disponibles.insert(0, anio_actual)

    resultado = []
    for anio_cursada in anios_carrera:
        cur.execute("""
            SELECT
                al.apellido || ', ' || al.nombre AS nombre_completo,
                al.dni,
                ROUND(AVG(cu.nota_cursada)::numeric, 2) AS promedio
            FROM alumnos al
            JOIN inscripciones i  ON i.alumno_id  = al.id
            JOIN materias m       ON m.id          = i.materia_id
            JOIN cursadas cu      ON cu.inscripcion_id = i.id
            WHERE m.carrera_id   = %s
              AND m.anio         = %s
              AND i.anio_lectivo = %s
              AND cu.nota_cursada IS NOT NULL
              AND cu.cerrada     = TRUE
            GROUP BY al.id, al.apellido, al.nombre, al.dni
            HAVING COUNT(cu.nota_cursada) > 0
            ORDER BY promedio DESC
            LIMIT 5
        """, (carrera_id, anio_cursada, anio_lectivo))
        alumnos = cur.fetchall()
        resultado.append({
            'anio_cursada': anio_cursada,
            'alumnos': [
                {'nombre': r[0], 'dni': r[1], 'promedio': float(r[2])}
                for r in alumnos
            ]
        })

    cur.close()
    conn.close()

    return jsonify({
        'anio_lectivo':       anio_lectivo,
        'anios_disponibles':  anios_disponibles,
        'bloques':            resultado
    })


@auth.route('/api/reportes/descargar-pdf', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_reportes_descargar_pdf():
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur  = conn.cursor()

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio_actual = int(cur.fetchone()[0])
    anio_lectivo = request.args.get('anio_lectivo', anio_actual, type=int)

    cur.execute("SELECT nombre FROM carreras WHERE id = %s", (carrera_id,))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        return jsonify({'error': 'Carrera no encontrada'}), 404
    nombre_carrera = row[0]

    # Años de cursada de la carrera
    cur.execute("""
        SELECT DISTINCT anio FROM materias
        WHERE carrera_id = %s AND activa = TRUE
        ORDER BY anio
    """, (carrera_id,))
    anios_carrera = [r[0] for r in cur.fetchall()]

    # Top 5 por año de cursada
    bloques = []
    for anio_cursada in anios_carrera:
        cur.execute("""
            SELECT
                al.apellido || ', ' || al.nombre AS nombre_completo,
                al.dni,
                ROUND(AVG(cu.nota_cursada)::numeric, 2) AS promedio
            FROM alumnos al
            JOIN inscripciones i  ON i.alumno_id  = al.id
            JOIN materias m       ON m.id          = i.materia_id
            JOIN cursadas cu      ON cu.inscripcion_id = i.id
            WHERE m.carrera_id   = %s
              AND m.anio         = %s
              AND i.anio_lectivo = %s
              AND cu.nota_cursada IS NOT NULL
            GROUP BY al.id, al.apellido, al.nombre, al.dni
            HAVING COUNT(cu.nota_cursada) > 0
            ORDER BY promedio DESC
            LIMIT 5
        """, (carrera_id, anio_cursada, anio_lectivo))
        alumnos = cur.fetchall()
        bloques.append({'anio_cursada': anio_cursada, 'alumnos': alumnos})

    cur.close()
    conn.close()

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm,   bottomMargin=1.5*cm
    )

    styles    = getSampleStyleSheet()
    st_titulo = ParagraphStyle('t', parent=styles['Normal'],
                                fontSize=13, fontName='Helvetica-Bold',
                                alignment=TA_CENTER, spaceAfter=4)
    st_sub    = ParagraphStyle('s', parent=styles['Normal'],
                                fontSize=10, fontName='Helvetica',
                                alignment=TA_CENTER, spaceAfter=2)
    st_head   = ParagraphStyle('h', parent=styles['Normal'],
                                fontSize=8, fontName='Helvetica-Bold',
                                alignment=TA_CENTER, textColor=colors.white)
    st_celda  = ParagraphStyle('c', parent=styles['Normal'],
                                fontSize=9, fontName='Helvetica',
                                alignment=TA_LEFT)
    st_num    = ParagraphStyle('n', parent=styles['Normal'],
                                fontSize=9, fontName='Helvetica',
                                alignment=TA_CENTER)
    st_anio   = ParagraphStyle('a', parent=styles['Normal'],
                                fontSize=10, fontName='Helvetica-Bold',
                                alignment=TA_LEFT, spaceAfter=4, spaceBefore=10)
    st_pos    = ParagraphStyle('pos', parent=styles['Normal'],
                                fontSize=9, fontName='Helvetica-Bold',
                                alignment=TA_CENTER)

    VERDE      = colors.HexColor('#1a4731')
    VERDE_CLARO= colors.HexColor('#e1f5ee')
    DORADO     = colors.HexColor('#FAC775')
    DORADO_T   = colors.HexColor('#633806')
    GRIS_PAR   = colors.HexColor('#F5F5F5')
    GRIS_BORDE = colors.HexColor('#CCCCCC')
    VERDE_T    = colors.HexColor('#0F6E56')

    elementos = []

    # Logo circular
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo_ies9.png')
    if os.path.exists(logo_path):
        try:
            from PIL import Image as PILImage, ImageDraw
            pil_img = PILImage.open(logo_path).convert('RGBA')
            size    = pil_img.size
            lado    = min(size)
            left    = (size[0] - lado) // 2
            top_c   = (size[1] - lado) // 2
            pil_img = pil_img.crop((left, top_c, left + lado, top_c + lado))
            margen  = int(lado * 0.03)
            mascara = PILImage.new('L', (lado, lado), 0)
            draw    = ImageDraw.Draw(mascara)
            draw.ellipse((margen, margen, lado - margen, lado - margen), fill=255)
            resultado_img = PILImage.new('RGBA', (lado, lado), (255, 255, 255, 0))
            resultado_img.paste(pil_img, mask=mascara)
            logo_buf = BytesIO()
            resultado_img.save(logo_buf, format='PNG')
            logo_buf.seek(0)
            logo_img = RLImage(logo_buf, width=1.8*cm, height=1.8*cm)
            logo_img.hAlign = 'CENTER'
            elementos.append(logo_img)
            elementos.append(Spacer(1, 0.2*cm))
        except Exception:
            pass

    # Encabezado
    elementos.append(Paragraph('Instituto de Educación Superior N° 9 "Juana Azurduy"', st_titulo))
    elementos.append(Paragraph('San Pedro de Jujuy — Jujuy', st_sub))
    elementos.append(Paragraph(nombre_carrera, st_sub))
    elementos.append(Paragraph(f'Mejores promedios de cursada — Año lectivo {anio_lectivo}', st_sub))
    elementos.append(Spacer(1, 0.5*cm))

    # Un bloque por año de cursada
    for bloque in bloques:
        anio_c  = bloque['anio_cursada']
        alumnos = bloque['alumnos']

        elementos.append(Paragraph(f'{anio_c}° Año', st_anio))

        enc = [
            Paragraph('#',               st_head),
            Paragraph('Alumno',          st_head),
            Paragraph('DNI',             st_head),
            Paragraph('Promedio',        st_head),
        ]
        tabla_filas = [enc]

        if not alumnos:
            tabla_filas.append([
                Paragraph('', st_num),
                Paragraph('Sin datos para este año', st_celda),
                Paragraph('', st_num),
                Paragraph('', st_num),
            ])
        else:
            for idx, al in enumerate(alumnos, 1):
                nombre_al, dni_al, promedio_al = al
                es_primero = idx == 1
                st_fila = ParagraphStyle(f'f{idx}', parent=st_celda,
                                         textColor=DORADO_T if es_primero else colors.black)
                st_num_fila = ParagraphStyle(f'fn{idx}', parent=st_num,
                                             textColor=DORADO_T if es_primero else colors.black,
                                             fontName='Helvetica-Bold' if es_primero else 'Helvetica')
                tabla_filas.append([
                    Paragraph(str(idx),        st_num_fila),
                    Paragraph(nombre_al,        st_fila),
                    Paragraph(str(dni_al),      st_num_fila),
                    Paragraph(str(promedio_al), st_num_fila),
                ])

        col_widths = [1.2*cm, 12*cm, 4*cm, 3*cm]
        tabla = Table(tabla_filas, colWidths=col_widths)
        estilo = TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), VERDE),
            ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
            ('ALIGN',         (0, 0), (-1,-1), 'CENTER'),
            ('ALIGN',         (1, 1), (1, -1), 'LEFT'),
            ('VALIGN',        (0, 0), (-1,-1), 'MIDDLE'),
            ('GRID',          (0, 0), (-1,-1), 0.5, GRIS_BORDE),
            ('TOPPADDING',    (0, 0), (-1,-1), 5),
            ('BOTTOMPADDING', (0, 0), (-1,-1), 5),
            ('LEFTPADDING',   (0, 0), (-1,-1), 6),
            ('RIGHTPADDING',  (0, 0), (-1,-1), 6),
        ])
        # Fila 1 (primer lugar) fondo dorado suave
        if len(tabla_filas) > 1 and alumnos:
            estilo.add('BACKGROUND', (0, 1), (-1, 1), DORADO)
        # Filas pares en gris suave
        for i in range(2, len(tabla_filas)):
            if i % 2 == 0:
                estilo.add('BACKGROUND', (0, i), (-1, i), GRIS_PAR)
        tabla.setStyle(estilo)
        elementos.append(tabla)
        elementos.append(Spacer(1, 0.4*cm))

    doc.build(elementos)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'mejores_promedios_{anio_lectivo}.pdf'
    )


# ================================================================
# API — MESAS DE EXAMEN
# ================================================================

@auth.route('/api/mesas', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_mesas_listar():
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])
    cur.execute("""
        SELECT me.id, me.numero_acta, m.nombre, m.anio, me.tipo,
               me.fecha_mesa, me.turno, me.cerrada,
               COUNT(ime.id) AS inscriptos
        FROM mesas_examen me
        JOIN materias m ON m.id = me.materia_id
        LEFT JOIN inscripciones_mesa ime ON ime.mesa_id = me.id
        WHERE me.carrera_id = %s AND me.anio_lectivo = %s
        GROUP BY me.id, me.numero_acta, m.nombre, m.anio, me.tipo,
                 me.fecha_mesa, me.turno, me.cerrada
        ORDER BY me.fecha_mesa DESC, me.numero_acta DESC
    """, (carrera_id, anio))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return jsonify([{
        'id': r[0], 'numero_acta': r[1], 'materia': r[2], 'anio_materia': r[3],
        'tipo': r[4], 'fecha_mesa': str(r[5]), 'turno': r[6],
        'cerrada': r[7], 'inscriptos': r[8]
    } for r in rows])


@auth.route('/api/mesas', methods=['POST'])
@login_requerido(['coordinador', 'preceptora'])
def api_mesas_crear():
    carrera_id = session.get('carrera_id')
    data = request.get_json()
    materia_id = data.get('materia_id')
    tipo       = data.get('tipo')
    fecha      = data.get('fecha_mesa')
    turno      = data.get('turno', 'Tarde')

    if not materia_id or not tipo or not fecha:
        return jsonify({'error': 'Faltan datos obligatorios'}), 400
    if tipo not in ('regular', 'libre'):
        return jsonify({'error': 'Tipo inválido'}), 400

    # ── Validación de fecha (hora del servidor, sincronizada por NTP) ──
    try:
        fecha_dt = datetime.strptime(fecha, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return jsonify({'error': 'Fecha inválida'}), 400

    hoy = date.today()

    # Tope superior: no tiene sentido una mesa a más de un año vista
    if fecha_dt > hoy + timedelta(days=365):
        return jsonify({'error': 'La fecha de la mesa no puede ser mayor a un año desde hoy.'}), 400

    # Fecha pasada: se permite (carga retroactiva de actas) pero pide confirmación
    if fecha_dt < hoy and not data.get('confirmar_retroactiva'):
        dias = (hoy - fecha_dt).days
        return jsonify({
            'requiere_confirmacion': True,
            'mensaje': f'La fecha indicada ya pasó (hace {dias} día{"s" if dias != 1 else ""}). '
                       f'Solo corresponde si estás cargando un acta de una mesa ya tomada. '
                       f'¿Confirmás que es una carga retroactiva?'
        }), 409

    conn = get_db()
    cur  = conn.cursor()

    # ── El plan manda: las materias de régimen sólo "Promoción"
    #    (prácticas profesionalizantes, EDI) no van a mesa de examen ──
    cur.execute("""SELECT nombre, regimen_aprobacion FROM materias
                   WHERE id = %s AND carrera_id = %s""", (materia_id, carrera_id))
    _mat = cur.fetchone()
    if not _mat:
        cur.close(); conn.close()
        return jsonify({'error': 'Materia no encontrada'}), 404
    if not admite_examen_final(_mat[1]):
        cur.close(); conn.close()
        return jsonify({'error': f'"{_mat[0]}" se aprueba únicamente por promoción '
                                 f'según el plan de estudios, no se rinde en mesa de examen.'}), 400

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio = int(cur.fetchone()[0])

    # Número de acta correlativo por carrera
    cur.execute("""
        SELECT COALESCE(MAX(numero_acta), 0) + 1
        FROM mesas_examen WHERE carrera_id = %s
    """, (carrera_id,))
    numero_acta = cur.fetchone()[0]

    try:
        cur.execute("""
            INSERT INTO mesas_examen
                (carrera_id, materia_id, tipo, fecha_mesa, turno, anio_lectivo, numero_acta)
            VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id
        """, (carrera_id, materia_id, tipo, fecha, turno, anio, numero_acta))
        mesa_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({'ok': True, 'id': mesa_id, 'numero_acta': numero_acta})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close(); conn.close()


@auth.route('/api/mesas/<int:mid>', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_mesas_detalle(mid):
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur  = conn.cursor()

    cur.execute("""
        SELECT me.id, me.numero_acta, m.nombre, m.anio, me.tipo,
               me.fecha_mesa, me.turno, me.cerrada, me.materia_id,
               me.motivo_cierre
        FROM mesas_examen me
        JOIN materias m ON m.id = me.materia_id
        WHERE me.id = %s AND me.carrera_id = %s
    """, (mid, carrera_id))
    mesa = cur.fetchone()
    if not mesa:
        cur.close(); conn.close()
        return jsonify({'error': 'Mesa no encontrada'}), 404

    cur.execute("""
        SELECT ime.id, a.id, a.apellido, a.nombre, a.dni,
               cu.condicion, ime.resultado, ime.nota_escrita,
               ime.nota_oral, ime.nota_final
        FROM inscripciones_mesa ime
        JOIN alumnos a ON a.id = ime.alumno_id
        LEFT JOIN inscripciones i ON i.alumno_id = a.id AND i.materia_id = %s
        LEFT JOIN cursadas cu ON cu.inscripcion_id = i.id
        WHERE ime.mesa_id = %s
        ORDER BY a.apellido, a.nombre
    """, (mesa[8], mid))
    inscriptos = cur.fetchall()
    cur.close(); conn.close()

    return jsonify({
        'id': mesa[0], 'numero_acta': mesa[1], 'materia': mesa[2],
        'anio_materia': mesa[3], 'tipo': mesa[4], 'fecha_mesa': str(mesa[5]),
        'turno': mesa[6], 'cerrada': mesa[7], 'materia_id': mesa[8],
        'motivo_cierre': mesa[9],
        'fecha_ya_paso': mesa[5] < date.today(),
        'inscriptos': [{
            'inscripcion_mesa_id': r[0], 'alumno_id': r[1],
            'apellido': r[2], 'nombre': r[3], 'dni': formatear_dni(r[4]),
            'condicion_cursada': r[5], 'resultado': r[6],
            'nota_escrita': float(r[7]) if r[7] else None,
            'nota_oral': float(r[8]) if r[8] else None,
            'nota_final': float(r[9]) if r[9] else None,
        } for r in inscriptos]
    })


@auth.route('/api/mesas/<int:mid>/inscribir', methods=['POST'])
@login_requerido(['coordinador', 'preceptora'])
def api_mesas_inscribir(mid):
    carrera_id = session.get('carrera_id')
    data = request.get_json()
    alumno_id = data.get('alumno_id')
    if not alumno_id:
        return jsonify({'error': 'Falta alumno_id'}), 400

    conn = get_db()
    cur  = conn.cursor()

    # Verificar que la mesa pertenece a esta carrera
    cur.execute("""SELECT materia_id, tipo, cerrada, fecha_mesa
                   FROM mesas_examen WHERE id = %s AND carrera_id = %s""",
                (mid, carrera_id))
    mesa = cur.fetchone()
    if not mesa:
        cur.close(); conn.close()
        return jsonify({'error': 'Mesa no encontrada'}), 404
    if mesa[2]:
        cur.close(); conn.close()
        return jsonify({'error': 'La mesa ya está cerrada'}), 400
    if mesa[3] < date.today():
        cur.close(); conn.close()
        return jsonify({'error': 'No se puede inscribir alumnos a una mesa cuya fecha ya pasó '
                                 f'({mesa[3].strftime("%d/%m/%Y")}).'}), 400

    # ── Revalidación en el backend ──────────────────────────────
    # La pantalla ya filtra los alumnos habilitados, pero la regla se
    # vuelve a verificar acá: la validación no puede depender de que el
    # navegador se haya comportado bien.
    materia_id_mesa = mesa[0]
    cur.execute("""
        SELECT cu.condicion, cu.cargado_en,
               (cu.cargado_en + INTERVAL '2 years')::date
        FROM inscripciones i
        JOIN cursadas cu ON cu.inscripcion_id = i.id
        WHERE i.alumno_id = %s AND i.materia_id = %s AND cu.cerrada = TRUE
        ORDER BY cu.cargado_en DESC LIMIT 1
    """, (alumno_id, materia_id_mesa))
    _cur_row = cur.fetchone()
    if not _cur_row:
        cur.close(); conn.close()
        return jsonify({'error': 'El alumno no tiene una cursada cerrada en esta materia.'}), 400

    _cond, _cargado, _vence = _cur_row
    if _vence and mesa[3] > _vence:
        cur.close(); conn.close()
        return jsonify({'error': f'La condición del alumno venció el '
                                 f'{_vence.strftime("%d/%m/%Y")}.'}), 400

    cur.execute("""
        SELECT COUNT(*) FROM examenes
        WHERE alumno_id = %s AND materia_id = %s AND fecha_mesa >= %s
    """, (alumno_id, materia_id_mesa, _cargado))
    if cur.fetchone()[0] >= 3:
        cur.close(); conn.close()
        return jsonify({'error': 'El alumno agotó los 3 intentos permitidos.'}), 400

    cur.execute("""
        SELECT m.nombre, m.orden FROM correlatividades co
        JOIN materias m ON m.id = co.requiere_materia_id
        WHERE co.materia_id = %s AND co.tipo = 'aprobada'
          AND co.requiere_materia_id NOT IN (
              SELECT materia_id FROM (
                  SELECT i.materia_id FROM inscripciones i
                  JOIN cursadas cu ON cu.inscripcion_id = i.id
                  WHERE i.alumno_id = %s
                    AND cu.condicion IN ('promocionado', 'aprobado')
                  UNION
                  SELECT materia_id FROM examenes
                  WHERE alumno_id = %s AND resultado = 'aprobado'
              ) sub
          )
    """, (materia_id_mesa, alumno_id, alumno_id))
    _faltantes = cur.fetchall()
    if _faltantes:
        cur.close(); conn.close()
        return jsonify({'error': 'Le falta aprobar: ' +
                        ', '.join(f'({o}) {n}' for n, o in _faltantes)}), 400

    try:
        cur.execute("""
            INSERT INTO inscripciones_mesa (mesa_id, alumno_id)
            VALUES (%s, %s)
        """, (mid, alumno_id))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        if 'unique' in str(e).lower():
            return jsonify({'error': 'El alumno ya está inscripto en esta mesa'}), 409
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close(); conn.close()


@auth.route('/api/mesas/<int:mid>/desinscribir/<int:alumno_id>', methods=['DELETE'])
@login_requerido(['coordinador', 'preceptora'])
def api_mesas_desinscribir(mid, alumno_id):
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("SELECT cerrada FROM mesas_examen WHERE id = %s AND carrera_id = %s", (mid, carrera_id))
    mesa = cur.fetchone()
    if not mesa:
        cur.close(); conn.close()
        return jsonify({'error': 'Mesa no encontrada'}), 404
    if mesa[0]:
        cur.close(); conn.close()
        return jsonify({'error': 'La mesa ya está cerrada'}), 400
    cur.execute("DELETE FROM inscripciones_mesa WHERE mesa_id = %s AND alumno_id = %s", (mid, alumno_id))
    conn.commit()
    cur.close(); conn.close()
    return jsonify({'ok': True})


@auth.route('/api/mesas/<int:mid>/resultados', methods=['POST'])
@login_requerido(['coordinador', 'preceptora'])
def api_mesas_cargar_resultados(mid):
    carrera_id = session.get('carrera_id')
    data = request.get_json()
    resultados = data.get('resultados', [])

    conn = get_db()
    cur  = conn.cursor()

    cur.execute("""SELECT materia_id, cerrada, fecha_mesa, anio_lectivo
                   FROM mesas_examen WHERE id = %s AND carrera_id = %s""",
                (mid, carrera_id))
    mesa = cur.fetchone()
    if not mesa:
        cur.close(); conn.close()
        return jsonify({'error': 'Mesa no encontrada'}), 404

    materia_id  = mesa[0]
    fecha_mesa  = mesa[2]
    anio        = mesa[3]   # año lectivo DE LA MESA, no el actual

    # No se pueden cargar resultados de un examen que todavía no se tomó
    if fecha_mesa > date.today():
        cur.close(); conn.close()
        return jsonify({'error': 'No se pueden cargar resultados de una mesa futura '
                                 f'({fecha_mesa.strftime("%d/%m/%Y")}).'}), 400

    try:
        for r in resultados:
            alumno_id    = r.get('alumno_id')
            resultado    = r.get('resultado')  # aprobado / desaprobado / ausente
            nota_escrita = r.get('nota_escrita')
            nota_oral    = r.get('nota_oral')
            nota_final   = r.get('nota_final')
            libro        = r.get('libro')
            folio        = r.get('folio')

            # Actualizar inscripcion_mesa
            cur.execute("""
                UPDATE inscripciones_mesa
                SET resultado = %s, nota_escrita = %s, nota_oral = %s, nota_final = %s
                WHERE mesa_id = %s AND alumno_id = %s
            """, (resultado, nota_escrita, nota_oral, nota_final, mid, alumno_id))

            # Insertar/actualizar en tabla examenes
            cur.execute("""
                INSERT INTO examenes (alumno_id, materia_id, anio_lectivo, fecha_mesa, nota, resultado)
                SELECT %s, %s, %s, me.fecha_mesa, %s, %s
                FROM mesas_examen me WHERE me.id = %s
                ON CONFLICT (alumno_id, materia_id, anio_lectivo, fecha_mesa)
                DO UPDATE SET nota = EXCLUDED.nota, resultado = EXCLUDED.resultado
            """, (alumno_id, materia_id, anio, nota_final, resultado, mid))

            # ── FIX: actualizar condición en cursadas + libro/folio si aprobado ──
            if resultado == 'aprobado':
                # FIX: se actualiza la cursada MÁS RECIENTE del alumno en esa
                # materia, sin importar en qué ciclo lectivo la haya cursado.
                # Antes filtraba por el año actual, así que un alumno que
                # regularizó en un ciclo anterior nunca quedaba como aprobado.
                cur.execute("""
                    UPDATE cursadas
                    SET condicion = 'aprobado',
                        libro     = %s,
                        folio     = %s
                    WHERE id = (
                        SELECT cu.id
                        FROM cursadas cu
                        JOIN inscripciones i ON i.id = cu.inscripcion_id
                        WHERE i.alumno_id  = %s
                          AND i.materia_id = %s
                        ORDER BY i.anio_lectivo DESC, cu.cargado_en DESC
                        LIMIT 1
                    )
                """, (libro, folio, alumno_id, materia_id))

                # ── Confirmar promociones provisorias ──
                # Al aprobar este final, el alumno destraba las materias
                # que había promocionado adeudando esta correlativa.
                # Sin esto quedarían marcadas y se caerían igual cuando
                # venza el plazo, pese a haber cumplido.
                cur.execute("""
                    UPDATE cursadas cu
                    SET promocion_provisoria = FALSE
                    FROM inscripciones i, correlatividades co
                    WHERE cu.inscripcion_id = i.id
                      AND i.alumno_id = %s
                      AND co.materia_id = i.materia_id
                      AND co.tipo = 'aprobada'
                      AND co.requiere_materia_id = %s
                      AND cu.promocion_provisoria
                      AND NOT cu.cerrada
                      -- Solo si NO le queda ninguna otra correlativa
                      -- 'aprobada' pendiente para esa materia.
                      AND NOT EXISTS (
                          SELECT 1
                          FROM correlatividades co2
                          WHERE co2.materia_id = i.materia_id
                            AND co2.tipo = 'aprobada'
                            AND co2.requiere_materia_id <> %s
                            AND co2.requiere_materia_id NOT IN (
                                SELECT i2.materia_id
                                FROM inscripciones i2
                                JOIN cursadas cu2 ON cu2.inscripcion_id = i2.id
                                WHERE i2.alumno_id = %s
                                  AND cu2.condicion IN ('promocionado', 'aprobado')
                                  AND NOT cu2.promocion_provisoria
                                UNION
                                SELECT e2.materia_id FROM examenes e2
                                WHERE e2.alumno_id = %s AND e2.resultado = 'aprobado'
                            )
                      )
                """, (alumno_id, materia_id, materia_id, alumno_id, alumno_id))

        # Cerrar la mesa
        cur.execute("UPDATE mesas_examen SET cerrada = TRUE WHERE id = %s", (mid,))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close(); conn.close()


@auth.route('/api/mesas/<int:mid>/cerrar-desierta', methods=['POST'])
@login_requerido(['coordinador', 'preceptora'])
def api_mesas_cerrar_desierta(mid):
    """
    Cierra una mesa a la que no se inscribió ningún alumno.
    Deja constancia del motivo: institucionalmente el acta de una mesa
    desierta se labra igual, no se borra el registro.
    """
    carrera_id = session.get('carrera_id')
    data   = request.get_json() or {}
    motivo = (data.get('motivo') or '').strip()

    if len(motivo) < 10:
        return jsonify({'error': 'Indicá el motivo del cierre (mínimo 10 caracteres).'}), 400

    conn = get_db()
    cur  = conn.cursor()

    cur.execute("""SELECT cerrada, fecha_mesa FROM mesas_examen
                   WHERE id = %s AND carrera_id = %s""", (mid, carrera_id))
    mesa = cur.fetchone()
    if not mesa:
        cur.close(); conn.close()
        return jsonify({'error': 'Mesa no encontrada'}), 404
    if mesa[0]:
        cur.close(); conn.close()
        return jsonify({'error': 'La mesa ya está cerrada.'}), 400

    cur.execute("SELECT COUNT(*) FROM inscripciones_mesa WHERE mesa_id = %s", (mid,))
    if cur.fetchone()[0] > 0:
        cur.close(); conn.close()
        return jsonify({'error': 'La mesa tiene alumnos inscriptos. '
                                 'Cerrala cargando los resultados.'}), 400

    usuario_id = session.get('usuario_id')
    try:
        cur.execute("""
            UPDATE mesas_examen
            SET cerrada = TRUE,
                motivo_cierre = %s
            WHERE id = %s
        """, (f"MESA DESIERTA — {motivo}", mid))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close(); conn.close()


@auth.route('/api/mesas/<int:mid>/alumnos-disponibles', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_mesas_alumnos_disponibles(mid):
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur  = conn.cursor()

    cur.execute("""
        SELECT materia_id, tipo, fecha_mesa
        FROM mesas_examen
        WHERE id = %s AND carrera_id = %s
    """, (mid, carrera_id))
    mesa = cur.fetchone()
    if not mesa:
        cur.close(); conn.close()
        return jsonify({'error': 'Mesa no encontrada'}), 404

    materia_id, tipo, fecha_mesa = mesa
    condicion = 'regular' if tipo == 'regular' else 'libre'

    # ------------------------------------------------------------
    # 1) Candidatos: cursada MÁS RECIENTE de cada alumno en esa
    #    materia, con la condición que pide la mesa y ya cerrada.
    # ------------------------------------------------------------
    cur.execute("""
        WITH ultima_cursada AS (
            SELECT DISTINCT ON (i.alumno_id)
                   i.alumno_id,
                   cu.condicion,
                   cu.cargado_en,
                   (cu.cargado_en + INTERVAL '2 years')::date AS vence_el
            FROM inscripciones i
            JOIN cursadas cu ON cu.inscripcion_id = i.id
            WHERE i.materia_id = %s
              AND cu.condicion = %s
              AND cu.cerrada   = TRUE
            ORDER BY i.alumno_id, cu.cargado_en DESC
        )
        SELECT a.id, a.apellido, a.nombre, a.dni,
               uc.condicion, uc.vence_el,
               (SELECT COUNT(*)
                  FROM examenes e
                 WHERE e.alumno_id  = a.id
                   AND e.materia_id = %s
                   AND e.fecha_mesa >= uc.cargado_en::date) AS intentos
        FROM ultima_cursada uc
        JOIN alumnos a ON a.id = uc.alumno_id
        WHERE a.carrera_id = %s
          AND a.activo     = TRUE
          AND a.id NOT IN (
              SELECT alumno_id FROM inscripciones_mesa WHERE mesa_id = %s
          )
          AND a.id NOT IN (
              SELECT alumno_id FROM examenes
              WHERE materia_id = %s AND resultado = 'aprobado'
          )
        ORDER BY a.apellido, a.nombre
    """, (materia_id, condicion, materia_id, carrera_id, mid, materia_id))
    candidatos = cur.fetchall()

    # ------------------------------------------------------------
    # 2) Correlatividades de tipo 'aprobada' que exige esta materia
    # ------------------------------------------------------------
    cur.execute("""
        SELECT co.requiere_materia_id, m.nombre, m.orden
        FROM correlatividades co
        JOIN materias m ON m.id = co.requiere_materia_id
        WHERE co.materia_id = %s AND co.tipo = 'aprobada'
    """, (materia_id,))
    requeridas = cur.fetchall()

    # ------------------------------------------------------------
    # 3) Qué materias tiene aprobadas cada candidato
    # ------------------------------------------------------------
    aprobadas_por_alumno = {}
    if requeridas and candidatos:
        ids_alumnos = tuple(c[0] for c in candidatos)
        cur.execute("""
            SELECT alumno_id, materia_id FROM (
                SELECT i.alumno_id, i.materia_id
                FROM inscripciones i
                JOIN cursadas cu ON cu.inscripcion_id = i.id
                WHERE i.alumno_id IN %s
                  AND cu.condicion IN ('promocionado', 'aprobado')
                UNION
                SELECT alumno_id, materia_id
                FROM examenes
                WHERE alumno_id IN %s AND resultado = 'aprobado'
            ) sub
        """, (ids_alumnos, ids_alumnos))
        for alumno_id, mat_id in cur.fetchall():
            aprobadas_por_alumno.setdefault(alumno_id, set()).add(mat_id)

    cur.close(); conn.close()

    # ------------------------------------------------------------
    # 4) Aplicar las reglas
    # ------------------------------------------------------------
    habilitados = []
    bloqueados  = []

    for aid, apellido, nombre, dni, cond, vence_el, intentos in candidatos:
        motivos = []

        etiqueta = 'Regularidad' if tipo == 'regular' else 'Condición libre'

        if vence_el and fecha_mesa > vence_el:
            motivos.append(
                f"{etiqueta} vencida el {vence_el.strftime('%d/%m/%Y')}"
            )
        if intentos >= 3:
            motivos.append(f"Agotó los 3 intentos ({intentos} rendidos)")

        aprobadas = aprobadas_por_alumno.get(aid, set())
        faltantes = [f"({orden}) {nom}"
                     for req_id, nom, orden in requeridas
                     if req_id not in aprobadas]
        if faltantes:
            motivos.append("Le falta aprobar: " + ", ".join(faltantes))

        alumno = {
            'id': aid,
            'apellido': apellido,
            'nombre': nombre,
            'dni': formatear_dni(dni),
            'condicion': cond,
            'vence_el': vence_el.strftime('%d/%m/%Y') if vence_el else None,
            'intentos': intentos,
            'intentos_restantes': max(0, 3 - intentos)
        }

        if motivos:
            alumno['motivos'] = motivos
            bloqueados.append(alumno)
        else:
            habilitados.append(alumno)

    return jsonify({
        'habilitados': habilitados,
        'bloqueados': bloqueados
    })


@auth.route('/api/mesas/<int:mid>/acta-pdf', methods=['GET'])
@login_requerido(['coordinador', 'preceptora'])
def api_mesas_acta_pdf(mid):
    carrera_id = session.get('carrera_id')
    conn = get_db()
    cur  = conn.cursor()

    cur.execute("""
        SELECT me.numero_acta, m.nombre, m.anio, me.tipo,
               me.turno, c.nombre, me.cerrada
        FROM mesas_examen me
        JOIN materias m ON m.id = me.materia_id
        JOIN carreras c ON c.id = me.carrera_id
        WHERE me.id = %s AND me.carrera_id = %s
    """, (mid, carrera_id))
    mesa = cur.fetchone()
    if not mesa:
        cur.close(); conn.close()
        return jsonify({'error': 'Mesa no encontrada'}), 404

    numero_acta, nombre_materia, anio_materia, tipo, turno, nombre_carrera, cerrada = mesa

    cur.execute("""
        SELECT a.apellido, a.nombre, a.dni,
               ime.nota_escrita, ime.nota_oral, ime.nota_final, ime.resultado
        FROM inscripciones_mesa ime
        JOIN alumnos a ON a.id = ime.alumno_id
        WHERE ime.mesa_id = %s
        ORDER BY a.apellido, a.nombre
    """, (mid,))
    inscriptos = cur.fetchall()
    cur.close(); conn.close()



    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.pdfgen import canvas as rl_canvas

    PAGE_W, PAGE_H = A4
    LM = 1.5*cm;  RM = 1.5*cm;  TM = 1.2*cm;  BM = 1.2*cm
    W  = PAGE_W - LM - RM
    NEGRO = colors.black
    GRIS  = colors.HexColor('#DDDDDD')
    LINE_H = 13
    FILAS_POR_HOJA = 30

    styles = getSampleStyleSheet()

    def P(txt, size=8, bold=False, align=TA_LEFT):
        fn = 'Helvetica-Bold' if bold else 'Helvetica'
        return Paragraph(txt, ParagraphStyle('_', parent=styles['Normal'],
            fontSize=size, fontName=fn, alignment=align,
            textColor=NEGRO, leading=size+2))

    tipo_label = 'REGULARES' if tipo == 'regular' else 'LIBRES'
    anio_label = f'{anio_materia}\u00b0'
    CW = [1.0*cm, 1.2*cm, 2.5*cm, 7.5*cm,
          1.1*cm, 1.1*cm, 1.2*cm, 1.2*cm, 1.2*cm]
    H_BOLD = dict(bold=True, size=6, align=TA_CENTER)

    total_alumnos = len(inscriptos)
    aprob  = sum(1 for r in inscriptos if r[6] == 'aprobado')
    aplaz  = sum(1 for r in inscriptos if r[6] == 'desaprobado')
    ausen  = sum(1 for r in inscriptos if r[6] == 'ausente')
    pp     = '.......................'
    vt  = str(total_alumnos) if cerrada else pp
    va  = str(aprob)         if cerrada else pp
    vl  = str(aplaz)         if cerrada else pp
    vau = str(ausen)         if cerrada else pp

    total_hojas = max(1, -(-total_alumnos // FILAS_POR_HOJA))
    paginas = []
    for h in range(total_hojas):
        ini = h * FILAS_POR_HOJA
        fin = min(ini + FILAS_POR_HOJA, total_alumnos)
        paginas.append(inscriptos[ini:fin])

    def build_encabezado(hoja_num):
        acta_label = f'N\u00b0 ACTA: <b>{numero_acta}</b>'
        if total_hojas > 1:
            acta_label += f'  \u2014  Hoja {hoja_num} de {total_hojas}'
        top = Table([[
            P(acta_label, 8),
            Table([
                [P('LIBRO: ____________________', 8), P('FOLIO: ____________', 8)],
                [P('DIA ______  MES _____________________  A\u00d1O _______', 8), ''],
            ], colWidths=[5*cm, 4*cm], style=TableStyle([
                ('SPAN',         (0,1),(1,1)),
                ('TOPPADDING',   (0,0),(-1,-1), 1),
                ('BOTTOMPADDING',(0,0),(-1,-1), 1),
                ('LEFTPADDING',  (0,0),(-1,-1), 2),
            ])),
        ]], colWidths=[9*cm, 9*cm])
        top.setStyle(TableStyle([
            ('VALIGN',       (0,0),(-1,-1), 'TOP'),
            ('TOPPADDING',   (0,0),(-1,-1), 0),
            ('BOTTOMPADDING',(0,0),(-1,-1), 0),
            ('LEFTPADDING',  (0,0),(-1,-1), 0),
            ('RIGHTPADDING', (0,0),(-1,-1), 0),
        ]))
        return [top, Spacer(1, 3*mm),
                P('<b>ACTA VOLANTE DE EXAMEN</b>', 13, align=TA_CENTER),
                Spacer(1, 1*mm),
                P('<b>INSTITUTO DE EDUCACI\u00d3N SUPERIOR N\u00ba 9 "JUANA AZURDUY"</b>', 9, align=TA_CENTER),
                P(nombre_carrera.upper(), 8, align=TA_CENTER),
                Spacer(1, 3*mm)]

    def build_datos_examen():
        info = Table([
            [P('EXAMEN DE ALUMNOS:', 9, bold=True),
             P(f'<b>{tipo_label}</b>', 9),
             P('<b>TURNO:</b>', 9, bold=True),
             P(f'<b>{turno.upper()}</b>', 9)],
            [P('<b>ASIGNATURA:</b>', 9, bold=True),
             P(f'<b>{nombre_materia.upper()}</b>', 9),
             P('<b>A\u00d1O:</b>', 9, bold=True),
             P(f'<b>{anio_label}  Div 1\u00b0</b>', 9)],
        ], colWidths=[4*cm, 9*cm, 2*cm, 3*cm])
        info.setStyle(TableStyle([
            ('TOPPADDING',   (0,0),(-1,-1), 3),
            ('BOTTOMPADDING',(0,0),(-1,-1), 3),
            ('LEFTPADDING',  (0,0),(-1,-1), 3),
            ('BOX',          (0,0),(-1,-1), 0.5, NEGRO),
            ('LINEBELOW',    (0,0),(-1,0),  0.5, NEGRO),
            ('LINEBEFORE',   (2,0),(2,-1),  0.5, NEGRO),
        ]))
        return [info, Spacer(1, 2*mm)]

    def build_tabla(alumnos_hoja, offset):
        h0 = [P('N\u00ba<br/>ORDEN',   **H_BOLD), P('N\u00ba<br/>PERMISO', **H_BOLD),
              P('D.N.I',               **H_BOLD), P('APELLIDO Y NOMBRE/S', **H_BOLD),
              P('CALIFICACIONES',      **H_BOLD), '', '',
              P('N\u00ba DE LAS BOL',  **H_BOLD), '']
        h1 = ['', '', '', '',
              P('ESCRITA', **H_BOLD), P('ORAL', **H_BOLD), P('PROMED.', **H_BOLD),
              P('ESCRITA', **H_BOLD), P('ORAL', **H_BOLD)]
        filas = [h0, h1]
        for i, al in enumerate(alumnos_hoja):
            apellido, nombre, dni, nota_e, nota_o, nota_f, res = al
            ne = str(int(nota_e)) if nota_e is not None else ''
            no = str(int(nota_o)) if nota_o is not None else ''
            nf = str(int(nota_f)) if nota_f is not None else ''
            filas.append([
                P(str(offset+i+1), size=8, align=TA_CENTER),
                P('', size=8),
                P(formatear_dni(dni), size=8, align=TA_CENTER),
                P(f'{apellido}, {nombre}', size=8),
                P(ne, size=8, align=TA_CENTER),
                P(no, size=8, align=TA_CENTER),
                P(nf, size=8, align=TA_CENTER),
                P('', size=8), P('', size=8),
            ])
        for i in range(len(alumnos_hoja), FILAS_POR_HOJA):
            filas.append([P(str(offset+i+1), size=8, align=TA_CENTER)] + ['']*8)
        t = Table(filas, colWidths=CW)
        t.setStyle(TableStyle([
            ('GRID',          (0,0), (-1,-1), 0.5, NEGRO),
            ('BACKGROUND',    (0,0), (-1,1),  GRIS),
            ('SPAN',          (0,0), (0,1)),  ('SPAN', (1,0), (1,1)),
            ('SPAN',          (2,0), (2,1)),  ('SPAN', (3,0), (3,1)),
            ('SPAN',          (4,0), (6,0)),  ('SPAN', (7,0), (8,0)),
            ('ALIGN',         (0,0), (-1,1),  'CENTER'),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME',      (0,0), (-1,1),  'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,1),  6),
            ('FONTSIZE',      (0,2), (-1,-1), 8),
            ('TOPPADDING',    (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ('LEFTPADDING',   (0,0), (-1,-1), 2),
            ('RIGHTPADDING',  (0,0), (-1,-1), 2),
            ('ROWHEIGHT',     (0,2), (-1,-1), 12),
        ]))
        return t

    # Y fija donde empieza el footer (desde abajo)
    Y_FOOTER = BM + 4*LINE_H      # línea más alta del footer

    def draw_footer_y_secretario(c, es_ultima, y_tabla=None):
        x  = LM
        # Footer: siempre en todas las hojas
        c.setFont('Helvetica', 7)
        c.drawString(x, Y_FOOTER,
            'PRESIDENTE: \u2026...................................................'
            '   VOCAL: \u2026..........................................................')
        c.drawString(x, Y_FOOTER - LINE_H,
            'VOCAL:\u2026.......................................................')
        c.setFont('Helvetica', 8)
        c.drawString(x, Y_FOOTER - 2*LINE_H,
            'SAN PEDRO DE JUJUY , ....... DE \u2026.................................DE 20\u2026.........')
        x_der = LM + 12.5*cm
        x_val = x_der + 3.4*cm
        c.setFont('Helvetica', 7)
        for idx, (lbl, val) in enumerate([
                ('TOTAL DE ALUMNOS:', vt), ('APROBADOS :', va),
                ('APLAZADOS :', vl),       ('AUSENTES :', vau)]):
            c.drawString(x_der, Y_FOOTER - idx*LINE_H, lbl)
            c.drawString(x_val, Y_FOOTER - idx*LINE_H, val)
        # Frase secretario: SOLO en última hoja, justo debajo de la tabla
        if es_ultima and y_tabla is not None:
            c.setFont('Helvetica', 7)
            c.drawString(x, y_tabla - 2*mm,
                'A continuaci\u00f3n del ultimo alumno deber\u00e1 firmar el secretario')

    def draw_elem(c, elem, y):
        _, h = elem.wrap(W, 9999)
        elem.drawOn(c, LM, y - h)
        return h

    buf = BytesIO()
    c   = rl_canvas.Canvas(buf, pagesize=A4)

    for idx_hoja, alumnos_hoja in enumerate(paginas):
        if idx_hoja > 0:
            c.showPage()
        hoja_num  = idx_hoja + 1
        es_ultima = (hoja_num == total_hojas)
        offset    = idx_hoja * FILAS_POR_HOJA

        y = PAGE_H - TM

        for e in build_encabezado(hoja_num):
            y -= draw_elem(c, e, y)

        for e in build_datos_examen():
            y -= draw_elem(c, e, y)

        tabla = build_tabla(alumnos_hoja, offset)
        _, th = tabla.wrapOn(c, W, 9999)
        tabla.drawOn(c, LM, y - th)
        y -= th

        draw_footer_y_secretario(c, es_ultima, y)

    c.save()
    buf.seek(0)
    nombre_archivo = f'acta_{numero_acta}_{nombre_materia.replace(" ","_")}_{tipo_label}.pdf'
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=nombre_archivo)


# ================================================================
# IMPORTACIÓN DE ALUMNOS DESDE GOOGLE SHEETS
# ================================================================
# Flujo:
# 1. Coordinador pega la URL pública de la Google Sheet (publicada como CSV)
# 2. POST /api/alumnos/importar-sheets/previsualizar → descarga el CSV,
#    valida cada fila, y devuelve:
#      - nuevos: filas que no existen en la DB (por DNI)
#      - duplicados: filas con DNI ya existente + diff de campos
#      - errores: filas con datos inválidos (sin DNI, DNI no numérico, etc.)
# 3. El coordinador revisa el diff en la pantalla y decide qué hacer
# 4. POST /api/alumnos/importar-sheets/confirmar → aplica los cambios:
#      - inserta los nuevos
#      - actualiza los duplicados que el coordinador aprobó

CAMPOS_SHEET = {
    # Nombres posibles en el encabezado del CSV → campo interno
    'apellido':       ['apellido', 'apellidos'],
    'nombre':         ['nombre', 'nombres'],
    'dni':            ['dni', 'documento', 'nro documento', 'número de documento', 'numero de documento', 'nro de documento'],
    'tipo_documento': ['tipo de documento', 'tipo documento', 'tipo_documento', 'tipo doc'],
    'cuil':           ['cuil', 'cuit', 'cuil/cuit', 'nro cuil', 'número de cuil', 'numero de cuil'],
    'email':          ['email', 'correo', 'correo electrónico', 'correo electronico', 'mail'],
    'celular':        ['celular', 'teléfono', 'telefono', 'tel', 'cel'],
    'direccion':      ['dirección', 'direccion', 'domicilio', 'calle'],
    'localidad':      ['localidad', 'ciudad', 'municipio'],
    'provincia':      ['provincia', 'jurisdicción', 'jurisdiccion', 'estado'],
    'fecha_nacimiento': ['fecha de nacimiento', 'fecha nacimiento', 'nacimiento', 'fecha_nacimiento'],
    'contacto_emergencia_nombre':    ['contacto emergencia', 'contacto de emergencia', 'familiar', 'nombre contacto'],
    'contacto_emergencia_telefono':  ['teléfono emergencia', 'telefono emergencia', 'tel contacto', 'tel emergencia'],
}

def _mapear_encabezados(headers):
    """Mapea los encabezados del CSV a los campos internos del sistema."""
    mapa = {}
    for i, h in enumerate(headers):
        h_norm = h.strip().lower()
        for campo, variantes in CAMPOS_SHEET.items():
            if h_norm in variantes:
                mapa[campo] = i
                break
    return mapa

def _convertir_url_sheet(url):
    """
    Convierte una URL de Google Sheets a su versión de exportación CSV.
    Acepta URLs en formato:
      - https://docs.google.com/spreadsheets/d/ID/edit#gid=0
      - https://docs.google.com/spreadsheets/d/ID/pub?output=csv
    """
    url = url.strip()
    if 'docs.google.com/spreadsheets' not in url:
        return None
    # Extraer el ID del spreadsheet
    import re
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
    if not m:
        return None
    sheet_id = m.group(1)
    # Extraer gid si existe
    gid_match = re.search(r'gid=(\d+)', url)
    gid = gid_match.group(1) if gid_match else '0'
    return f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}'


@auth.route('/api/alumnos/importar-sheets/previsualizar', methods=['POST'])
@login_requerido(['coordinador'])
def api_importar_sheets_previsualizar():
    carrera_id = session.get('carrera_id')
    data = request.get_json()
    url_raw   = (data.get('url') or '').strip()
    csv_texto = data.get('csv_texto', '').strip()

    if not url_raw and not csv_texto:
        return jsonify({'error': 'Ingresá la URL de la Google Sheet o subí un archivo CSV'}), 400

    if csv_texto:
        # Modo CSV manual — el texto ya viene del cliente
        contenido = csv_texto
    else:
        # Modo URL — descargar desde Google Sheets
        url_csv = _convertir_url_sheet(url_raw)
        if not url_csv:
            return jsonify({'error': 'La URL no parece ser una Google Sheet válida. Asegurate de copiar el link completo.'}), 400
        try:
            req = urllib.request.Request(url_csv, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=15) as resp:
                contenido = resp.read().decode('utf-8-sig')
        except urllib.error.HTTPError as e:
            if e.code == 403:
                return jsonify({'error': 'No se pudo acceder a la Sheet. Asegurate de que esté publicada como "Cualquier persona con el link puede ver".'}), 400
            return jsonify({'error': f'Error al descargar la Sheet (HTTP {e.code})'}), 400
        except Exception as e:
            return jsonify({'error': f'No se pudo conectar: {str(e)}'}), 400

    # Parsear CSV
    reader = csv.reader(contenido.splitlines())
    filas = list(reader)
    if len(filas) < 2:
        return jsonify({'error': 'La Sheet está vacía o solo tiene encabezados'}), 400

    encabezados = filas[0]
    mapa = _mapear_encabezados(encabezados)

    if 'apellido' not in mapa or 'nombre' not in mapa or 'dni' not in mapa:
        return jsonify({
            'error': f'No se encontraron las columnas obligatorias (Apellido, Nombre, DNI). '
                     f'Columnas detectadas: {", ".join(encabezados)}'
        }), 400

    # Cargar alumnos existentes en la DB (clave: documento normalizado por tipo + número)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, apellido, nombre, dni, email, celular, direccion, localidad,
               fecha_nacimiento, contacto_emergencia_nombre, contacto_emergencia_telefono,
               tipo_documento, cuil
        FROM alumnos WHERE carrera_id = %s
    """, (carrera_id,))
    existentes = {limpiar_documento(str(r[3])): {
        'id': r[0], 'apellido': r[1], 'nombre': r[2], 'dni': r[3],
        'email': r[4], 'celular': r[5], 'direccion': r[6], 'localidad': r[7],
        'fecha_nacimiento': str(r[8]) if r[8] else None,
        'contacto_emergencia_nombre': r[9], 'contacto_emergencia_telefono': r[10],
        'tipo_documento': r[11] or 'DNI', 'cuil': r[12],
    } for r in cur.fetchall()}
    cur.close()
    conn.close()

    nuevos = []
    duplicados = []
    errores = []

    for i, fila in enumerate(filas[1:], start=2):
        if not any(c.strip() for c in fila):
            continue  # fila vacía

        def get(campo):
            idx = mapa.get(campo)
            if idx is None or idx >= len(fila):
                return None
            v = fila[idx].strip()
            return v if v else None

        apellido = get('apellido')
        nombre   = get('nombre')
        dni_raw  = get('dni')
        dni      = limpiar_documento(dni_raw) if dni_raw else None

        # Tipo de documento: si no viene en el CSV, asumir DNI
        tipo_doc_raw = get('tipo_documento')
        tipo_doc = (tipo_doc_raw or 'DNI').strip().upper()
        # Aceptar también nombres extendidos
        if tipo_doc in ('DNI ARGENTINO', 'DNI ARG', 'NACIONAL'):
            tipo_doc = 'DNI'
        elif tipo_doc in ('DNI EXTRANJERO', 'DNI EXT', 'EXTRANJERO'):
            tipo_doc = 'DNI_EXT'
        elif tipo_doc in ('PASAPORTE', 'PASS'):
            tipo_doc = 'PAS'
        elif tipo_doc in ('CEDULA', 'CÉDULA', 'CEDULA DE IDENTIDAD', 'CEDULA IDENTIDAD'):
            tipo_doc = 'CI'

        cuil_raw = get('cuil')
        cuil     = limpiar_cuil(cuil_raw) if cuil_raw else None

        if not apellido or not nombre or not dni:
            errores.append({'fila': i, 'motivo': 'Faltan datos obligatorios (Apellido, Nombre o Documento)',
                            'datos': {'apellido': apellido, 'nombre': nombre, 'dni': dni_raw}})
            continue

        # Validar documento según tipo
        err_doc = validar_documento(tipo_doc, dni)
        if err_doc:
            errores.append({'fila': i, 'motivo': f'{err_doc}: "{dni_raw}"',
                            'datos': {'apellido': apellido, 'nombre': nombre, 'dni': dni_raw}})
            continue

        # Validar CUIL (si vino)
        err_cuil = validar_cuil(cuil)
        if err_cuil:
            errores.append({'fila': i, 'motivo': f'{err_cuil}: "{cuil_raw}"',
                            'datos': {'apellido': apellido, 'nombre': nombre, 'dni': dni_raw, 'cuil': cuil_raw}})
            continue

        # Provincia (opcional, normalizar si vino)
        provincia_raw = get('provincia')
        provincia = provincia_raw.strip() if provincia_raw else None
        # Aceptar variantes comunes de Jujuy
        if provincia and provincia.lower() in ('jujuy', 'jujui', 'sa.s.s. de jujuy'):
            provincia = 'Jujuy'
        if provincia and provincia.lower() in ('caba', 'capital federal', 'capital', 'ciudad de buenos aires'):
            provincia = 'Ciudad Autónoma de Buenos Aires'
        err_prov = validar_provincia(provincia)
        if err_prov:
            errores.append({'fila': i, 'motivo': f'{err_prov}: "{provincia_raw}"',
                            'datos': {'apellido': apellido, 'nombre': nombre, 'dni': dni_raw, 'provincia': provincia_raw}})
            continue

        # Fecha de nacimiento (validar si vino)
        fecha_nac_raw = get('fecha_nacimiento')
        fecha_nac = fecha_nac_raw
        if fecha_nac_raw:
            # Aceptar formatos DD/MM/AAAA y DD-MM-AAAA además del ISO
            fnr = fecha_nac_raw.strip()
            m = re.match(r'^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$', fnr)
            if m:
                fecha_nac = f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
        err_fnac = validar_fecha_nacimiento(fecha_nac)
        if err_fnac:
            errores.append({'fila': i, 'motivo': f'{err_fnac}: "{fecha_nac_raw}"',
                            'datos': {'apellido': apellido, 'nombre': nombre, 'dni': dni_raw, 'fecha_nacimiento': fecha_nac_raw}})
            continue

        nuevo = {
            'apellido': apellido, 'nombre': nombre, 'dni': dni,
            'tipo_documento': tipo_doc, 'cuil': cuil,
            'email':    get('email'), 'celular': get('celular'),
            'direccion': get('direccion'), 'localidad': get('localidad'),
            'provincia': provincia,
            'fecha_nacimiento': fecha_nac,
            'contacto_emergencia_nombre':   get('contacto_emergencia_nombre'),
            'contacto_emergencia_telefono': get('contacto_emergencia_telefono'),
        }

        if dni in existentes:
            viejo = existentes[dni]
            diff = []
            CAMPOS_DIFF = ['apellido','nombre','email','celular','direccion','localidad','provincia',
                           'fecha_nacimiento','contacto_emergencia_nombre','contacto_emergencia_telefono',
                           'tipo_documento','cuil']
            for campo in CAMPOS_DIFF:
                v_viejo = viejo.get(campo)
                v_nuevo = nuevo.get(campo)
                if v_nuevo and str(v_nuevo) != str(v_viejo or ''):
                    diff.append({'campo': campo, 'antes': v_viejo, 'despues': v_nuevo})
            duplicados.append({
                'fila': i, 'id': viejo['id'],
                'apellido': viejo['apellido'], 'nombre': viejo['nombre'], 'dni': dni,
                'diff': diff, 'datos_nuevos': nuevo,
                'actualizar': len(diff) > 0  # por defecto: actualizar si hay cambios
            })
        else:
            nuevos.append({'fila': i, **nuevo})

    return jsonify({
        'ok': True,
        'nuevos': nuevos,
        'duplicados': duplicados,
        'errores': errores,
        'resumen': {
            'total_filas': len(filas) - 1,
            'nuevos': len(nuevos),
            'duplicados': len(duplicados),
            'errores': len(errores),
        }
    })


@auth.route('/api/alumnos/importar-sheets/confirmar', methods=['POST'])
@login_requerido(['coordinador'])
def api_importar_sheets_confirmar():
    carrera_id = session.get('carrera_id')
    data = request.get_json()

    # Chequeo de ventana — misma regla que carga manual
    estado_ventana = get_estado_inscripciones()
    if not estado_ventana['abierto']:
        return jsonify({
            'error': f'No se pueden importar alumnos. {estado_ventana["motivo"]}.',
            'ventana_cerrada': True
        }), 403

    nuevos     = data.get('nuevos', [])
    duplicados = data.get('duplicados', [])  # solo los que tienen actualizar=True

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    anio_lectivo = int(cur.fetchone()[0])

    insertados  = 0
    actualizados = 0
    errores_conf = []

    # Insertar nuevos EN LOTES (execute_values) — mucho más rápido que uno por uno
    if nuevos:
        valores = []
        for a in nuevos:
            valores.append((
                carrera_id, a['apellido'], a['nombre'], a['dni'],
                (a.get('tipo_documento') or 'DNI'),
                a.get('cuil'),
                a.get('email'), a.get('celular'), a.get('direccion'), a.get('localidad'),
                a.get('provincia'),
                a.get('fecha_nacimiento') or None,
                a.get('contacto_emergencia_nombre'), a.get('contacto_emergencia_telefono'),
                anio_lectivo
            ))
        try:
            # page_size=1000 inserta de a 1000 registros por operación
            execute_values(cur, """
                INSERT INTO alumnos (
                    carrera_id, apellido, nombre, dni, tipo_documento, cuil,
                    email, celular, direccion, localidad, provincia,
                    fecha_nacimiento, contacto_emergencia_nombre,
                    contacto_emergencia_telefono, anio_ingreso
                ) VALUES %s
            """, valores, page_size=1000)
            insertados = len(valores)
        except Exception as e:
            conn.rollback()
            # Si falla el lote, reintentar uno por uno para identificar los problemáticos
            insertados = 0
            for a in nuevos:
                try:
                    cur.execute("""
                        INSERT INTO alumnos (
                            carrera_id, apellido, nombre, dni, tipo_documento, cuil,
                            email, celular, direccion, localidad, provincia,
                            fecha_nacimiento, contacto_emergencia_nombre,
                            contacto_emergencia_telefono, anio_ingreso
                        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        carrera_id, a['apellido'], a['nombre'], a['dni'],
                        (a.get('tipo_documento') or 'DNI'), a.get('cuil'),
                        a.get('email'), a.get('celular'), a.get('direccion'), a.get('localidad'),
                        a.get('provincia'),
                        a.get('fecha_nacimiento') or None,
                        a.get('contacto_emergencia_nombre'), a.get('contacto_emergencia_telefono'),
                        anio_lectivo
                    ))
                    insertados += 1
                except Exception as e2:
                    conn.rollback()
                    errores_conf.append({'dni': a['dni'], 'motivo': str(e2)})
                    continue

    # Actualizar duplicados aprobados
    for a in duplicados:
        if not a.get('actualizar'):
            continue
        d = a.get('datos_nuevos', {})
        try:
            cur.execute("""
                UPDATE alumnos SET
                    apellido = COALESCE(%s, apellido),
                    nombre   = COALESCE(%s, nombre),
                    tipo_documento = COALESCE(%s, tipo_documento),
                    cuil     = COALESCE(%s, cuil),
                    email    = COALESCE(%s, email),
                    celular  = COALESCE(%s, celular),
                    direccion = COALESCE(%s, direccion),
                    localidad = COALESCE(%s, localidad),
                    provincia = COALESCE(%s, provincia),
                    fecha_nacimiento = COALESCE(%s, fecha_nacimiento),
                    contacto_emergencia_nombre   = COALESCE(%s, contacto_emergencia_nombre),
                    contacto_emergencia_telefono = COALESCE(%s, contacto_emergencia_telefono)
                WHERE id = %s AND carrera_id = %s
            """, (
                d.get('apellido'), d.get('nombre'),
                d.get('tipo_documento'), d.get('cuil'),
                d.get('email'), d.get('celular'),
                d.get('direccion'), d.get('localidad'),
                d.get('provincia'),
                d.get('fecha_nacimiento') or None,
                d.get('contacto_emergencia_nombre'),
                d.get('contacto_emergencia_telefono'),
                a['id'], carrera_id
            ))
            actualizados += 1
        except Exception as e:
            errores_conf.append({'dni': a.get('dni'), 'motivo': str(e)})

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({
        'ok': True,
        'insertados': insertados,
        'actualizados': actualizados,
        'errores': errores_conf,
    })


# ============================================================
# Estadísticas del panel de administrador
# ============================================================
@auth.route('/api/stats/admin', methods=['GET'])
@login_requerido(['admin'])
def api_stats_admin():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT valor FROM configuracion WHERE clave = 'anio_lectivo_actual'")
    row = cur.fetchone()
    anio = row[0] if row else '2026'
    cur.close()
    conn.close()
    return jsonify({'anio': anio})
